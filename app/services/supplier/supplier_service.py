"""Supplier management: CRUD, search, trade matching, blacklist, Excel I/O.

Single-user/global for now (organization_id stays NULL; auth is Phase 15).
Trade-category matching is done in Python — SQLAlchemy JSON .contains() does
not work on SQLite/aiosqlite.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.supplier import Supplier

# Editable supplier fields shared by create/update (keeps DRY).
_SETTABLE = (
    "name", "name_ar", "code", "emails", "trade_categories", "phone", "fax",
    "address", "website", "contact_name", "contact_email", "contact_phone",
    "region", "country", "rating", "preferred_language", "preferred_format",
    "notes",
)


class SupplierService:
    async def _next_code(self, db: AsyncSession) -> str:
        count = (await db.execute(select(func.count(Supplier.id)))).scalar() or 0
        return f"SUP-{count + 1:04d}"

    async def create(
        self,
        db: AsyncSession,
        *,
        name: str,
        emails: list[str],
        trade_categories: list[str],
        **fields,
    ) -> Supplier:
        code = fields.pop("code", None) or await self._next_code(db)
        supplier = Supplier(
            organization_id=None,
            name=name,
            code=code,
            emails=emails or [],
            trade_categories=trade_categories or [],
            **{k: v for k, v in fields.items() if k in _SETTABLE},
        )
        db.add(supplier)
        await db.commit()
        await db.refresh(supplier)
        return supplier

    async def get(self, db: AsyncSession, supplier_id: int) -> Supplier | None:
        return await db.get(Supplier, supplier_id)

    async def update(
        self, db: AsyncSession, supplier_id: int, **fields
    ) -> Supplier | None:
        supplier = await db.get(Supplier, supplier_id)
        if supplier is None:
            return None
        for key, value in fields.items():
            if value is None:
                continue
            if key in _SETTABLE or key == "is_active":
                setattr(supplier, key, value)
        await db.commit()
        await db.refresh(supplier)
        return supplier

    async def list_suppliers(
        self,
        db: AsyncSession,
        *,
        query: str | None = None,
        trade: str | None = None,
        region: str | None = None,
        is_active: bool | None = True,
        min_rating: float | None = None,
    ) -> list[Supplier]:
        stmt = select(Supplier)
        if query:
            term = f"%{query}%"
            stmt = stmt.where(
                or_(
                    Supplier.name.ilike(term),
                    Supplier.code.ilike(term),
                    Supplier.contact_name.ilike(term),
                )
            )
        if region:
            stmt = stmt.where(Supplier.region == region)
        if is_active is not None:
            stmt = stmt.where(Supplier.is_active == is_active)
        if min_rating is not None:
            stmt = stmt.where(Supplier.rating >= min_rating)
        stmt = stmt.order_by(Supplier.name)
        rows = list((await db.execute(stmt)).scalars().all())
        # Trade filter in Python (JSON .contains is not portable to SQLite).
        if trade:
            rows = [s for s in rows if trade in (s.trade_categories or [])]
        return rows

    async def suppliers_for_trade(
        self, db: AsyncSession, trade: str, limit: int = 50
    ) -> list[Supplier]:
        stmt = (
            select(Supplier)
            .where(Supplier.is_active.is_(True), Supplier.is_blacklisted.is_(False))
            .order_by(Supplier.rating.desc().nullslast(), Supplier.name)
        )
        rows = [
            s
            for s in (await db.execute(stmt)).scalars().all()
            if trade in (s.trade_categories or [])
        ]
        return rows[:limit]

    async def blacklist(
        self, db: AsyncSession, supplier_id: int, reason: str
    ) -> Supplier | None:
        supplier = await db.get(Supplier, supplier_id)
        if supplier is None:
            return None
        supplier.is_blacklisted = True
        supplier.is_active = False
        supplier.blacklist_reason = reason
        await db.commit()
        await db.refresh(supplier)
        return supplier

    _COLUMN_CANDIDATES = {
        "name": ("name", "supplier_name", "vendor", "supplier"),
        "email": ("email", "emails", "email_address", "e-mail", "e_mail"),
        "trade": ("trade", "trades", "trade_category", "category", "specialization"),
        "phone": ("phone", "telephone", "tel", "mobile"),
        "contact": ("contact", "contact_name", "contact_person", "person"),
        "region": ("region", "area", "location"),
        "country": ("country",),
        "address": ("address", "full_address"),
        "website": ("website", "web", "url"),
    }

    async def import_excel(
        self, db: AsyncSession, file_path: str, update_existing: bool = False
    ) -> dict:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - surface a clean message
            raise ValueError(f"Failed to read Excel file: {exc}") from exc
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            raise ValueError("Excel file is empty")

        normalized = [_norm_header(h) for h in header]
        col = {
            key: _find_col(normalized, candidates)
            for key, candidates in self._COLUMN_CANDIDATES.items()
        }
        if col["name"] is None:
            raise ValueError("Required column 'name' not found")

        imported = updated = skipped = 0
        errors: list[str] = []
        # Capture the supplier count ONCE up front. Calling _count() inside the
        # loop would trigger autoflush of pending db.add()s and double-count,
        # producing gapped codes. base + imported + 1 stays correct & gap-free.
        base_count = await self._count(db)

        for row_idx, row in enumerate(rows, start=2):
            try:
                name = _cell(row, col["name"])
                if not name:
                    skipped += 1
                    continue
                emails = _split_multi(_cell(row, col["email"]), at_only=True)
                trades = [
                    _norm_trade(t)
                    for t in _split_multi(_cell(row, col["trade"]))
                    if t
                ]
                existing = (
                    await db.execute(select(Supplier).where(Supplier.name == name))
                ).scalar_one_or_none()
                if existing is not None:
                    if not update_existing:
                        skipped += 1
                        continue
                    if emails:
                        existing.emails = emails
                    if trades:
                        existing.trade_categories = trades
                    for fld, idx in (("phone", col["phone"]), ("contact_name", col["contact"]),
                                     ("region", col["region"]), ("country", col["country"]),
                                     ("address", col["address"]), ("website", col["website"])):
                        val = _cell(row, idx)
                        if val:
                            setattr(existing, fld, val)
                    updated += 1
                    continue

                supplier = Supplier(
                    organization_id=None,
                    name=name,
                    code=f"SUP-{base_count + imported + 1:04d}",
                    emails=emails,
                    trade_categories=trades,
                    phone=_cell(row, col["phone"]),
                    contact_name=_cell(row, col["contact"]),
                    region=_cell(row, col["region"]),
                    country=_cell(row, col["country"]),
                    address=_cell(row, col["address"]),
                    website=_cell(row, col["website"]),
                )
                db.add(supplier)
                imported += 1
            except Exception as exc:  # noqa: BLE001 - per-row resilience
                errors.append(f"Row {row_idx}: {exc}")

        await db.commit()
        wb.close()
        return {
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:10],
            "total_errors": len(errors),
        }

    async def _count(self, db: AsyncSession) -> int:
        return (await db.execute(select(func.count(Supplier.id)))).scalar() or 0

    async def export_excel(self, db: AsyncSession, output_path: str) -> str:
        suppliers = (
            await db.execute(select(Supplier).order_by(Supplier.name))
        ).scalars().all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"
        headers = ["Code", "Name", "Email(s)", "Trades", "Contact", "Phone",
                   "Region", "Country", "Rating", "Active"]
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = fill
            cell.font = font
        for r, s in enumerate(suppliers, 2):
            ws.cell(row=r, column=1, value=s.code)
            ws.cell(row=r, column=2, value=s.name)
            ws.cell(row=r, column=3, value=", ".join(s.emails or []))
            ws.cell(row=r, column=4, value=", ".join(s.trade_categories or []))
            ws.cell(row=r, column=5, value=s.contact_name)
            ws.cell(row=r, column=6, value=s.phone)
            ws.cell(row=r, column=7, value=s.region)
            ws.cell(row=r, column=8, value=s.country)
            ws.cell(row=r, column=9, value=s.rating)
            ws.cell(row=r, column=10, value="yes" if s.is_active else "no")
        for col_letter, width in zip("ABCDEFGHIJ", (12, 30, 35, 30, 20, 15, 15, 15, 10, 8)):
            ws.column_dimensions[col_letter].width = width
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path


def _norm_header(value) -> str:
    return str(value or "").lower().strip().replace(" ", "_")


def _find_col(normalized: list[str], candidates: tuple[str, ...]) -> int | None:
    for cand in candidates:
        if cand in normalized:
            return normalized.index(cand)
    return None


def _cell(row, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    text = str(val).strip()
    return text or None


def _split_multi(value: str | None, *, at_only: bool = False) -> list[str]:
    if not value:
        return []
    parts = [p.strip() for p in value.replace(";", ",").split(",")]
    parts = [p for p in parts if p]
    if at_only:
        parts = [p for p in parts if "@" in p]
    return parts


def _norm_trade(value: str) -> str:
    return value.strip().lower().replace(" ", "_")
