"""Write BOQ unit rates back into the client's Excel template, formula-preserving.

The workbook is loaded with openpyxl's defaults (NOT read_only, NOT
data_only) so every formula cell is preserved verbatim as its "=..." string.
Only the detected rate column is written, addressed by each item's
client_row_index (1-based Excel row), so total/amount formulas keep working
(Excel recalculates them on open).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

# Specific header aliases that unambiguously identify a unit-rate column. These
# are tried FIRST so a generic "Price" column (which is often a total) can never
# shadow an explicit unit-rate column.
_SPECIFIC_RATE_ALIASES = (
    "rate", "unit rate", "unit price", "unitprice", "unit_rate",
    "u.rate", "u rate", "u.price", "rate (usd)", "unit rate (usd)",
)
# The generic alias, tried only as a second pass when no specific alias matched.
_GENERIC_RATE_ALIASES = ("price",)
_SHEET_HINTS = ("boq", "bill", "quantity", "pricing", "boqs")
_MAX_HEADER_SCAN = 20


def _norm(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _pick_sheet(wb):
    for name in wb.sheetnames:
        if any(h in name.lower() for h in _SHEET_HINTS):
            return wb[name]
    return wb[wb.sheetnames[0]]


def detect_rate_column(ws) -> int | None:
    """Return the 1-based column index of the rate header, or None.

    Two-pass: specific unit-rate aliases first; only if none are found is the
    generic "price" alias considered. A header that normalizes to a total/amount
    label is never matched (those are not rate columns).
    """
    for aliases in (_SPECIFIC_RATE_ALIASES, _GENERIC_RATE_ALIASES):
        for r in range(1, min(ws.max_row, _MAX_HEADER_SCAN) + 1):
            for c in range(1, ws.max_column + 1):
                if _norm(ws.cell(row=r, column=c).value) in aliases:
                    return c
    return None


def populate_template(
    template_path: str,
    output_path: str,
    row_rates: dict[int, float],
    rate_column: int | None = None,
) -> dict:
    """Write rates into the rate column at the given 1-based rows; keep formulas.

    Args:
        template_path: the client's original .xlsx.
        output_path: where to write the populated copy.
        row_rates: {client_row_index (1-based) -> unit_rate}.
        rate_column: optional explicit 1-based rate column (else auto-detected).

    Returns: {"written": int, "rate_column": int, "skipped_formula": int}.
    Raises ValueError if the rate column cannot be determined.
    """
    wb = load_workbook(template_path)  # defaults preserve formulas
    try:
        ws = _pick_sheet(wb)
        col = rate_column or detect_rate_column(ws)
        if col is None:
            raise ValueError(
                "Could not detect a rate column in the template; pass rate_column explicitly"
            )
        written = 0
        skipped_formula = 0
        for row_idx, rate in row_rates.items():
            if row_idx is None:
                continue
            cell = ws.cell(row=row_idx, column=col)
            # Never clobber an existing formula cell (e.g. a misdetected total).
            if isinstance(cell.value, str) and cell.value.startswith("="):
                skipped_formula += 1
                continue
            cell.value = rate
            written += 1
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return {"written": written, "rate_column": col, "skipped_formula": skipped_formula}
    finally:
        wb.close()
