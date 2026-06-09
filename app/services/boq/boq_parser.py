"""BOQ Excel parser (openpyxl): workbook -> structured, unit-normalized rows."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import load_workbook

from app.schemas.rules import RulesConfig
from app.services.boq.sheet_select import pick_sheet

# Header-cell aliases -> canonical column key.
_COLUMN_ALIASES: dict[str, list[str]] = {
    "line_number": ["item", "item no", "no", "no.", "s.no", "sno", "ref", "line"],
    "description": ["description", "desc", "item description", "particulars", "work item"],
    "unit": ["unit", "uom", "u/m", "u.o.m"],
    "quantity": ["quantity", "qty", "qnty", "q'ty", "quantities"],
    "section": ["section", "division", "category", "trade", "bill"],
}

_MAX_HEADER_SCAN = 20


@dataclass
class ParsedBoqRow:
    """One parsed BOQ line (a priced item, not a section header)."""

    line_number: str | None
    section: str | None
    description: str
    unit: str | None
    quantity: float | None
    client_row_index: int  # 1-based Excel row number


def _norm(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def _find_header(ws) -> tuple[int, dict[int, str]]:
    """Return (1-based header row index, {col_index: canonical_key}).

    Picks the first row (within the scan window) that maps both a description
    column and at least one of unit/quantity. Falls back to row 1.
    """
    for r in range(1, min(ws.max_row, _MAX_HEADER_SCAN) + 1):
        col_map: dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            cell = _norm(ws.cell(row=r, column=c).value)
            if not cell:
                continue
            for key, aliases in _COLUMN_ALIASES.items():
                if cell in aliases and key not in col_map.values():
                    col_map[c] = key
                    break
        if "description" in col_map.values() and (
            "quantity" in col_map.values() or "unit" in col_map.values()
        ):
            return r, col_map
    return 1, {}


def _standardize_unit(raw: object, rules: RulesConfig) -> str | None:
    if raw is None or str(raw).strip() == "":
        return None
    key = str(raw).strip().lower()
    return rules.measurement.unit_mappings.get(key, str(raw).strip())


def parse_boq_workbook(file_path: str, rules: RulesConfig) -> list[ParsedBoqRow]:
    """Parse a BOQ workbook into priced rows; section headers propagate down."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = pick_sheet(wb)
        header_row, col_map = _find_header(ws)
        if not col_map:
            return []
        col_by_key = {v: k for k, v in col_map.items()}
        desc_col = col_by_key.get("description")
        qty_col = col_by_key.get("quantity")
        unit_col = col_by_key.get("unit")
        line_col = col_by_key.get("line_number")

        rows: list[ParsedBoqRow] = []
        current_section: str | None = None

        for r in range(header_row + 1, ws.max_row + 1):
            desc = ws.cell(row=r, column=desc_col).value if desc_col else None
            if desc is None or str(desc).strip() == "":
                continue
            description = str(desc).strip()
            qty_val = ws.cell(row=r, column=qty_col).value if qty_col else None
            quantity = _coerce_float(qty_val)

            # Row with a description but no numeric quantity = section header.
            if quantity is None:
                current_section = description
                continue

            rows.append(
                ParsedBoqRow(
                    line_number=(
                        str(ws.cell(row=r, column=line_col).value).strip()
                        if line_col and ws.cell(row=r, column=line_col).value is not None
                        else None
                    ),
                    section=current_section,
                    description=description,
                    unit=_standardize_unit(
                        ws.cell(row=r, column=unit_col).value if unit_col else None,
                        rules,
                    ),
                    quantity=quantity,
                    client_row_index=r,
                )
            )
        return rows
    finally:
        wb.close()


def _coerce_float(value: object) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None
