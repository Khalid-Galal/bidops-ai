"""XLSX parser using openpyxl with multi-sheet support.

Parses Excel spreadsheets into the uniform ParsedDocument structure.
Each worksheet becomes a separate PageContent entry. Cell data is
preserved as pipe-delimited tables.

Uses openpyxl with read_only=True and data_only=True for memory
efficiency (avoids loading entire workbook into memory) and to get
computed cell values instead of formulas.
"""

import asyncio
import time
from pathlib import Path

from openpyxl import load_workbook

from app.services.parsing.base import PageContent, ParsedDocument, ParserInterface


def _cell_to_str(value) -> str:
    """Convert any cell value to a string, handling None, dates, numbers."""
    if value is None:
        return ""
    return str(value)


def _load_and_extract(file_path: str) -> dict:
    """Synchronous workbook loading and data extraction.

    Separated into its own function so it can be run via asyncio.to_thread()
    without blocking the event loop.

    Args:
        file_path: Absolute path to the XLSX/XLS file.

    Returns:
        Dict with keys: sheet_names, sheets_data (list of dicts per sheet).
    """
    wb = load_workbook(file_path, data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    sheets_data: list[dict] = []

    for sheet_idx, sheet_name in enumerate(sheet_names):
        sheet = wb[sheet_name]
        rows: list[list[str]] = []

        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows (all cells None).
            if not any(cell is not None for cell in row):
                continue
            row_values = [_cell_to_str(cell) for cell in row]
            rows.append(row_values)

        sheets_data.append({
            "sheet_name": sheet_name,
            "sheet_idx": sheet_idx,
            "rows": rows,
        })

    wb.close()

    return {
        "sheet_names": sheet_names,
        "sheets_data": sheets_data,
    }


class XlsxParser(ParserInterface):
    """Parse XLSX/XLS files using openpyxl.

    Each sheet is treated as a separate "page". Cell values are converted
    to strings and organized into tables with headers derived from the
    first row of each sheet.

    Attributes:
        supported_extensions: [".xlsx", ".xls"]
    """

    supported_extensions: list[str] = [".xlsx", ".xls"]

    async def parse(self, file_path: str) -> ParsedDocument:
        """Parse an XLSX file and return a uniform ParsedDocument.

        The sync openpyxl load_workbook() call is wrapped in
        asyncio.to_thread() to avoid blocking the event loop.

        Args:
            file_path: Absolute path to the XLSX/XLS file.

        Returns:
            ParsedDocument with each sheet as a page and tables extracted.
        """
        start_ms = time.perf_counter()
        warnings: list[str] = []

        try:
            # Run sync openpyxl operations in a thread.
            raw = await asyncio.to_thread(_load_and_extract, file_path)

            sheet_names: list[str] = raw["sheet_names"]
            sheets_data: list[dict] = raw["sheets_data"]

            all_text_parts: list[str] = []
            all_tables: list[dict] = []
            pages: list[PageContent] = []

            for sheet_info in sheets_data:
                sheet_name = sheet_info["sheet_name"]
                sheet_idx = sheet_info["sheet_idx"]
                rows = sheet_info["rows"]

                # Build text representation.
                sheet_text_parts = [f"=== Sheet: {sheet_name} ==="]
                for row in rows:
                    sheet_text_parts.append(" | ".join(row))

                sheet_text = "\n".join(sheet_text_parts)
                all_text_parts.append(sheet_text)

                # Build table dict.
                if rows:
                    headers = rows[0]
                    data = rows[1:] if len(rows) > 1 else rows
                    table_dict = {
                        "page": sheet_idx + 1,
                        "sheet": sheet_name,
                        "headers": headers,
                        "data": data,
                        "rows": len(rows),
                        "cols": len(headers),
                    }
                    all_tables.append(table_dict)
                    page_tables = [table_dict]
                else:
                    page_tables = []
                    warnings.append(f"Sheet '{sheet_name}' is empty (skipped)")

                pages.append(
                    PageContent(
                        page_number=sheet_idx + 1,
                        text=sheet_text,
                        tables=page_tables,
                    )
                )

            full_text = "\n\n".join(all_text_parts)
            processing_time_ms = int((time.perf_counter() - start_ms) * 1000)

            return ParsedDocument(
                filename=Path(file_path).name,
                content_type="xlsx",
                full_text=full_text,
                pages=pages,
                tables=all_tables,
                metadata={
                    "sheet_count": len(sheet_names),
                    "sheet_names": sheet_names,
                },
                page_count=len(sheet_names),
                processing_time_ms=processing_time_ms,
                warnings=warnings,
            )

        except Exception as exc:
            processing_time_ms = int((time.perf_counter() - start_ms) * 1000)
            print(f"[XlsxParser] Error parsing {file_path}: {exc}")
            return ParsedDocument(
                filename=Path(file_path).name,
                content_type="xlsx",
                full_text="",
                pages=[],
                tables=[],
                metadata={},
                page_count=0,
                processing_time_ms=processing_time_ms,
                warnings=[f"Parse error: {exc}"],
            )
