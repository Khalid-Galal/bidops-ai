"""Excel export service for generating styled project reports.

Generates an .xlsx workbook with two sheets:
- Project Summary: All 13 extracted fields with confidence and citations.
- Requirements Checklist: All checklist items grouped by category.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select

from app.database import async_session_factory
from app.models.project import Project
from app.schemas.checklist import RequirementsChecklist
from app.schemas.extraction import ProjectSummary

# Human-readable labels for the 13 summary fields
FIELD_LABELS: dict[str, str] = {
    "project_name": "Project Name",
    "project_owner": "Project Owner",
    "location": "Location",
    "submission_deadline": "Submission Deadline",
    "bid_validity_period": "Bid Validity Period",
    "pre_bid_meeting_date": "Pre-Bid Meeting Date",
    "scope_of_work": "Scope of Work",
    "contract_type": "Contract Type",
    "tender_bond": "Tender Bond",
    "advance_payment": "Advance Payment",
    "retention_percentage": "Retention Percentage",
    "payment_terms": "Payment Terms",
    "stakeholders": "Stakeholders",
}

# Header styling
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="left", vertical="center")


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths based on cell content length."""
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_length:
                    max_length = cell_len
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def _style_header_row(ws, headers: list[str]) -> None:
    """Write and style a header row."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


async def generate_excel_report(project_id: int) -> BytesIO:
    """Generate a styled Excel report for the given project.

    Args:
        project_id: Database ID of the project.

    Returns:
        BytesIO buffer containing the .xlsx workbook.

    Raises:
        ValueError: If the project does not exist.
    """
    async with async_session_factory() as session:
        result = await session.execute(
            select(Project).where(Project.id == project_id)
        )
        project = result.scalar_one_or_none()
        if project is None:
            raise ValueError(f"Project with id {project_id} not found")

    wb = Workbook()

    # ── Summary Sheet ───────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Project Summary"

    _style_header_row(ws_summary, [
        "Field", "Value", "Confidence", "Source Document", "Page", "Requires Review",
    ])

    if project.summary_json:
        summary = ProjectSummary.model_validate_json(project.summary_json)
        for field_key, label in FIELD_LABELS.items():
            field = getattr(summary, field_key, None)
            if field is None:
                ws_summary.append([label, "", "", "", "", ""])
                continue
            first_citation = field.citations[0] if field.citations else None
            ws_summary.append([
                label,
                field.value or "",
                field.confidence_level,
                first_citation.document_name if first_citation else "",
                first_citation.page_number if first_citation else "",
                "Yes" if field.requires_review else "No",
            ])
    else:
        ws_summary.append(["No extraction results available", "", "", "", "", ""])

    _auto_fit_columns(ws_summary)

    # ── Checklist Sheet ─────────────────────────────────────────
    ws_checklist = wb.create_sheet("Requirements Checklist")

    checklist_headers = [
        "#", "Requirement", "Description", "Category", "Mandatory",
        "Confidence", "Source", "Page", "Status",
    ]
    # Write header manually (since ws is not empty-first-row for create_sheet)
    ws_checklist.append(checklist_headers)
    for cell in ws_checklist[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT

    if project.checklist_json:
        checklist = RequirementsChecklist.model_validate_json(project.checklist_json)
        row_num = 0
        categories = [
            ("requirements", "Requirements"),
            ("submission_documents", "Submission Documents"),
            ("eligibility_criteria", "Eligibility Criteria"),
        ]
        for cat_key, cat_label in categories:
            items = getattr(checklist, cat_key, [])
            for item in items:
                row_num += 1
                # Determine checked status from the raw JSON dict
                checked = getattr(item, "checked", False) if hasattr(item, "checked") else False
                citation = item.citation
                ws_checklist.append([
                    row_num,
                    item.requirement,
                    item.description,
                    cat_label,
                    "Yes" if item.is_mandatory else "No",
                    item.confidence_level,
                    citation.document_name if citation else "",
                    citation.page_number if citation else "",
                    "Checked" if checked else "Unchecked",
                ])
    else:
        ws_checklist.append(["", "No checklist results available", "", "", "", "", "", "", ""])

    _auto_fit_columns(ws_checklist)

    # ── Save to buffer ──────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
