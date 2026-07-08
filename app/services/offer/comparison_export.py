"""Render an offer-comparison matrix to an .xlsx workbook (openpyxl)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADERS = [
    "Rank", "Supplier", "Total Price", "Currency", "VAT", "Validity (days)",
    "Delivery (weeks)", "Delivery Terms", "Payment Terms", "Commercial Score",
    "Technical Score", "Overall Score", "Status", "Exclusions", "Deviations",
]
_WIDTHS = [6, 28, 14, 9, 16, 14, 16, 18, 18, 16, 16, 14, 14, 11, 11]


def _vat_label(offer: dict) -> str:
    """Human-readable VAT status for the comparison matrix (see item 24's
    ex-VAT normalization in scoring_service._ex_vat_price for the ranking
    basis; this column just shows what was extracted)."""
    included = offer.get("vat_included")
    amount = offer.get("vat_amount")
    if included is True:
        return f"Incl ({amount})" if amount is not None else "Incl"
    if included is False:
        return "Excl"
    return "?"


def build_comparison_workbook(comparison: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Offer Comparison"

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_font = Font(bold=True, color="9C0006")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    last_col = get_column_letter(len(_HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"Offer Comparison - {comparison.get('package_name', '')}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Total Offers:"
    ws["B3"] = comparison.get("total_offers", 0)
    ws["C3"] = "Lowest Price:"
    ws["D3"] = comparison.get("price_min")
    ws["E3"] = comparison.get("currency", "")

    header_row = 5
    warnings = comparison.get("warnings") or []
    if warnings:
        ws.merge_cells(f"A4:{last_col}4")
        ws["A4"] = "WARNING: " + " | ".join(warnings)
        ws["A4"].font = warn_font
        ws["A4"].fill = warn_fill
        ws["A4"].alignment = Alignment(horizontal="left", wrap_text=True)

    for col, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row_idx, offer in enumerate(comparison.get("offers", []), header_row + 1):
        values = [
            offer.get("rank") or "-",
            offer.get("supplier_name") or "",
            offer.get("total_price") if offer.get("total_price") is not None else "-",
            offer.get("currency") or "",
            _vat_label(offer),
            offer.get("validity_days") if offer.get("validity_days") is not None else "-",
            offer.get("delivery_weeks") if offer.get("delivery_weeks") is not None else "-",
            offer.get("delivery_terms") or "-",
            offer.get("payment_terms") or "-",
            round(offer.get("commercial_score") or 0, 1),
            round(offer.get("technical_score") or 0, 1),
            round(offer.get("overall_score") or 0, 1),
            offer.get("status") or "",
            offer.get("exclusions_count", len(offer.get("exclusions") or [])),
            offer.get("deviations_count", len(offer.get("deviations") or [])),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            if offer.get("rank") == 1:
                cell.fill = best_fill

    for col, width in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    _write_exclusions_sheet(wb, comparison, header_font, header_fill, border)
    return wb


def _write_exclusions_sheet(wb: Workbook, comparison: dict, header_font, header_fill, border) -> None:
    """Verbatim exclusions/deviations per offer -- the counts-only columns on
    the main sheet don't tell an evaluator WHAT was excluded/deviated."""
    ws = wb.create_sheet("Exclusions & Deviations")
    headers = ["Supplier", "Type", "Text"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    row_idx = 2
    for offer in comparison.get("offers", []):
        supplier_name = offer.get("supplier_name") or ""
        entries = [("Exclusion", t) for t in (offer.get("exclusions") or [])]
        entries += [("Deviation", t) for t in (offer.get("deviations") or [])]
        if not entries:
            entries = [("-", "None stated")]
        for kind, text in entries:
            ws.cell(row=row_idx, column=1, value=supplier_name).border = border
            ws.cell(row=row_idx, column=2, value=kind).border = border
            cell = ws.cell(row=row_idx, column=3, value=text)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
            row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 80


def export_comparison_excel(comparison: dict, output_path: str) -> str:
    wb = build_comparison_workbook(comparison)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
