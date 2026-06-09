"""Historical price learning: corpus management + benchmark suggestions.

Matching reuses the deterministic Phase 11 fuzzy matcher; an optional
semantic_scorer(a, b) -> float can be injected to blend an embedding score
(the local sentence-transformer is a dependency but is intentionally not wired
into the default/tested path). Pure logic — no Gemini key.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from statistics import mean, median

from openpyxl import load_workbook
from sqlalchemy import delete, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.boq import BOQItem
from app.models.historical import HistoricalPrice
from app.models.project import Project
from app.services.pricing.line_item_matcher import DEFAULT_THRESHOLD, match_score

_DEFAULT_TOP_K = 5

# Editable corpus fields shared by add/import (keeps DRY).
_SETTABLE = (
    "description", "description_ar", "unit", "rate", "currency", "trade_category",
)


class HistoricalService:
    def __init__(self, semantic_scorer=None) -> None:
        self._semantic_scorer = semantic_scorer

    async def add(
        self,
        db: AsyncSession,
        *,
        description: str,
        rate: float,
        source: str = "manual",
        source_project_id: int | None = None,
        recorded_at: datetime | None = None,
        **fields,
    ) -> HistoricalPrice:
        rec = HistoricalPrice(
            description=description,
            rate=rate,
            source=source,
            source_project_id=source_project_id,
            recorded_at=recorded_at or datetime.now(timezone.utc),
            **{k: v for k, v in fields.items() if k in _SETTABLE},
        )
        db.add(rec)
        await db.commit()
        await db.refresh(rec)
        return rec

    async def list_records(
        self, db: AsyncSession, *, trade: str | None = None, limit: int = 200
    ) -> list[HistoricalPrice]:
        stmt = select(HistoricalPrice)
        if trade:
            stmt = stmt.where(HistoricalPrice.trade_category == trade)
        stmt = stmt.order_by(HistoricalPrice.id.desc()).limit(limit)
        return list((await db.execute(stmt)).scalars().all())

    @staticmethod
    def _benchmark(rates: list[float], currency: str | None) -> dict:
        if not rates:
            return {
                "count": 0, "min": None, "max": None, "avg": None,
                "median": None, "suggested_rate": None, "currency": currency,
            }
        med = round(median(rates), 2)
        return {
            "count": len(rates),
            "min": round(min(rates), 2),
            "max": round(max(rates), 2),
            "avg": round(mean(rates), 2),
            "median": med,
            "suggested_rate": med,
            "currency": currency,
        }

    async def suggest(
        self,
        db: AsyncSession,
        description: str,
        *,
        unit: str | None = None,
        trade: str | None = None,
        top_k: int = _DEFAULT_TOP_K,
        min_score: float = DEFAULT_THRESHOLD,
        exclude_project_id: int | None = None,
    ) -> dict:
        stmt = select(HistoricalPrice)
        if trade:
            stmt = stmt.where(HistoricalPrice.trade_category == trade)
        if exclude_project_id is not None:
            stmt = stmt.where(
                or_(
                    HistoricalPrice.source_project_id != exclude_project_id,
                    HistoricalPrice.source_project_id.is_(None),
                )
            )
        records = list((await db.execute(stmt)).scalars().all())

        scored: list[tuple[HistoricalPrice, float]] = []
        for rec in records:
            score = match_score(description, rec.description)
            if self._semantic_scorer is not None:
                score = max(score, float(self._semantic_scorer(description, rec.description)))
            if score >= min_score:
                scored.append((rec, round(score, 4)))
        scored.sort(key=lambda rs: rs[1], reverse=True)
        top = scored[:top_k]

        rates = [rec.rate for rec, _ in top]
        currency = next((rec.currency for rec, _ in top if rec.currency), None)
        matches = [
            {
                "historical_id": rec.id,
                "description": rec.description,
                "unit": rec.unit,
                "rate": rec.rate,
                "currency": rec.currency,
                "source": rec.source,
                "source_project_id": rec.source_project_id,
                "similarity": score,
            }
            for rec, score in top
        ]
        return {
            "query": description,
            "trade": trade,
            "benchmark": self._benchmark(rates, currency),
            "matches": matches,
        }

    _COLUMN_CANDIDATES = {
        "description": ("description", "desc", "item", "item description", "particulars"),
        "unit": ("unit", "uom", "u/m"),
        "rate": ("rate", "unit rate", "unit price", "price", "amount"),
        "trade": ("trade", "trade category", "category", "division"),
        "currency": ("currency", "ccy"),
    }

    async def import_excel(
        self, db: AsyncSession, file_path: str, *, source: str | None = None
    ) -> dict:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - surface a clean message
            raise ValueError(f"Failed to read Excel file: {exc}") from exc
        rows = None
        src = source or f"import:{path.name}"
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                raise ValueError("Excel file is empty")
            normalized = [_norm_header(h) for h in header]
            col = {
                key: _find_col(normalized, cands)
                for key, cands in self._COLUMN_CANDIDATES.items()
            }
            if col["description"] is None or col["rate"] is None:
                raise ValueError("Required columns 'description' and 'rate' not found")

            imported = skipped = 0
            errors: list[str] = []
            for row_idx, row in enumerate(rows, start=2):
                try:
                    description = _cell(row, col["description"])
                    rate = _coerce_float(_cell(row, col["rate"]))
                    if not description or rate is None or rate <= 0:
                        skipped += 1
                        continue
                    trade = _cell(row, col["trade"])
                    db.add(
                        HistoricalPrice(
                            description=description,
                            unit=_cell(row, col["unit"]),
                            rate=rate,
                            currency=_cell(row, col["currency"]),
                            trade_category=_norm_trade(trade) if trade else None,
                            source=src,
                            recorded_at=datetime.now(timezone.utc),
                        )
                    )
                    imported += 1
                except Exception as exc:  # noqa: BLE001 - per-row resilience
                    errors.append(f"Row {row_idx}: {exc}")
            await db.commit()
        finally:
            if rows is not None:
                rows.close()
            wb.close()
        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors[:10],
            "total_errors": len(errors),
        }

    async def index_project(self, db: AsyncSession, project_id: int) -> dict:
        project = await db.get(Project, project_id)
        if project is None:
            raise ValueError(f"Project {project_id} not found")
        # Idempotent: clear this project's prior snapshot first.
        await db.execute(
            delete(HistoricalPrice).where(HistoricalPrice.source_project_id == project_id)
        )
        items = list(
            (
                await db.execute(
                    select(BOQItem).where(
                        BOQItem.project_id == project_id,
                        BOQItem.unit_rate.is_not(None),
                        BOQItem.is_excluded.is_(False),
                    )
                )
            ).scalars().all()
        )
        now = datetime.now(timezone.utc)
        for item in items:
            db.add(
                HistoricalPrice(
                    description=item.description,
                    description_ar=item.description_ar,
                    unit=item.unit,
                    rate=item.unit_rate,
                    currency=item.currency,
                    trade_category=item.trade_category,
                    source=f"project:{project.name}",
                    source_project_id=project_id,
                    recorded_at=now,
                )
            )
        await db.commit()
        return {"project_id": project_id, "indexed": len(items)}

    async def record_feedback(
        self,
        db: AsyncSession,
        *,
        description: str,
        accepted_rate: float,
        unit: str | None = None,
        currency: str | None = None,
        trade_category: str | None = None,
    ) -> HistoricalPrice:
        return await self.add(
            db,
            description=description,
            rate=accepted_rate,
            unit=unit,
            currency=currency,
            trade_category=trade_category,
            source="feedback",
        )


def _norm_header(value: object) -> str:
    return str(value or "").lower().strip()


def _find_col(normalized: list[str], candidates: tuple[str, ...]) -> int | None:
    for cand in candidates:
        if cand in normalized:
            return normalized.index(cand)
    return None


def _cell(row, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    text = str(val).strip()
    return text or None


def _coerce_float(value: str | None) -> float | None:
    if value is None:
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _norm_trade(value: str) -> str:
    return value.strip().lower().replace(" ", "_")
