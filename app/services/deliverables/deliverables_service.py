"""Assemble the client-ready deliverables bundle for a project.

Collects: Pricing_Summary.xlsx (full cost rollup + by-trade), Pricing_Gaps.xlsx,
one offer-comparison matrix per package with offers, the Packages Register
(if exported), package briefs, and a manifest.json. Rebuilds are idempotent
(the project folder is recreated from scratch). Pure logic — no LLM."""

from __future__ import annotations

import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.package import Package
from app.models.project import Project
from app.services.indirects.indirects_service import IndirectsService
from app.services.offer.comparison_export import export_comparison_excel
from app.services.offer.scoring_service import ScoringService
from app.services.packaging.package_exporter import PackageExporter
from app.services.pricing.pricing_service import PricingService

_SAFE = re.compile(r"[^\w\-]+")


def _safe_name(text: str) -> str:
    return _SAFE.sub("_", text).strip("_") or "package"


class DeliverablesService:
    def __init__(
        self,
        output_root: Path | str = "data/deliverables",
        package_exporter: PackageExporter | None = None,
    ) -> None:
        self._root = Path(output_root)
        self._package_exporter = package_exporter

    def project_dir(self, project_id: int) -> Path:
        return self._root / f"project_{project_id}"

    async def build(
        self,
        db: AsyncSession,
        project_id: int,
        *,
        duration_months: int = 0,
        location: str = "default",
    ) -> dict:
        """Assemble the deliverables bundle for a project.

        duration_months/location feed the indirects rollup that drives
        Pricing_Summary.xlsx and the manifest grand_total.

        Note: the Packages Register is copied AS-IS from the packaging export
        and may predate re-generated packages — re-run the packaging export to
        refresh it before building deliverables.
        """
        project = await db.get(Project, project_id)
        if project is None:
            raise ValueError(f"Project {project_id} not found")

        dest = self.project_dir(project_id)
        if dest.exists():
            shutil.rmtree(dest)  # idempotent rebuild
        dest.mkdir(parents=True)

        files: list[str] = []

        pricing_svc = PricingService()
        pricing = await pricing_svc.pricing_summary(db, project_id)
        cost = await IndirectsService().project_cost_summary(
            db, project_id, duration_months=duration_months, location=location
        )
        self._write_pricing_workbook(cost, pricing, dest / "Pricing_Summary.xlsx")
        files.append("Pricing_Summary.xlsx")

        gaps = await pricing_svc.gaps_report(db, project_id)
        self._write_gaps_workbook(gaps, dest / "Pricing_Gaps.xlsx")
        files.append("Pricing_Gaps.xlsx")

        packages = list(
            (
                await db.execute(
                    select(Package)
                    .where(Package.project_id == project_id)
                    .order_by(Package.code)
                )
            ).scalars().all()
        )
        comparisons = 0
        scoring = ScoringService()
        for package in packages:
            comparison = await scoring.compare(db, package.id)
            if comparison["total_offers"] == 0:
                continue
            name = f"Comparison_{_safe_name(package.code)}.xlsx"
            export_comparison_excel(comparison, str(dest / name))
            files.append(name)
            comparisons += 1

        exporter = self._package_exporter or PackageExporter()
        register = exporter.register_path(project_id)
        if register.exists():
            shutil.copy2(register, dest / "Packages_Register.xlsx")
            files.append("Packages_Register.xlsx")

        briefs = 0
        briefs_dir = dest / "Briefs"
        for package in packages:
            if package.brief_path and Path(package.brief_path).exists():
                briefs_dir.mkdir(exist_ok=True)
                # Exporter names every brief Package_Brief.html/.pdf (parent
                # folder disambiguates); prefix the package code so flattening
                # into Briefs/ cannot silently overwrite earlier copies.
                target = briefs_dir / (
                    f"{_safe_name(package.code)}_{Path(package.brief_path).name}"
                )
                shutil.copy2(package.brief_path, target)
                files.append(f"Briefs/{target.name}")
                briefs += 1

        manifest = {
            "project_id": project_id,
            "project_name": project.name,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "grand_total": cost["grand_total"],
            "currency": cost["currency"],
            "duration_months": duration_months,
            "location": location,
            "files": sorted(files),
            "comparisons": comparisons,
            "briefs": briefs,
        }
        (dest / "manifest.json").write_text(
            json.dumps(manifest, indent=2), encoding="utf-8"
        )
        files.append("manifest.json")

        return {
            "project_id": project_id,
            "folder": str(dest),
            "files": sorted(files),
            "comparisons": comparisons,
            "briefs": briefs,
        }

    @staticmethod
    def _write_pricing_workbook(cost: dict, pricing: dict, path: Path) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Cost Summary"
        rows = [
            ("Direct Cost", cost["direct_cost"]),
            ("Total Indirects", cost["indirects"]["total_indirects"]),
            ("Cost Base (direct + indirects)", cost["total_cost_base"]),
            ("Overhead", cost["markups"]["overhead"]),
            ("Profit", cost["markups"]["profit"]),
            ("Contingency", cost["markups"]["contingency"]),
            ("Risk", cost["markups"]["risk"]),
            ("Markup Total", cost["markups"]["markup_total"]),
            ("Selling Before VAT", cost["selling_before_vat"]),
            (f"VAT ({cost['vat_rate']:.0%})", cost["vat_amount"]),
            ("GRAND TOTAL", cost["grand_total"]),
            ("Currency", cost["currency"]),
        ]
        for label, value in rows:
            ws.append([label, value])
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18

        trade_ws = wb.create_sheet("By Trade")
        trade_ws.append(["Trade", "Items", "Total", "% of Direct"])
        for t in pricing["by_trade"]:
            trade_ws.append([t["trade"], t["count"], t["total"], t["percentage"]])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)

    @staticmethod
    def _write_gaps_workbook(gaps: dict, path: Path) -> None:
        wb = Workbook()
        first = True
        for title, key in (
            ("Unpriced", "unpriced"),
            ("Needs Review", "needs_review"),
            ("Excluded", "excluded"),
        ):
            ws = wb.active if first else wb.create_sheet()
            ws.title = title
            first = False
            ws.append(["ID", "Line", "Description", "Trade", "Reason"])
            for g in gaps[key]:
                ws.append([
                    g["id"], g["line_number"], g["description"],
                    g["trade_category"], g["reason"],
                ])
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
