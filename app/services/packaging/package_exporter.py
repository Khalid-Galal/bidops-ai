"""Generates on-disk package deliverables: folders, BOQ subsets, linked docs,
briefs (HTML always, PDF when WeasyPrint is usable), and a master register."""

from __future__ import annotations

import html
import logging
import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.boq import BOQItem
from app.models.document import Document
from app.models.package import Package, PackageDocument

logger = logging.getLogger(__name__)

_SUBFOLDERS = ("BOQ", "Documents", "Offers", "Clarifications")


def _safe_name(name: str) -> str:
    """Filesystem-safe folder/file component (blocks path traversal)."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", name or "").strip("._")
    if cleaned in ("", ".", ".."):
        return "package"
    return cleaned


class PackageExporter:
    """Writes package folder trees + BOQ subsets + briefs + a master register."""

    def __init__(self, output_root: Path | str = "data/packages") -> None:
        self._root = Path(output_root)

    def register_path(self, project_id: int) -> Path:
        """Canonical on-disk location of a project's master register."""
        return self._root / f"project_{project_id}" / "Packages_Register.xlsx"

    async def export_project(self, db: AsyncSession, project_id: int) -> dict:
        packages = (
            await db.execute(
                select(Package).where(Package.project_id == project_id).order_by(Package.code)
            )
        ).scalars().all()

        project_dir = self._root / f"project_{project_id}"
        project_dir.mkdir(parents=True, exist_ok=True)

        exported: list[dict] = []
        briefs_pdf = 0
        for pkg in packages:
            info = await self._export_package(db, pkg, project_dir)
            exported.append(info)
            if info["brief_pdf"]:
                briefs_pdf += 1

        register_path = self._write_register(project_id, packages)
        await db.commit()  # persist folder_path/brief_path set on packages

        return {
            "project_id": project_id,
            "packages_exported": len(packages),
            "register_path": str(register_path),
            "briefs_pdf": briefs_pdf,
            "packages": exported,
        }

    async def _export_package(
        self, db: AsyncSession, pkg: Package, project_dir: Path
    ) -> dict:
        pkg_dir = project_dir / _safe_name(pkg.code)
        for sub in _SUBFOLDERS:
            (pkg_dir / sub).mkdir(parents=True, exist_ok=True)

        items = (
            await db.execute(
                select(BOQItem)
                .where(BOQItem.package_id == pkg.id)
                .order_by(BOQItem.client_row_index)
            )
        ).scalars().all()

        self._write_boq_subset(pkg_dir / "BOQ" / f"BOQ_{_safe_name(pkg.code)}.xlsx", pkg, items)

        links = (
            await db.execute(
                select(PackageDocument, Document)
                .join(Document, Document.id == PackageDocument.document_id)
                .where(
                    PackageDocument.package_id == pkg.id,
                    Document.is_superseded.is_(False),
                )
                .order_by(PackageDocument.relevance_score.desc())
            )
        ).all()
        self._attach_documents(pkg_dir / "Documents", links)

        brief_html = pkg_dir / "Package_Brief.html"
        self._write_brief_html(brief_html, pkg, items, links)
        brief_pdf = self._try_write_brief_pdf(brief_html, pkg_dir / "Package_Brief.pdf")

        pkg.folder_path = str(pkg_dir)
        pkg.brief_path = str(brief_pdf or brief_html)
        return {
            "package_id": pkg.id,
            "code": pkg.code,
            "folder_path": str(pkg_dir),
            "brief_pdf": bool(brief_pdf),
            "documents_linked": len(links),
        }

    def _write_boq_subset(self, path: Path, pkg: Package, items: list[BOQItem]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"
        # Header row first (row 1) so data items occupy rows 2..N.
        headers = ["Line", "Section", "Description", "Unit", "Quantity"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for it in items:
            ws.append([it.line_number, it.section, it.description, it.unit, it.quantity])
        for col, width in {"A": 10, "B": 24, "C": 60, "D": 10, "E": 14}.items():
            ws.column_dimensions[col].width = width
        wb.save(path)

    def _attach_documents(self, docs_dir: Path, links: list) -> None:
        # Prune previously-copied files (manifest + prior copies) so a re-export
        # with changed linked documents doesn't leave stale files behind.
        for existing in docs_dir.iterdir():
            if existing.is_file():
                existing.unlink()
        lines = ["Linked documents for this package", "=" * 40, ""]
        for pd, doc in links:
            src = Path(doc.file_path)
            status = "COPIED"
            if src.exists():
                try:
                    shutil.copy2(src, docs_dir / _safe_name(doc.filename))
                except OSError as exc:  # pragma: no cover
                    status = f"COPY_FAILED ({exc})"
            else:
                status = "MISSING (source file not found)"
            score = f"{pd.relevance_score:.2f}" if pd.relevance_score is not None else "n/a"
            lines.append(f"- {doc.filename} [{status}] relevance={score}")
            if pd.excerpt:
                lines.append(f"    excerpt: {pd.excerpt}")
        (docs_dir / "linked_manifest.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _write_brief_html(
        self, path: Path, pkg: Package, items: list[BOQItem], links: list
    ) -> None:
        def esc(v) -> str:
            return html.escape(str(v if v is not None else ""))

        rows = "".join(
            f"<tr><td>{esc(i.line_number)}</td><td>{esc(i.description)}</td>"
            f"<td>{esc(i.unit)}</td><td>{esc(i.quantity)}</td></tr>"
            for i in items
        )
        doc_rows = "".join(
            f"<li>{esc(doc.filename)} (relevance "
            f"{esc(f'{pd.relevance_score:.2f}' if pd.relevance_score is not None else 'n/a')})</li>"
            for pd, doc in links
        )
        document = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Package Brief - {esc(pkg.code)}</title>
<style>body{{font-family:sans-serif;margin:2em}}table{{border-collapse:collapse;width:100%}}
td,th{{border:1px solid #999;padding:4px;text-align:left}}h1{{font-size:1.4em}}</style></head>
<body>
<h1>Package Brief: {esc(pkg.name)}</h1>
<p><strong>Code:</strong> {esc(pkg.code)} &nbsp; <strong>Trade:</strong> {esc(pkg.trade_category)}
 &nbsp; <strong>Items:</strong> {esc(pkg.total_items)} &nbsp; <strong>Status:</strong> {esc(pkg.status)}</p>
<h2>Bill of Quantities</h2>
<table><tr><th>Line</th><th>Description</th><th>Unit</th><th>Qty</th></tr>{rows}</table>
<h2>Linked Documents</h2>
<ul>{doc_rows or '<li>None linked</li>'}</ul>
</body></html>"""
        path.write_text(document, encoding="utf-8")

    def _try_write_brief_pdf(self, html_path: Path, pdf_path: Path) -> Path | None:
        """Render the brief HTML to PDF if WeasyPrint + native libs are usable."""
        try:
            import weasyprint

            weasyprint.HTML(string=html_path.read_text(encoding="utf-8")).write_pdf(
                str(pdf_path)
            )
            return pdf_path
        except (ImportError, OSError) as exc:
            logger.info("Package brief PDF skipped (WeasyPrint unavailable): %s", exc)
            return None

    def _write_register(self, project_id: int, packages: list[Package]) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Packages Register"
        headers = [
            "Code", "Name", "Trade", "Items", "Status", "Deadline",
            "Estimated Value", "Currency", "Offers Received", "Folder", "Brief",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for pkg in packages:
            ws.append([
                pkg.code, pkg.name, pkg.trade_category, pkg.total_items,
                pkg.status,
                pkg.submission_deadline.strftime("%Y-%m-%d") if pkg.submission_deadline else "",
                pkg.estimated_value if pkg.estimated_value is not None else "",
                pkg.currency or "",
                pkg.offers_received,
                pkg.folder_path or "", pkg.brief_path or "",
            ])
        widths = {
            "A": 22, "B": 28, "C": 16, "D": 8, "E": 12, "F": 14,
            "G": 16, "H": 10, "I": 14, "J": 40, "K": 40,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        path = self.register_path(project_id)
        wb.save(path)
        return path
