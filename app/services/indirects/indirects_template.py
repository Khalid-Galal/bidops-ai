"""Fill a client's indirects Excel template, formula-preserving.

Rows are matched by fuzzy label similarity (Phase 11 matcher; component names
have underscores normalized to spaces). Only the detected amount column is
written; any target cell holding a formula string is skipped. Loaded with
openpyxl defaults so formulas elsewhere survive (pivot tables / VBA are not
round-tripped — .xlsm is rejected at the API layer)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.services.boq.sheet_select import pick_sheet
from app.services.pricing.line_item_matcher import match_score

# Priority-ordered: the FIRST alias found wins, so "description" beats "item"
# (an "Item" number column must not be mistaken for the label column) and
# "amount" beats "total" (a row-sum "Total" column must not be the write target).
_AMOUNT_ALIASES = ("amount", "value", "cost", "price", "total", "rate")
_LABEL_ALIASES = (
    "description", "particulars", "indirect item", "indirects", "indirect",
    "component", "item",
)
_MAX_HEADER_SCAN = 20
_MATCH_THRESHOLD = 0.45


def _norm(value: object) -> str:
    return str(value).strip().lower() if value is not None else ""


def detect_columns(ws) -> tuple[int | None, int | None, int | None]:
    """Return (label_col, amount_col, header_row) from the first header row
    that names an amount-like column. Aliases are matched in PRIORITY order
    (not column order); label falls back to column 1. header_row is the row
    where the amount alias was found, or None when nothing was detected."""
    for r in range(1, min(ws.max_row, _MAX_HEADER_SCAN) + 1):
        headers: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            cell = _norm(ws.cell(row=r, column=c).value)
            if cell and cell not in headers:
                headers[cell] = c
        amount_col = next((headers[a] for a in _AMOUNT_ALIASES if a in headers), None)
        if amount_col is not None:
            label_col = next((headers[a] for a in _LABEL_ALIASES if a in headers), None)
            return (label_col or 1), amount_col, r
    return None, None, None


def _pick_indirects_sheet(wb):
    """Prefer a sheet explicitly named for indirects; else reuse the BOQ
    heuristic (whose hints would otherwise miss e.g. ["Summary", "Indirects"]
    and fall back to the first sheet)."""
    for name in wb.sheetnames:
        if "indirect" in name.lower():
            return wb[name]
    return pick_sheet(wb)


def populate_indirects_template(
    template_path: str,
    output_path: str,
    components: dict[str, float],
    *,
    amount_column: int | None = None,
    label_column: int | None = None,
) -> dict:
    """Write each component amount next to its best-matching row label.

    Components are assigned via GLOBAL best matching: every (row, component)
    candidate scoring >= threshold is ranked by score, so an exact later row
    cannot be starved by a generic earlier one. Each component is written at
    most once; the detected header row and any row whose label IS a header
    alias are never written. Formula cells are never overwritten.
    Raises ValueError if no amount column can be determined.
    """
    wb = load_workbook(template_path)  # defaults preserve formulas
    try:
        ws = _pick_indirects_sheet(wb)
        det_label, det_amount, header_row = detect_columns(ws)
        label_col = label_column or det_label or 1
        amount_col = amount_column or det_amount
        if amount_col is None:
            raise ValueError(
                "Could not detect an amount column in the template; "
                "pass amount_column explicitly"
            )

        # Collect every candidate (row, component) pairing above the threshold.
        start_row = (header_row + 1) if header_row is not None else 1
        candidates: list[tuple[float, int, str]] = []
        for r in range(start_row, ws.max_row + 1):
            label = ws.cell(row=r, column=label_col).value
            if label is None or str(label).strip() == "":
                continue
            label_text = str(label)
            # Defense for the explicit-override path (header row unknown):
            # never treat a literal header label as a data row.
            if _norm(label_text) in _LABEL_ALIASES or _norm(label_text) in _AMOUNT_ALIASES:
                continue
            for name in components:
                # underscores are word chars to the matcher: normalize to spaces
                score = match_score(label_text, name.replace("_", " "))
                if score >= _MATCH_THRESHOLD:
                    candidates.append((score, r, name))

        # Global best assignment: highest score first (row, then name break ties
        # deterministically); each row and each component used at most once.
        candidates.sort(key=lambda c: (-c[0], c[1], c[2]))
        assignments: dict[int, str] = {}
        used_components: set[str] = set()
        for score, r, name in candidates:
            if r in assignments or name in used_components:
                continue
            assignments[r] = name
            used_components.add(name)

        remaining = dict(components)
        written = skipped_formula = 0
        for r in sorted(assignments):
            name = assignments[r]
            target = ws.cell(row=r, column=amount_col)
            if isinstance(target.value, str) and target.value.startswith("="):
                skipped_formula += 1
                remaining.pop(name)  # row exists but is formula-driven
                continue
            target.value = round(remaining.pop(name), 2)
            written += 1

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return {
            "written": written,
            "skipped_formula": skipped_formula,
            "amount_column": amount_col,
            "label_column": label_col,
            "unmatched_components": sorted(remaining),
        }
    finally:
        wb.close()
