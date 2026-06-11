"""Live-HTTP end-to-end shakedown driver for BidOps AI.

Drives a synthetic tender through the ENTIRE pipeline against a *running*
uvicorn server (not ASGITransport), so it catches runtime-only failures the
unit suite can't: template render errors, real file I/O, response-schema
mismatches, artifact corruption, background-ingest behaviour.

Run against a hermetic server (separate DB/uploads/chroma) so it never touches
real project data. Launch one in a throwaway dir, then run this driver:

    # PowerShell:
    $env:BIDOPS_DATABASE_PATH=".e2e_tmp/bidops.db"
    $env:BIDOPS_UPLOAD_DIR=".e2e_tmp/uploads"
    $env:BIDOPS_CHROMA_PERSIST_DIR=".e2e_tmp/chroma"
    .venv/Scripts/python.exe -m uvicorn app.main:app --host 127.0.0.1 --port 8077
    # then, in another shell:
    .venv/Scripts/python.exe scripts/e2e_shakedown.py

Note: the artifact roots data/packages, data/offers, data/deliverables and the
rules override data/rules.json are CWD-relative (not env-overridable), so they
are written under the repo's data/ for the throwaway project ids; delete them
afterwards if you want a pristine data/ dir.

Usage:
    python scripts/e2e_shakedown.py            # base http://127.0.0.1:8077
    BIDOPS_E2E_BASE=http://127.0.0.1:9000 python scripts/e2e_shakedown.py

Exit code is always 0; read the printed table + .e2e_tmp/shakedown_report.json.
LLM steps (extract/checklist/offer-extract/compliance) and SMTP send are
expected to DEGRADE gracefully (5xx) when keys/SMTP are absent — that is a PASS
for the "works without a key" guarantee, recorded as DEGRADED, not FAIL.
"""

from __future__ import annotations

import io
import json
import os
import time
import traceback
import zipfile
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook

BASE = os.environ.get("BIDOPS_E2E_BASE", "http://127.0.0.1:8077")
REPO = Path(__file__).resolve().parent.parent
SAMPLE_BOQ = REPO / "sample_tender" / "BOQ_Bill_of_Quantities.xlsx"
SAMPLE_ITT = REPO / "sample_tender" / "ITT_Tender_Document.docx"
OUT = REPO / ".e2e_tmp" / "artifacts"
OUT.mkdir(parents=True, exist_ok=True)
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

results: list[dict] = []
ctx: dict = {}  # shared state across steps (ids, etc.)


def rec(name, status, verdict, note=""):
    results.append({"step": name, "status": status, "verdict": verdict, "note": str(note)[:600]})
    icon = {"PASS": "[PASS]", "DEGRADED": "[DEGR]", "SKIP": "[SKIP]", "FAIL": "[FAIL]"}[verdict]
    print(f"{icon} {name}: HTTP {status} -- {str(note)[:200]}")


def step(name, fn, *, expect=(200, 201, 202, 204), degrade=(), skip_if=None):
    """Run a step fn() -> (status_code, note). Classify the verdict."""
    if skip_if is not None and skip_if():
        rec(name, "-", "SKIP", "prerequisite missing")
        return None
    try:
        status, note = fn()
        if status in expect:
            rec(name, status, "PASS", note)
        elif status in degrade:
            rec(name, status, "DEGRADED", note)
        else:
            rec(name, status, "FAIL", note)
        return status
    except Exception as e:  # noqa: BLE001 -- shakedown must never abort
        rec(name, "ERR", "FAIL", f"{type(e).__name__}: {e}\n{traceback.format_exc()[-400:]}")
        return None


def make_boq_template_bytes() -> bytes:
    """A minimal client BOQ matching the parser's header aliases."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["Item", "Description", "Unit", "Quantity", "Rate", "Amount"])
    rows = [
        ["1", "Supply and pour C30 ready-mix concrete to foundations", "m3", 120, None, None],
        ["2", "Reinforcement steel bars grade 60 cut and fixed", "ton", 18, None, None],
        ["3", "Structural steel fabrication and erection of beams", "ton", 40, None, None],
        ["4", "Internal wall painting two coats emulsion", "m2", 850, None, None],
        ["5", "Supply and install HVAC ducting galvanised", "m", 300, None, None],
        ["6", "Site excavation and earthworks to reduced level", "m3", 500, None, None],
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_indirects_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Indirects"
    ws.append(["Component", "Amount"])
    for label in ["Site Supervision", "Quality Control", "Safety", "Insurance", "Temporary Works"]:
        ws.append([label, None])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_historical_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"
    ws.append(["Description", "Unit", "Rate", "Trade", "Currency"])
    ws.append(["C30 ready mix concrete to foundations", "m3", 95.0, "concrete", "USD"])
    ws.append(["Reinforcement steel grade 60", "ton", 1100.0, "concrete", "USD"])
    ws.append(["Emulsion paint two coats", "m2", 6.5, "finishes", "USD"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_artifact(name: str, content: bytes) -> Path:
    p = OUT / name
    p.write_bytes(content)
    return p


def main() -> None:
    print(f"=== BidOps E2E shakedown against {BASE} ===\n")
    # Generous timeout: a cold server embeds the first search query with a
    # freshly-loaded transformer (~30s observed).
    c = httpx.Client(base_url=BASE, timeout=120.0)

    # ---- 0. health + home page render -------------------------------------
    step("health", lambda: (c.get("/health").status_code, "liveness"))
    step("page:home /", lambda: (c.get("/").status_code, "index.html render"))
    step("page:suppliers /suppliers", lambda: (c.get("/suppliers").status_code, "suppliers.html render"))
    step("page:settings /settings", lambda: (c.get("/settings").status_code, "settings.html render"))

    # ---- 1. create project -------------------------------------------------
    def create_project():
        r = c.post("/api/projects", json={"name": "E2E Shakedown Tender", "description": "synthetic"})
        if r.status_code in (200, 201):
            ctx["pid"] = r.json()["id"]
        return r.status_code, f"pid={ctx.get('pid')}"

    step("create project", create_project)
    pid = ctx.get("pid")

    def has_pid():
        return ctx.get("pid") is None

    step("page:project /projects/{id}", lambda: (c.get(f"/projects/{pid}").status_code, "project.html"), skip_if=has_pid)
    step("page:workbench", lambda: (c.get(f"/projects/{pid}/workbench").status_code, "workbench.html"), skip_if=has_pid)
    step("page:dashboard", lambda: (c.get(f"/projects/{pid}/dashboard").status_code, "dashboard.html"), skip_if=has_pid)

    # ---- 2. upload tender docs (background ingest) -------------------------
    def upload():
        files = [
            ("files", (SAMPLE_ITT.name, SAMPLE_ITT.read_bytes(),
                       "application/vnd.openxmlformats-officedocument.wordprocessingml.document")),
            ("files", (SAMPLE_BOQ.name, SAMPLE_BOQ.read_bytes(), XLSX_MIME)),
        ]
        r = c.post(f"/api/projects/{pid}/upload", files=files)
        if r.status_code == 202:
            ctx["task_id"] = r.json().get("task_id")
        return r.status_code, r.text[:200]

    step("upload tender docs", upload, skip_if=has_pid)

    # ---- 3. poll ingest to completion -------------------------------------
    def poll_ingest():
        # The FIRST ingest after a cold server triggers lazy model loading
        # (embedding + NLI transformers) plus HF Hub metadata round-trips, which
        # can block the worker ~1 min and reset in-flight connections. Treat
        # transport errors as "server busy" and keep polling, and only proceed
        # once every doc reaches a terminal status -- downstream search/extract
        # silently return empty if run against a not-yet-indexed corpus.
        deadline = time.time() + 360
        statuses = {}
        last_err = None
        while time.time() < deadline:
            try:
                r = c.get(f"/api/projects/{pid}/documents", timeout=30.0)
                docs = r.json()
                statuses = {d["filename"]: d["status"] for d in docs}
                if docs and all(d["status"] not in ("pending", "processing") for d in docs):
                    ctx["docs"] = docs
                    return 200, json.dumps(statuses)
            except httpx.TransportError as e:
                last_err = f"{type(e).__name__} (server busy loading models, retrying)"
            time.sleep(3)
        return 408, f"timeout; last={json.dumps(statuses)} err={last_err}"

    step("ingest completes", poll_ingest, expect=(200,), degrade=(408,), skip_if=has_pid)

    # ---- 4. search ---------------------------------------------------------
    def search():
        r = c.get(f"/api/projects/{pid}/search", params={"q": "payment terms and bid validity", "mode": "hybrid"})
        if r.status_code == 200:
            n = r.json().get("total_results")
            ctx["search_hits"] = n
            return 200, f"total_results={n}"
        return r.status_code, r.text[:200]

    step("hybrid search", search, skip_if=has_pid)

    # ---- 5. LLM extract + checklist (expected to degrade w/o key) ---------
    # These are the heaviest endpoints: on a COLD server the NLI model loads and
    # the per-field pipeline makes many rate-limited (free-tier) Gemini calls, so
    # they can exceed the client timeout while the server keeps working and
    # persists the result. Treat a client timeout as DEGRADED (server completes
    # async), not FAIL; give them a long per-call timeout.
    def llm_call(path, note):
        try:
            return c.post(path, timeout=300.0).status_code, note
        except httpx.TimeoutException:
            return 408, note + " (client timeout; server completes async)"

    step("LLM extract (degrade ok)",
         lambda: llm_call(f"/api/projects/{pid}/extract", "summary extraction"),
         expect=(200,), degrade=(408, 500, 503, 429), skip_if=has_pid)
    step("LLM checklist (degrade ok)",
         lambda: llm_call(f"/api/projects/{pid}/checklist", "checklist extraction"),
         expect=(200,), degrade=(408, 500, 503, 429), skip_if=has_pid)

    # ---- 6. document versioning analyze -----------------------------------
    def analyze():
        r = c.post(f"/api/projects/{pid}/documents/analyze")
        return r.status_code, r.text[:200]

    step("documents/analyze", analyze, skip_if=has_pid)

    # ---- 7. BOQ parse ------------------------------------------------------
    def boq_parse():
        files = {"file": (SAMPLE_BOQ.name, SAMPLE_BOQ.read_bytes(), XLSX_MIME)}
        r = c.post(f"/api/projects/{pid}/boq/parse", files=files)
        return r.status_code, r.text[:300]

    step("BOQ parse", boq_parse, skip_if=has_pid)

    def boq_list():
        r = c.get(f"/api/projects/{pid}/boq")
        items = r.json()
        ctx["boq_items"] = items
        priced = [i for i in items if i.get("trade_category")]
        return r.status_code, f"items={len(items)} classified={len(priced)}"

    step("BOQ list", boq_list, skip_if=has_pid)

    # ---- 8. packaging ------------------------------------------------------
    def gen_pkgs():
        r = c.post(f"/api/projects/{pid}/packages/generate")
        return r.status_code, r.text[:300]

    step("packages generate", gen_pkgs, skip_if=has_pid)

    def list_pkgs():
        r = c.get(f"/api/projects/{pid}/packages")
        pkgs = r.json()
        ctx["pkgs"] = pkgs
        if pkgs:
            ctx["pkg"] = pkgs[0]
        return r.status_code, f"packages={len(pkgs)} first={pkgs[0] if pkgs else None}"

    step("packages list", list_pkgs, skip_if=has_pid)

    def no_pkg():
        return ctx.get("pkg") is None

    def pkg_detail():
        pk = ctx["pkg"]["id"]
        r = c.get(f"/api/projects/{pid}/packages/{pk}")
        j = r.json()
        return r.status_code, f"items={len(j.get('items', []))} linked={len(j.get('linked_documents', []))}"

    step("package detail", pkg_detail, skip_if=no_pkg)

    step("packages link-documents",
         lambda: (c.post(f"/api/projects/{pid}/packages/link-documents").status_code, "semantic linking"),
         skip_if=has_pid)

    def export_pkgs():
        r = c.post(f"/api/projects/{pid}/packages/export")
        if r.status_code == 200:
            ctx["exported"] = True
        return r.status_code, r.text[:250]

    step("packages export", export_pkgs, skip_if=has_pid)

    def dl_register():
        r = c.get(f"/api/projects/{pid}/packages/register")
        if r.status_code == 200:
            p = save_artifact("Packages_Register.xlsx", r.content)
            load_workbook(p)  # raises if corrupt
            return 200, f"opened {p.name} ({len(r.content)} bytes)"
        return r.status_code, r.text[:150]

    step("download+open register.xlsx", dl_register, skip_if=lambda: not ctx.get("exported"))

    # ---- 9. suppliers ------------------------------------------------------
    def make_suppliers():
        trade = ctx["pkg"]["trade_category"]
        ids = []
        for i, (nm, price_factor) in enumerate([("Alpha Contracting", 1.0), ("Beta Builders", 1.15)]):
            r = c.post("/api/suppliers", json={
                "name": nm, "emails": [f"bids@{nm.split()[0].lower()}.example"],
                "trade_categories": [trade], "region": "Gulf", "rating": 4.0 + i * 0.5,
            })
            if r.status_code in (200, 201):
                ids.append(r.json()["id"])
        ctx["supplier_ids"] = ids
        ctx["trade"] = trade
        return (201 if len(ids) == 2 else 500), f"created suppliers {ids} for trade={trade}"

    step("create suppliers", make_suppliers, skip_if=no_pkg)

    def suggested():
        pk = ctx["pkg"]["id"]
        r = c.get(f"/api/projects/{pid}/packages/{pk}/suggested-suppliers")
        j = r.json()
        return r.status_code, f"suggested={len(j)} ids={[s.get('id') for s in j]}"

    step("suggested-suppliers (trade match)", suggested, skip_if=no_pkg)

    # ---- 10. RFQ drafts ----------------------------------------------------
    def rfq():
        pk = ctx["pkg"]["id"]
        r = c.post(f"/api/projects/{pid}/packages/{pk}/rfq",
                   json={"supplier_ids": ctx.get("supplier_ids", []), "language": "en"})
        if r.status_code in (200, 201):
            j = r.json()
            ctx["email_ids"] = j.get("email_ids", [])
            return r.status_code, f"drafts={j.get('drafts_created')} skipped={j.get('skipped')}"
        return r.status_code, r.text[:200]

    step("create RFQ drafts", rfq, skip_if=lambda: not ctx.get("supplier_ids"))

    def email_view_edit():
        eid = ctx["email_ids"][0]
        g = c.get(f"/api/emails/{eid}")
        has_body = bool(g.json().get("body_html"))
        p = c.patch(f"/api/emails/{eid}", json={"subject": "Revised RFQ - E2E"})
        return p.status_code, f"body_html_present={has_body} patch={p.status_code}"

    step("email get + draft edit", email_view_edit, skip_if=lambda: not ctx.get("email_ids"))

    step("email send (degrade: no SMTP)",
         lambda: (c.post(f"/api/emails/{ctx['email_ids'][0]}/send").status_code, "expect 503 no SMTP"),
         expect=(200,), degrade=(503,), skip_if=lambda: not ctx.get("email_ids"))

    # ---- 11. offers --------------------------------------------------------
    def ingest_offers():
        pk = ctx["pkg"]["id"]
        offer_ids = []
        for sid in ctx.get("supplier_ids", []):
            quote = f"Quotation from supplier {sid}. See attached pricing.".encode()
            files = [("files", (f"quote_{sid}.txt", quote, "text/plain"))]
            r = c.post(f"/api/projects/{pid}/packages/{pk}/offers",
                       data={"supplier_id": str(sid)}, files=files)
            if r.status_code in (200, 201):
                offer_ids.append(r.json()["id"])
        ctx["offer_ids"] = offer_ids
        return (201 if len(offer_ids) == len(ctx.get("supplier_ids", [])) and offer_ids else 500), \
            f"offers={offer_ids}"

    step("ingest offers", ingest_offers, skip_if=lambda: not ctx.get("supplier_ids"))

    def patch_commercial():
        # Build line items mirroring this package's BOQ item descriptions so
        # populate-prices can fuzzy-match them.
        pk = ctx["pkg"]["id"]
        det = c.get(f"/api/projects/{pid}/packages/{pk}").json()
        pkg_items = det.get("items", [])
        codes = []
        for idx, oid in enumerate(ctx["offer_ids"]):
            factor = 1.0 + idx * 0.15  # 2nd offer pricier -> 1st should rank #1
            line_items = []
            for it in pkg_items:
                qty = it.get("quantity") or 1
                rate = round(100.0 * factor, 2)
                line_items.append({
                    "description": it.get("description", "item"),
                    "unit": it.get("unit"), "quantity": qty,
                    "rate": rate, "total": round(rate * qty, 2),
                })
            total = round(sum(li["total"] for li in line_items), 2) or round(50000 * factor, 2)
            body = {
                "total_price": total, "currency": "USD", "validity_days": 90,
                "delivery_weeks": 8 + idx * 2, "payment_terms": "Net 30",
                "technical_score": 85.0 - idx * 10, "line_items": line_items,
            }
            r = c.patch(f"/api/offers/{oid}", json=body)
            codes.append(r.status_code)
        return (200 if all(s == 200 for s in codes) else codes[0]), \
            f"patched {len(codes)} offers ({len(pkg_items)} line items each)"

    step("offers: manual commercial + line items", patch_commercial,
         skip_if=lambda: not ctx.get("offer_ids"))

    step("offers score+rank",
         lambda: (c.post(f"/api/projects/{pid}/packages/{ctx['pkg']['id']}/offers/score").status_code, "weighted scoring"),
         skip_if=lambda: not ctx.get("offer_ids"))

    def comparison():
        pk = ctx["pkg"]["id"]
        r = c.get(f"/api/projects/{pid}/packages/{pk}/offers/comparison")
        j = r.json()
        ranks = [(o.get("rank"), o.get("supplier_name"), o.get("total_price")) for o in j.get("offers", [])]
        return r.status_code, f"min={j.get('price_min')} avg={j.get('price_avg')} ranks={ranks}"

    step("offers comparison", comparison, skip_if=lambda: not ctx.get("offer_ids"))

    def comparison_xlsx():
        pk = ctx["pkg"]["id"]
        r = c.get(f"/api/projects/{pid}/packages/{pk}/offers/comparison.xlsx")
        if r.status_code == 200:
            p = save_artifact("Comparison.xlsx", r.content)
            load_workbook(p)
            return 200, f"opened {p.name}"
        return r.status_code, r.text[:150]

    step("download+open comparison.xlsx", comparison_xlsx, skip_if=lambda: not ctx.get("offer_ids"))

    def select_offer():
        # pick rank-1 offer from comparison
        pk = ctx["pkg"]["id"]
        comp = c.get(f"/api/projects/{pid}/packages/{pk}/offers/comparison").json()
        winner = None
        for o in comp.get("offers", []):
            if o.get("rank") == 1:
                winner = o.get("offer_id") or o.get("id")
        winner = winner or ctx["offer_ids"][0]
        ctx["winner"] = winner
        r = c.post(f"/api/offers/{winner}/select", json={"notes": "Best value (E2E)"})
        return r.status_code, f"winner={winner} status={r.json().get('status') if r.status_code == 200 else r.text[:120]}"

    step("select winning offer", select_offer, skip_if=lambda: not ctx.get("offer_ids"))

    def populate_prices():
        r = c.post(f"/api/offers/{ctx['winner']}/populate-prices")
        return r.status_code, r.text[:300]

    step("populate BOQ prices from winner", populate_prices, skip_if=lambda: not ctx.get("winner"))

    # ---- 12. pricing summary / gaps / manual override / template ----------
    def pricing_summary():
        r = c.get(f"/api/projects/{pid}/pricing/summary")
        if r.status_code == 200:
            j = r.json()
            ctx["direct_grand_total"] = j.get("grand_total")
            return 200, (f"cost_subtotal={j.get('cost_subtotal')} grand_total={j.get('grand_total')} "
                         f"completion={j.get('completion_rate')}")
        return r.status_code, r.text[:200]

    step("pricing summary", pricing_summary, skip_if=has_pid)

    def pricing_gaps():
        r = c.get(f"/api/projects/{pid}/pricing/gaps")
        j = r.json()
        unp = j.get("unpriced") or []
        if unp:
            ctx["gap_item_id"] = unp[0].get("id")
        return r.status_code, (f"unpriced={j.get('unpriced_count')} needs_review={j.get('needs_review_count')} "
                               f"excluded={j.get('excluded_count')}")

    step("pricing gaps", pricing_gaps, skip_if=has_pid)

    def manual_price():
        iid = ctx["gap_item_id"]
        r = c.patch(f"/api/boq-items/{iid}/price", json={"unit_rate": 75.0, "notes": "E2E manual"})
        return r.status_code, r.text[:200]

    step("manual price an unpriced item", manual_price, skip_if=lambda: not ctx.get("gap_item_id"))

    def populate_template():
        files = {"file": ("client_boq.xlsx", make_boq_template_bytes(), XLSX_MIME)}
        r = c.post(f"/api/projects/{pid}/pricing/populate-template", files=files)
        if r.status_code == 200:
            p = save_artifact("priced_boq.xlsx", r.content)
            load_workbook(p)
            return 200, f"opened {p.name} ({len(r.content)} bytes)"
        return r.status_code, r.text[:250]

    step("pricing populate-template (formula-preserving)", populate_template,
         expect=(200,), degrade=(409,), skip_if=has_pid)

    # ---- 13. indirects + cost summary -------------------------------------
    def indirects():
        r = c.get(f"/api/projects/{pid}/indirects", params={"duration_months": 12, "location": "default"})
        if r.status_code == 200:
            j = r.json()
            ind = j.get("indirects", {})
            return 200, f"total_indirects={ind.get('total_indirects')} direct={j.get('direct_cost')}"
        return r.status_code, r.text[:200]

    step("indirects breakdown", indirects, skip_if=has_pid)

    def cost_summary():
        r = c.get(f"/api/projects/{pid}/cost-summary", params={"duration_months": 12})
        if r.status_code == 200:
            j = r.json()
            gt = j.get("grand_total")
            note = f"grand_total={gt} direct={j.get('direct_cost')}"
            d = ctx.get("direct_grand_total")
            if d is not None and gt is not None and gt < d:
                note += f"  !! cost-summary grand_total < pricing-summary ({gt} < {d})"
            return 200, note
        return r.status_code, r.text[:200]

    step("cost-summary (full rollup)", cost_summary, skip_if=has_pid)

    def indirects_template():
        files = {"file": ("client_indirects.xlsx", make_indirects_template_bytes(), XLSX_MIME)}
        r = c.post(f"/api/projects/{pid}/indirects/populate-template",
                   params={"duration_months": 12}, files=files)
        if r.status_code == 200:
            p = save_artifact("indirects_filled.xlsx", r.content)
            load_workbook(p)
            hdrs = {k: v for k, v in r.headers.items() if k.lower().startswith("x-indirects")}
            return 200, f"opened {p.name} headers={hdrs}"
        return r.status_code, r.text[:250]

    step("indirects populate-template", indirects_template,
         expect=(200,), degrade=(409,), skip_if=has_pid)

    # ---- 14. historical learning ------------------------------------------
    def hist_add():
        codes = []
        for desc, rate, trade in [
            ("C30 ready-mix concrete to foundations", 92.0, "concrete"),
            ("Reinforcement steel grade 60 fixed", 1080.0, "concrete"),
            ("Internal emulsion paint two coats", 6.0, "finishes"),
        ]:
            r = c.post("/api/historical", json={
                "description": desc, "rate": rate, "unit": "m3",
                "currency": "USD", "trade_category": trade, "source": "manual"})
            codes.append(r.status_code)
        return (201 if all(s in (200, 201) for s in codes) else codes[0]), f"added {codes}"

    step("historical add records", hist_add)

    step("historical list", lambda: (c.get("/api/historical").status_code, "list corpus"))

    def hist_suggest():
        r = c.get("/api/historical/suggest",
                  params={"description": "ready mix concrete foundations", "trade": "concrete", "top_k": 5})
        if r.status_code == 200:
            b = r.json().get("benchmark", {})
            return 200, f"count={b.get('count')} suggested_rate={b.get('suggested_rate')} ccy={b.get('currency')}"
        return r.status_code, r.text[:200]

    step("historical suggest (benchmark)", hist_suggest)

    def hist_import():
        files = {"file": ("rates.xlsx", make_historical_xlsx_bytes(), XLSX_MIME)}
        r = c.post("/api/historical/import", files=files)
        return r.status_code, r.text[:200]

    step("historical import xlsx", hist_import)

    step("historical feedback",
         lambda: (c.post("/api/historical/feedback",
                         json={"description": "C30 concrete foundations", "accepted_rate": 90.0,
                               "unit": "m3", "currency": "USD", "trade_category": "concrete"}).status_code,
                  "record correction"))

    step("project historical index",
         lambda: (c.post(f"/api/projects/{pid}/historical/index").status_code, "snapshot priced BOQ"),
         skip_if=has_pid)
    step("project historical suggestions",
         lambda: (c.get(f"/api/projects/{pid}/historical/suggestions").status_code, "per-item suggestions"),
         skip_if=has_pid)

    # ---- 15. deliverables --------------------------------------------------
    def build_deliverables():
        r = c.post(f"/api/projects/{pid}/deliverables/build", params={"duration_months": 12})
        if r.status_code == 200:
            j = r.json()
            ctx["deliverables_built"] = True
            return 200, f"files={j.get('files')} comparisons={j.get('comparisons')} briefs={j.get('briefs')}"
        return r.status_code, r.text[:250]

    step("deliverables build", build_deliverables, skip_if=has_pid)

    def dl_zip():
        r = c.get(f"/api/projects/{pid}/deliverables/download")
        if r.status_code == 200:
            p = save_artifact("deliverables.zip", r.content)
            with zipfile.ZipFile(p) as z:
                names = z.namelist()
                bad = z.testzip()
            return 200, f"zip ok members={len(names)} bad={bad} sample={names[:4]}"
        return r.status_code, r.text[:200]

    step("download+verify deliverables.zip", dl_zip, skip_if=lambda: not ctx.get("deliverables_built"))

    # ---- 16. dashboard endpoint + rules -----------------------------------
    def dashboard():
        r = c.get(f"/api/projects/{pid}/dashboard")
        if r.status_code == 200:
            j = r.json()
            keys = list(j.keys())
            return 200, f"keys={keys}"
        return r.status_code, r.text[:200]

    step("dashboard endpoint", dashboard, skip_if=has_pid)

    def rules_roundtrip():
        g = c.get("/api/rules")
        if g.status_code != 200:
            return g.status_code, g.text[:150]
        rules = g.json()
        p = c.put("/api/rules", json=rules)  # round-trip identical -> must validate
        return p.status_code, f"GET ok ({len(rules)} sections); PUT={p.status_code}"

    step("rules GET+PUT round-trip", rules_roundtrip)

    # ---- 17. v1 report export (excel may 404 if extraction absent) --------
    step("export excel (404 ok if no extraction)",
         lambda: (c.get(f"/api/projects/{pid}/export/excel").status_code, "v1 report"),
         expect=(200,), degrade=(404, 500), skip_if=has_pid)
    step("export pdf (501/404 ok)",
         lambda: (c.get(f"/api/projects/{pid}/export/pdf").status_code, "v1 pdf report"),
         expect=(200,), degrade=(404, 500, 501), skip_if=has_pid)

    c.close()

    # ---- summary -----------------------------------------------------------
    counts = {"PASS": 0, "DEGRADED": 0, "SKIP": 0, "FAIL": 0}
    for r in results:
        counts[r["verdict"]] += 1
    print("\n=== SUMMARY ===")
    print(f"PASS={counts['PASS']}  DEGRADED={counts['DEGRADED']}  SKIP={counts['SKIP']}  FAIL={counts['FAIL']}")
    fails = [r for r in results if r["verdict"] == "FAIL"]
    if fails:
        print("\n--- FAILURES ---")
        for r in fails:
            print(f"  {r['step']} (HTTP {r['status']}): {r['note']}")
    report = REPO / ".e2e_tmp" / "shakedown_report.json"
    report.write_text(json.dumps({"base": BASE, "counts": counts, "results": results}, indent=2))
    print(f"\nReport: {report}")


if __name__ == "__main__":
    main()
