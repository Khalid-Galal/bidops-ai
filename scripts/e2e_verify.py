"""Phase 0 end-to-end verification driver for BidOps AI v1.

Exercises the real running server (http://127.0.0.1:8000) against the synthetic
bilingual tender set:  create project -> upload DOCX+XLSX -> poll parsing ->
search (3 modes) -> extract summary -> generate checklist -> export Excel/PDF.

Also reproduces the Excel "checked"-state bug: PATCH a checklist item to checked,
re-export Excel, and report whether the Status column reflects it (pre-fix => no).

Writes artifacts + a JSON report to verification_out/.
Run (background recommended):  .venv/Scripts/python.exe scripts/e2e_verify.py
"""

from __future__ import annotations

import json
import sys
import time
from pathlib import Path

import httpx

BASE = "http://127.0.0.1:8000"
ROOT = Path(__file__).resolve().parent.parent
TENDER = ROOT / "sample_tender"
OUT = ROOT / "verification_out"
OUT.mkdir(parents=True, exist_ok=True)

report: dict = {"steps": [], "ok": True}


def log(msg: str):
    print(msg, flush=True)


def step(name: str, ok: bool, detail):
    report["steps"].append({"name": name, "ok": ok, "detail": detail})
    if not ok:
        report["ok"] = False
    log(f"[{'PASS' if ok else 'FAIL'}] {name}: {detail}")


def poll(client: httpx.Client, fn, ready, timeout: float, interval: float, label: str):
    """Poll fn() until ready(result) True or timeout. Resilient to transient
    request errors (server busy during model load) — treats them as not-ready."""
    start = time.time()
    last = None
    while time.time() - start < timeout:
        try:
            last = fn()
        except Exception as e:  # noqa: BLE001 - server may be busy/blocked briefly
            log(f"    .. {label} request error ({type(e).__name__}), retrying")
            time.sleep(interval)
            continue
        if ready(last):
            return last
        elapsed = int(time.time() - start)
        log(f"    .. {label} waiting ({elapsed}s) -> {_short(last)}")
        time.sleep(interval)
    return last


def _short(x):
    s = json.dumps(x, default=str) if not isinstance(x, str) else x
    return s[:160]


def main():
    with httpx.Client(timeout=httpx.Timeout(120.0)) as client:
        # 1. health
        r = client.get(f"{BASE}/health")
        step("health", r.status_code == 200, r.text.strip())

        # 2. create project
        r = client.post(f"{BASE}/api/projects", json={
            "name": "New Cairo Medical Center (E2E verify)",
            "description": "Phase 0 synthetic bilingual tender verification",
        })
        step("create_project", r.status_code == 201, f"HTTP {r.status_code}")
        if r.status_code != 201:
            return finish()
        pid = r.json()["id"]
        report["project_id"] = pid

        # 3. upload docs
        files = []
        for fp in [TENDER / "ITT_Tender_Document.docx", TENDER / "BOQ_Bill_of_Quantities.xlsx"]:
            files.append(("files", (fp.name, fp.read_bytes())))
        r = client.post(f"{BASE}/api/projects/{pid}/upload", files=files)
        step("upload", r.status_code in (200, 202), f"HTTP {r.status_code} {_short(r.text)}")

        # 4. poll parsing (docling first-run downloads models -> generous timeout)
        def docs():
            return client.get(f"{BASE}/api/projects/{pid}/documents").json()
        def parsed(d):
            return bool(d) and all(x.get("status") in ("completed", "failed") for x in d)
        d = poll(client, docs, parsed, timeout=900, interval=8, label="parsing")
        statuses = {x.get("filename"): x.get("status") for x in (d or [])}
        all_done = parsed(d)
        any_completed = any(s == "completed" for s in statuses.values())
        step("parse_documents", all_done and any_completed, statuses)

        # 5. search (3 modes) — verify retrieval works
        search_res = {}
        for mode in ("keyword", "semantic", "hybrid"):
            try:
                r = client.get(f"{BASE}/api/projects/{pid}/search",
                               params={"q": "tender bond performance guarantee", "mode": mode, "limit": 5})
                n = r.json().get("total_results", 0) if r.status_code == 200 else -1
                search_res[mode] = n
            except Exception as e:
                search_res[mode] = f"err:{e}"
        step("search_modes", all(isinstance(v, int) and v >= 0 for v in search_res.values()), search_res)

        # 6. extract summary (POST runs the pipeline INLINE and returns the result)
        try:
            r = client.post(f"{BASE}/api/projects/{pid}/extract", timeout=900)
            log(f"    POST /extract -> HTTP {r.status_code}")
            s = r.json() if r.status_code == 200 else None
        except Exception as e:  # noqa: BLE001
            log(f"    POST /extract error: {e}; falling back to GET poll")
            s = None
        def summ():
            return client.get(f"{BASE}/api/projects/{pid}/extract").json()
        def summ_done(x):
            return isinstance(x, dict) and x.get("status") in ("completed", "failed")
        if not summ_done(s):
            s = poll(client, summ, summ_done, timeout=900, interval=6, label="summary")
        summary = s.get("summary") if isinstance(s, dict) else None
        fields = summary.get("fields") if isinstance(summary, dict) else None
        # summary schema may nest fields differently; capture whatever is present
        (OUT / "summary.json").write_text(json.dumps(s, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        filled = 0
        sample_fields = {}
        if isinstance(summary, dict):
            for k, v in summary.items():
                if isinstance(v, dict) and "value" in v:
                    if v.get("value"):
                        filled += 1
                        if len(sample_fields) < 5:
                            sample_fields[k] = {"value": str(v.get("value"))[:80],
                                                 "confidence": v.get("confidence"),
                                                 "level": v.get("confidence_level"),
                                                 "citations": len(v.get("citations") or [])}
        step("extract_summary",
             summ_done(s) and s.get("status") == "completed" and filled > 0,
             {"status": s.get("status") if isinstance(s, dict) else None,
              "fields_filled": filled, "sample": sample_fields})

        # 7. checklist (POST runs INLINE)
        try:
            r = client.post(f"{BASE}/api/projects/{pid}/checklist", timeout=1200)
            log(f"    POST /checklist -> HTTP {r.status_code}")
            c = r.json() if r.status_code == 200 else None
        except Exception as e:  # noqa: BLE001
            log(f"    POST /checklist error: {e}; falling back to GET poll")
            c = None
        def chk():
            return client.get(f"{BASE}/api/projects/{pid}/checklist").json()
        def chk_done(x):
            return isinstance(x, dict) and x.get("status") in ("completed", "failed")
        if not chk_done(c):
            c = poll(client, chk, chk_done, timeout=1200, interval=8, label="checklist")
        (OUT / "checklist.json").write_text(json.dumps(c, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
        cl = c.get("checklist") if isinstance(c, dict) else None
        counts = {}
        if isinstance(cl, dict):
            counts = {k: len(v) for k, v in cl.items() if isinstance(v, list)}
        total = c.get("total_requirements") if isinstance(c, dict) else 0
        step("extract_checklist",
             chk_done(c) and c.get("status") == "completed" and (total or 0) > 0,
             {"status": c.get("status") if isinstance(c, dict) else None,
              "total": total, "by_list": counts})

        # 8. export excel (should work)
        r = client.get(f"{BASE}/api/projects/{pid}/export/excel")
        xlsx_ok = r.status_code == 200 and len(r.content) > 0
        if xlsx_ok:
            (OUT / "export.xlsx").write_bytes(r.content)
        step("export_excel", xlsx_ok, f"HTTP {r.status_code} bytes={len(r.content)}")

        # 9. export pdf (expect graceful 501 — weasyprint/Pango absent on Windows)
        r = client.get(f"{BASE}/api/projects/{pid}/export/pdf")
        step("export_pdf_graceful", r.status_code in (200, 501),
             f"HTTP {r.status_code} (501 expected = graceful WeasyPrint degradation)")

        # 10. reproduce Excel "checked" bug: PATCH one requirement to checked, re-export
        bug_detail = {"patched": False, "excel_reflects_checked": None}
        if isinstance(cl, dict) and cl.get("requirements"):
            r = client.patch(f"{BASE}/api/projects/{pid}/checklist/items",
                             json={"category": "requirements", "index": 0, "updates": {"checked": True}})
            bug_detail["patched"] = r.status_code in (200, 204)
            bug_detail["patch_http"] = r.status_code
            # re-export and inspect the Status column of row 1
            r2 = client.get(f"{BASE}/api/projects/{pid}/export/excel")
            if r2.status_code == 200:
                p = OUT / "export_after_patch.xlsx"
                p.write_bytes(r2.content)
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(p)
                    ws = wb["Requirements Checklist"] if "Requirements Checklist" in wb.sheetnames else wb[wb.sheetnames[-1]]
                    # Status is last column; find first data row
                    statuses_col = [row[-1] for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0]]
                    bug_detail["first_rows_status"] = statuses_col[:5]
                    bug_detail["excel_reflects_checked"] = any(s == "Checked" for s in statuses_col)
                except Exception as e:
                    bug_detail["excel_read_err"] = str(e)
        # This step PASSES if we successfully observed the behavior (bug is expected pre-fix)
        step("excel_checked_bug_probe", bug_detail.get("patched", False), bug_detail)

    finish()


def finish():
    (OUT / "report.json").write_text(json.dumps(report, indent=2, ensure_ascii=False, default=str), encoding="utf-8")
    log("\n==== E2E REPORT ====")
    for s in report["steps"]:
        log(f"  [{'PASS' if s['ok'] else 'FAIL'}] {s['name']}")
    log(f"OVERALL: {'PASS' if report['ok'] else 'FAIL (see report.json)'}")
    sys.exit(0 if report["ok"] else 2)


if __name__ == "__main__":
    main()
