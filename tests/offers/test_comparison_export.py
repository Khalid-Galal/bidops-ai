import openpyxl

from app.services.offer.comparison_export import export_comparison_excel

COMPARISON = {
    "package_id": 1,
    "package_name": "HVAC Works",
    "total_offers": 2,
    "evaluated_offers": 2,
    "currency": "USD",
    "price_min": 100.0,
    "price_max": 150.0,
    "price_avg": 125.0,
    "offers": [
        {"offer_id": 1, "supplier_id": 1, "supplier_name": "CoolAir", "total_price": 100.0,
         "currency": "USD", "validity_days": 90, "delivery_weeks": 4,
         "delivery_terms": "DDP Site", "payment_terms": "Net 30",
         "vat_included": True, "vat_amount": 12.5,
         "commercial_score": 100.0, "technical_score": 50.0, "overall_score": 75.0, "rank": 1,
         "status": "evaluated", "exclusions": [], "deviations": ["Alternate chiller brand"],
         "exclusions_count": 0, "deviations_count": 1},
        {"offer_id": 2, "supplier_id": 2, "supplier_name": "HawaCo", "total_price": 150.0,
         "currency": "USD", "validity_days": 60, "delivery_weeks": 8,
         "delivery_terms": None, "payment_terms": "Net 60",
         "vat_included": False, "vat_amount": None,
         "commercial_score": 66.7, "technical_score": 50.0, "overall_score": 55.8, "rank": 2,
         "status": "evaluated", "exclusions": ["excludes copper piping"], "deviations": [],
         "exclusions_count": 1, "deviations_count": 0},
    ],
    "warnings": [],
}


def test_export_writes_matrix(tmp_path):
    out = export_comparison_excel(COMPARISON, str(tmp_path / "cmp.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.title == "Offer Comparison"
    # header row contains the key columns
    header = [c.value for c in ws[5]]
    assert "Rank" in header and "Supplier" in header and "Overall Score" in header
    assert "VAT" in header and "Delivery Terms" in header
    # first data row is the rank-1 supplier
    row6 = [c.value for c in ws[6]]
    assert "CoolAir" in row6
    assert 75.0 in row6
    assert "DDP Site" in row6
    assert "Incl (12.5)" in row6


def test_export_writes_exclusions_deviations_sheet(tmp_path):
    out = export_comparison_excel(COMPARISON, str(tmp_path / "cmp.xlsx"))
    wb = openpyxl.load_workbook(out)
    assert "Exclusions & Deviations" in wb.sheetnames
    ws = wb["Exclusions & Deviations"]
    rows = [tuple(r) for r in ws.iter_rows(min_row=2, values_only=True)]
    assert ("CoolAir", "Deviation", "Alternate chiller brand") in rows
    assert ("HawaCo", "Exclusion", "excludes copper piping") in rows


def test_export_warning_banner_when_present(tmp_path):
    data = dict(COMPARISON)
    data["warnings"] = ["Offers use different currencies (EGP, USD) -- ..."]
    out = export_comparison_excel(data, str(tmp_path / "cmp.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert "WARNING" in str(ws["A4"].value)
    # header row must stay at row 5 regardless of the warning banner
    assert ws["A5"].value == "Rank"
