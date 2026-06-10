from pathlib import Path

from openpyxl import load_workbook


async def _seed(db, tmp_path):
    """Project + package + 2 BOQ items + a real doc file linked to the package."""
    from app.models.project import Project
    from app.models.document import Document
    from app.models.package import Package, PackageDocument
    from app.models.boq import BOQItem

    project = Project(name="New Cairo Medical Center")
    db.add(project)
    await db.flush()

    # a real file on disk for the linked document
    src = tmp_path / "concrete_spec.pdf"
    src.write_bytes(b"%PDF-1.4 fake spec content")
    doc = Document(
        project_id=project.id, filename="concrete_spec.pdf",
        file_path=str(src), file_type=".pdf", file_size=src.stat().st_size,
    )
    db.add(doc)
    await db.flush()

    pkg = Package(project_id=project.id, name="Concrete Works",
                  code="PKG-P0001-CON-001", trade_category="concrete", total_items=2)
    db.add(pkg)
    await db.flush()
    db.add_all([
        BOQItem(project_id=project.id, package_id=pkg.id, line_number="2.1",
                description="Reinforced concrete C35/45", unit="m3", quantity=5400,
                client_row_index=1, trade_category="concrete"),
        BOQItem(project_id=project.id, package_id=pkg.id, line_number="2.2",
                description="Reinforcement steel B500B", unit="ton", quantity=4900,
                client_row_index=2, trade_category="concrete"),
        BOQItem(project_id=project.id, package_id=None, line_number="9.9",
                description="Unassigned item", unit="no", quantity=1,
                client_row_index=3, trade_category=None),
    ])
    db.add(PackageDocument(package_id=pkg.id, document_id=doc.id,
                           relevance_score=0.9, relevance_reason="match",
                           excerpt="Concrete grade C35/45"))
    await db.commit()
    return project.id, pkg.id


async def test_export_creates_folders_boq_subset_docs_brief_register(db_session, tmp_path):
    from app.services.packaging.package_exporter import PackageExporter

    out = tmp_path / "packages_out"
    pid, pkg_id = await _seed(db_session, tmp_path)

    result = await PackageExporter(output_root=out).export_project(db_session, pid)

    assert result["packages_exported"] == 1
    pkg_dir = Path(result["packages"][0]["folder_path"])
    assert pkg_dir.is_dir()
    # folder tree
    for sub in ("BOQ", "Documents", "Offers", "Clarifications"):
        assert (pkg_dir / sub).is_dir()
    # BOQ subset xlsx with the 2 package items (not the unassigned one)
    boq_files = list((pkg_dir / "BOQ").glob("*.xlsx"))
    assert boq_files, "BOQ subset xlsx missing"
    wb = load_workbook(boq_files[0])
    ws = wb.active
    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    assert len(data_rows) == 2
    # linked document copied + manifest written
    assert (pkg_dir / "Documents" / "concrete_spec.pdf").exists()
    assert (pkg_dir / "Documents" / "linked_manifest.txt").exists()
    # brief HTML always written
    assert (pkg_dir / "Package_Brief.html").exists()
    brief_html = (pkg_dir / "Package_Brief.html").read_text(encoding="utf-8")
    assert "Concrete Works" in brief_html and "C35/45" in brief_html
    # master register at project root
    register = Path(result["register_path"])
    assert register.exists()
    rwb = load_workbook(register)
    rws = rwb.active
    reg_rows = [r for r in rws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    assert any("PKG-P0001-CON-001" in str(r) for r in reg_rows)
    # package paths persisted
    from app.models.package import Package
    refreshed = await db_session.get(Package, pkg_id)
    assert refreshed.folder_path == str(pkg_dir)


async def test_export_missing_linked_file_is_graceful(db_session, tmp_path):
    from app.models.project import Project
    from app.models.document import Document
    from app.models.package import Package, PackageDocument
    from app.services.packaging.package_exporter import PackageExporter

    project = Project(name="P")
    db_session.add(project)
    await db_session.flush()
    doc = Document(project_id=project.id, filename="gone.pdf",
                   file_path=str(tmp_path / "does_not_exist.pdf"),
                   file_type=".pdf", file_size=0)
    db_session.add(doc)
    await db_session.flush()
    pkg = Package(project_id=project.id, name="MEP", code="PKG-P-MEP-001",
                  trade_category="mep", total_items=0)
    db_session.add(pkg)
    await db_session.flush()
    db_session.add(PackageDocument(package_id=pkg.id, document_id=doc.id))
    await db_session.commit()

    result = await PackageExporter(output_root=tmp_path / "o").export_project(db_session, project.id)
    # must not raise; the missing file is noted in the manifest, brief still written
    assert result["packages_exported"] == 1
    pkg_dir = Path(result["packages"][0]["folder_path"])
    assert not (pkg_dir / "Documents" / "gone.pdf").exists()
    manifest = (pkg_dir / "Documents" / "linked_manifest.txt").read_text(encoding="utf-8")
    assert "gone.pdf" in manifest and "MISSING" in manifest.upper()


async def test_exporter_excludes_superseded(db_session, tmp_path):
    """A superseded document linked to a package must not appear in the export
    (no copied file, absent from the manifest and brief)."""
    from app.models.project import Project
    from app.models.document import Document
    from app.models.package import Package, PackageDocument
    from app.services.packaging.package_exporter import PackageExporter

    project = Project(name="P")
    db_session.add(project)
    await db_session.flush()

    live_src = tmp_path / "live.pdf"
    live_src.write_bytes(b"%PDF live")
    stale_src = tmp_path / "stale.pdf"
    stale_src.write_bytes(b"%PDF stale")
    live = Document(project_id=project.id, filename="live.pdf",
                    file_path=str(live_src), file_type=".pdf", file_size=8)
    stale = Document(project_id=project.id, filename="stale.pdf",
                     file_path=str(stale_src), file_type=".pdf", file_size=9,
                     is_superseded=True,
                     supersede_reason="auto:superseded by newer revision")
    db_session.add_all([live, stale])
    await db_session.flush()
    pkg = Package(project_id=project.id, name="MEP", code="PKG-1",
                  trade_category="mep", total_items=0)
    db_session.add(pkg)
    await db_session.flush()
    db_session.add_all([
        PackageDocument(package_id=pkg.id, document_id=live.id, relevance_score=0.9),
        PackageDocument(package_id=pkg.id, document_id=stale.id, relevance_score=0.95),
    ])
    await db_session.commit()

    result = await PackageExporter(output_root=tmp_path / "o").export_project(
        db_session, project.id
    )
    pkg_dir = Path(result["packages"][0]["folder_path"])
    assert (pkg_dir / "Documents" / "live.pdf").exists()
    assert not (pkg_dir / "Documents" / "stale.pdf").exists()
    manifest = (pkg_dir / "Documents" / "linked_manifest.txt").read_text(encoding="utf-8")
    assert "live.pdf" in manifest
    assert "stale.pdf" not in manifest
    brief = (pkg_dir / "Package_Brief.html").read_text(encoding="utf-8")
    assert "stale.pdf" not in brief


def test_safe_name_blocks_traversal():
    from app.services.packaging.package_exporter import _safe_name
    for evil in ("..", "../..", "  ../  ", ".", "/etc/passwd", "..\\..\\x"):
        out = _safe_name(evil)
        assert out not in ("", ".", "..")
        assert "/" not in out and "\\" not in out
        assert out != ".."
