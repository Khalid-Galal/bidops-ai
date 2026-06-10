import openpyxl
import pytest

from app.services.indirects.indirects_template import (
    detect_columns,
    populate_indirects_template,
)


def _make_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indirects"
    ws.append(["Item", "Description", "Amount"])           # row 1 header
    ws.append([1, "Site Supervision", None])               # row 2
    ws.append([2, "Temporary Works", None])                # row 3
    ws.append([3, "Project Manager", None])                # row 4
    ws.append([4, "Total Indirects", "=SUM(C2:C4)"])       # row 5 formula
    wb.save(path)
    return str(path)


def test_detect_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Description", "Amount"])
    label_col, amount_col, header_row = detect_columns(ws)
    assert amount_col == 3
    assert label_col == 2
    assert header_row == 1


def test_header_row_not_overwritten(tmp_path):
    """A header like 'Indirects | Amount' fuzzy-matches component names; the
    populate loop must start BELOW the header, never writing into it."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Indirects", "Amount"])        # row 1: header
    ws.append(["Total Indirects", None])      # row 2: data
    src = tmp_path / "hdr.xlsx"
    wb.save(src)
    result = populate_indirects_template(
        str(src), str(tmp_path / "out.xlsx"), {"total_indirects": 510.0}
    )
    assert result["written"] == 1
    out = openpyxl.load_workbook(tmp_path / "out.xlsx")
    ws2 = out[out.sheetnames[0]]
    assert ws2.cell(row=1, column=2).value == "Amount"  # header intact
    assert ws2.cell(row=2, column=2).value == 510.0


def test_specific_label_wins_over_generic(tmp_path):
    """Global best assignment: 'Safety Officer' (exact, later row) must not be
    stolen by the generic 'Safety' row above it."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Description", "Amount"])
    ws.append(["Safety", None])           # row 2
    ws.append(["Safety Officer", None])   # row 3
    src = tmp_path / "spec.xlsx"
    wb.save(src)
    result = populate_indirects_template(
        str(src), str(tmp_path / "out.xlsx"),
        {"safety": 100.0, "safety_officer": 30000.0},
    )
    assert result["written"] == 2
    out = openpyxl.load_workbook(tmp_path / "out.xlsx")
    ws2 = out[out.sheetnames[0]]
    assert ws2.cell(row=2, column=2).value == 100.0
    assert ws2.cell(row=3, column=2).value == 30000.0

    # Asymmetric case: only safety_officer present -> must land on ITS row (3).
    result = populate_indirects_template(
        str(src), str(tmp_path / "out2.xlsx"), {"safety_officer": 30000.0}
    )
    assert result["written"] == 1
    out = openpyxl.load_workbook(tmp_path / "out2.xlsx")
    ws3 = out[out.sheetnames[0]]
    assert ws3.cell(row=2, column=2).value is None
    assert ws3.cell(row=3, column=2).value == 30000.0


def test_detect_amount_beats_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Description", "Total", "Amount"])
    label_col, amount_col, header_row = detect_columns(ws)
    assert amount_col == 3  # "amount" outranks "total" regardless of column order
    assert label_col == 1
    assert header_row == 1


def test_header_below_row_one(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ACME Constructions - Indirect Costs Template"])  # row 1: title junk
    ws.append([])                                                # row 2: blank
    ws.append(["Description", "Amount"])                         # row 3: real header
    ws.append(["Site Supervision", None])                        # row 4: data
    src = tmp_path / "deep.xlsx"
    wb.save(src)
    label_col, amount_col, header_row = detect_columns(ws)
    assert (label_col, amount_col, header_row) == (1, 2, 3)
    result = populate_indirects_template(
        str(src), str(tmp_path / "out.xlsx"), {"site_supervision": 180.0}
    )
    assert result["written"] == 1
    out = openpyxl.load_workbook(tmp_path / "out.xlsx")
    ws2 = out[out.sheetnames[0]]
    assert ws2.cell(row=4, column=2).value == 180.0


def test_picks_indirects_sheet(tmp_path):
    wb = openpyxl.Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Some summary text"])
    ind = wb.create_sheet("Indirects")
    ind.append(["Description", "Amount"])
    ind.append(["Site Supervision", None])
    src = tmp_path / "multi.xlsx"
    wb.save(src)
    result = populate_indirects_template(
        str(src), str(tmp_path / "out.xlsx"), {"site_supervision": 180.0}
    )
    assert result["written"] == 1
    out = openpyxl.load_workbook(tmp_path / "out.xlsx")
    assert out["Indirects"].cell(row=2, column=2).value == 180.0
    assert all(c.value is None for row in out["Summary"].iter_rows() for c in row
               if c.value == 180.0)  # Summary untouched by the write
    assert out["Summary"].cell(row=1, column=1).value == "Some summary text"


def test_populate_matches_labels_and_preserves_formula(tmp_path):
    src = _make_template(tmp_path / "ind.xlsx")
    out = str(tmp_path / "out.xlsx")
    components = {
        "site_supervision": 660.0,
        "temporary_works": 440.0,
        "project_manager": 30000.0,
        "total_indirects": 31100.0,  # the Total row holds a formula -> skipped
    }
    result = populate_indirects_template(src, out, components)
    assert result["written"] == 3
    assert result["skipped_formula"] == 1
    assert result["unmatched_components"] == []
    wb = openpyxl.load_workbook(out)
    ws = wb["Indirects"]
    assert ws.cell(row=2, column=3).value == 660.0
    assert ws.cell(row=3, column=3).value == 440.0
    assert ws.cell(row=4, column=3).value == 30000.0
    assert ws.cell(row=5, column=3).value == "=SUM(C2:C4)"  # formula intact


def test_populate_reports_unmatched(tmp_path):
    src = _make_template(tmp_path / "i.xlsx")
    out = str(tmp_path / "o.xlsx")
    result = populate_indirects_template(src, out, {"helicopter_rental": 5.0})
    assert result["written"] == 0
    assert result["unmatched_components"] == ["helicopter_rental"]


def test_populate_explicit_columns(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Site Supervision", None])   # no header row at all
    p = tmp_path / "nohdr.xlsx"
    wb.save(p)
    result = populate_indirects_template(
        str(p), str(tmp_path / "o2.xlsx"), {"site_supervision": 99.0},
        amount_column=2, label_column=1,
    )
    assert result["written"] == 1


def test_populate_raises_without_amount_column(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Just", "Words"])
    p = tmp_path / "bad.xlsx"
    wb.save(p)
    with pytest.raises(ValueError):
        populate_indirects_template(str(p), str(tmp_path / "o3.xlsx"), {"x": 1.0})


import io

import httpx
import pytest_asyncio
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine


@pytest_asyncio.fixture
async def ind_client():
    from app.database import get_db
    from app.main import app
    from app.models import Base
    from app.models.boq import BOQItem
    from app.models.project import Project

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    async with factory() as seed:
        project = Project(name="Metro")
        seed.add(project)
        await seed.flush()
        seed.add(BOQItem(project_id=project.id, line_number="1", description="AC",
                         unit="no", quantity=5, client_row_index=2, trade_category="mep",
                         unit_rate=1200, total_price=6000, currency="USD"))
        await seed.commit()
        pid = project.id

    async def _override():
        async with factory() as session:
            yield session

    app.dependency_overrides[get_db] = _override
    client = httpx.AsyncClient(transport=httpx.ASGITransport(app=app), base_url="http://test")
    yield client, pid
    await client.aclose()
    app.dependency_overrides.clear()
    await engine.dispose()


def _template_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Indirects"
    ws.append(["Item", "Description", "Amount"])
    ws.append([1, "Site Supervision", None])
    ws.append([2, "Total Indirects", "=SUM(C2:C2)"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_populate_template_endpoint(ind_client):
    client, pid = ind_client
    async with client as c:
        r = await c.post(
            f"/api/projects/{pid}/indirects/populate-template",
            files={"file": ("ind.xlsx", _template_bytes(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200, r.text
        assert r.headers["X-Indirects-Written"] == "1"
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Indirects"]
        # site_supervision = 0.03 * 6000 = 180.0 (default rules)
        assert ws.cell(row=2, column=3).value == 180.0
        assert ws.cell(row=3, column=3).value == "=SUM(C2:C2)"  # formula intact


async def test_populate_template_409_nothing_matched(ind_client):
    client, pid = ind_client
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Description", "Amount"])
    ws.append(["Quantum Flux", None])  # matches no indirect component
    buf = io.BytesIO()
    wb.save(buf)
    async with client as c:
        r = await c.post(
            f"/api/projects/{pid}/indirects/populate-template",
            files={"file": ("ind.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 409
    assert "No template rows matched" in r.json()["detail"]


async def test_populate_template_oversized_413(ind_client, monkeypatch):
    import app.api.indirects as ind_api

    monkeypatch.setattr(ind_api, "_MAX_UPLOAD_BYTES", 10)
    client, pid = ind_client
    async with client as c:
        r = await c.post(
            f"/api/projects/{pid}/indirects/populate-template",
            files={"file": ("ind.xlsx", _template_bytes(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 413


async def test_populate_template_409_no_components(ind_client):
    """A project with no priced items and duration_months=0 has no indirect
    amounts at all -> 409 before the template is even opened."""
    client, _ = ind_client
    async with client as c:
        created = await c.post("/api/projects", json={"name": "Empty"})
        assert created.status_code == 201, created.text
        empty_pid = created.json()["id"]
        r = await c.post(
            f"/api/projects/{empty_pid}/indirects/populate-template",
            files={"file": ("ind.xlsx", _template_bytes(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 409
    assert "No indirect amounts" in r.json()["detail"]


async def test_populate_template_400_corrupt_xlsx(ind_client):
    client, pid = ind_client
    async with client as c:
        r = await c.post(
            f"/api/projects/{pid}/indirects/populate-template",
            files={"file": ("bad.xlsx", b"not a zip",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 400


async def test_populate_template_duration_months(ind_client, monkeypatch):
    import app.api.indirects as ind_api
    from app.schemas.rules import DurationBasedRole, RulesConfig
    from app.services.indirects.indirects_service import IndirectsService

    rules = RulesConfig()  # percentage_based defaults to {}
    rules.indirects.duration_based = {
        "project_manager": DurationBasedRole(monthly_rate=5000)
    }

    class _FakeRulesService:
        def load(self):
            return rules

    monkeypatch.setattr(
        ind_api, "IndirectsService",
        lambda: IndirectsService(rules_service=_FakeRulesService()),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Description", "Amount"])
    ws.append(["Project Manager", None])
    buf = io.BytesIO()
    wb.save(buf)

    client, pid = ind_client
    async with client as c:
        r = await c.post(
            f"/api/projects/{pid}/indirects/populate-template?duration_months=6",
            files={"file": ("ind.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert r.status_code == 200, r.text
        out = openpyxl.load_workbook(io.BytesIO(r.content))
        ws2 = out[out.sheetnames[0]]
        assert ws2.cell(row=2, column=2).value == 30000.0  # 6 * 5000


async def test_populate_template_rejects_non_xlsx(ind_client):
    client, pid = ind_client
    async with client as c:
        r = await c.post(
            f"/api/projects/{pid}/indirects/populate-template",
            files={"file": ("x.txt", b"nope", "text/plain")},
        )
    assert r.status_code == 400


async def test_populate_template_404_missing_project(ind_client):
    client, _ = ind_client
    async with client as c:
        r = await c.post(
            "/api/projects/999999/indirects/populate-template",
            files={"file": ("i.xlsx", _template_bytes(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert r.status_code == 404
