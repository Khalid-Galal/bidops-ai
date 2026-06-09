import io

import httpx
import openpyxl
import pytest_asyncio
from sqlalchemy import select
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine


@pytest_asyncio.fixture
async def pricing_client():
    from app.database import get_db
    from app.main import app
    from app.models import Base
    from app.models.base import OfferStatus
    from app.models.boq import BOQItem
    from app.models.package import Package
    from app.models.project import Project
    from app.models.supplier import Supplier, SupplierOffer

    engine = create_async_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    factory = async_sessionmaker(engine, expire_on_commit=False)

    async with factory() as seed:
        project = Project(name="Metro")
        seed.add(project)
        await seed.flush()
        package = Package(project_id=project.id, name="HVAC", code="PKG-001-MEP", trade_category="mep")
        seed.add(package)
        await seed.flush()
        seed.add_all([
            BOQItem(project_id=project.id, package_id=package.id, line_number="1",
                    description="Split AC unit supply & installation", unit="no",
                    quantity=5, client_row_index=2, trade_category="mep"),
            BOQItem(project_id=project.id, package_id=package.id, line_number="2",
                    description="VRF outdoor condensing unit", unit="no",
                    quantity=2, client_row_index=3, trade_category="mep"),
        ])
        supplier = Supplier(name="CoolAir", emails=[], trade_categories=["mep"])
        seed.add(supplier)
        await seed.flush()
        offer = SupplierOffer(
            package_id=package.id, supplier_id=supplier.id,
            status=OfferStatus.SELECTED.value, file_paths=[], currency="USD",
            line_items=[
                {"description": "Supply and install split AC unit", "rate": 1200},
                {"description": "VRF outdoor condensing unit", "rate": 8000},
            ],
        )
        seed.add(offer)
        await seed.commit()
        ids = {"project": project.id, "package": package.id, "offer": offer.id}
        boq = (await seed.execute(select(BOQItem).order_by(BOQItem.id))).scalars().all()
        ids["item0"] = boq[0].id

    async def _override():
        async with factory() as session:
            yield session

    app.dependency_overrides[get_db] = _override
    client = httpx.AsyncClient(transport=httpx.ASGITransport(app=app), base_url="http://test")
    yield client, ids
    await client.aclose()
    app.dependency_overrides.clear()
    await engine.dispose()


async def test_populate_then_summary_and_gaps(pricing_client):
    client, ids = pricing_client
    async with client as c:
        pop = await c.post(f"/api/offers/{ids['offer']}/populate-prices")
        assert pop.status_code == 200, pop.text
        assert pop.json()["items_populated"] == 2

        summ = await c.get(f"/api/projects/{ids['project']}/pricing/summary")
        assert summ.status_code == 200
        assert summ.json()["cost_subtotal"] == 22000.0
        assert summ.json()["grand_total"] == round(22000.0 * 1.26, 2)

        gaps = await c.get(f"/api/projects/{ids['project']}/pricing/gaps")
        assert gaps.status_code == 200
        assert gaps.json()["unpriced_count"] == 0


async def test_populate_prices_404_missing_offer(pricing_client):
    client, ids = pricing_client
    async with client as c:
        r = await c.post("/api/offers/999999/populate-prices")
    assert r.status_code == 404


async def test_manual_price_override(pricing_client):
    client, ids = pricing_client
    async with client as c:
        r = await c.patch(f"/api/boq-items/{ids['item0']}/price",
                          json={"unit_rate": 1500, "notes": "negotiated"})
        assert r.status_code == 200
        assert r.json()["unit_rate"] == 1500
        assert r.json()["total_price"] == 7500  # qty 5
        assert (await c.patch("/api/boq-items/999999/price", json={"unit_rate": 1})).status_code == 404


async def test_populate_template_download_preserves_formulas(pricing_client):
    client, ids = pricing_client
    # build a client template whose rows 2/3 match the seeded client_row_index
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append([1, "Split AC unit", "no", 5, None, "=D2*E2"])
    ws.append([2, "VRF unit", "no", 2, None, "=D3*E3"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    async with client as c:
        await c.post(f"/api/offers/{ids['offer']}/populate-prices")
        resp = await c.post(
            f"/api/projects/{ids['project']}/pricing/populate-template",
            files={"file": ("client.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    out = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws2 = out["BOQ"]
    assert ws2.cell(row=2, column=5).value == 1200.0  # rate written
    assert ws2.cell(row=2, column=6).value == "=D2*E2"  # formula preserved


async def test_populate_template_409_before_pricing(pricing_client):
    # No populate-prices first -> no priced rows -> 409.
    client, ids = pricing_client
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])
    ws.append([1, "Split AC unit", "no", 5, None, "=D2*E2"])
    buf = io.BytesIO()
    wb.save(buf)
    async with client as c:
        resp = await c.post(
            f"/api/projects/{ids['project']}/pricing/populate-template",
            files={"file": ("client.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 409, resp.text


async def test_populate_template_rejects_non_xlsx(pricing_client):
    client, ids = pricing_client
    async with client as c:
        await c.post(f"/api/offers/{ids['offer']}/populate-prices")
        resp = await c.post(
            f"/api/projects/{ids['project']}/pricing/populate-template",
            files={"file": ("x.txt", b"not a spreadsheet", "text/plain")},
        )
    assert resp.status_code == 400, resp.text


async def test_populate_template_400_no_rate_column(pricing_client):
    client, ids = pricing_client
    # valid xlsx but header has no rate-like column -> 400 (not 500)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Description", "Qty"])
    ws.append([1, "Split AC unit", 5])
    buf = io.BytesIO()
    wb.save(buf)
    async with client as c:
        await c.post(f"/api/offers/{ids['offer']}/populate-prices")
        resp = await c.post(
            f"/api/projects/{ids['project']}/pricing/populate-template",
            files={"file": ("client.xlsx", buf.getvalue(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
    assert resp.status_code == 400, resp.text
