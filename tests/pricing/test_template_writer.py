import openpyxl
import pytest

from app.services.pricing.template_writer import detect_rate_column, populate_template


def _make_template(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])  # row 1 header
    # row 2 and row 3 data; Amount has a formula referencing Qty*Rate
    ws.append([1, "Split AC unit", "no", 5, None, "=D2*E2"])
    ws.append([2, "VRF unit", "no", 2, None, "=D3*E3"])
    wb.save(path)
    return str(path)


def test_detect_rate_column():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Description", "Unit", "Qty", "Unit Rate", "Amount"])
    assert detect_rate_column(ws) == 5


def test_populate_writes_rates_and_preserves_formulas(tmp_path):
    src = _make_template(tmp_path / "client.xlsx")
    out = str(tmp_path / "out.xlsx")
    result = populate_template(src, out, {2: 1200.0, 3: 8000.0})
    assert result["written"] == 2
    assert result["rate_column"] == 5
    wb = openpyxl.load_workbook(out)  # data_only=False -> formulas kept as strings
    ws = wb["BOQ"]
    assert ws.cell(row=2, column=5).value == 1200.0
    assert ws.cell(row=3, column=5).value == 8000.0
    # the Amount column formulas must survive untouched
    assert ws.cell(row=2, column=6).value == "=D2*E2"
    assert ws.cell(row=3, column=6).value == "=D3*E3"


def test_populate_explicit_rate_column(tmp_path):
    src = _make_template(tmp_path / "c.xlsx")
    out = str(tmp_path / "o.xlsx")
    result = populate_template(src, out, {2: 99.0}, rate_column=5)
    assert result["written"] == 1
    wb = openpyxl.load_workbook(out)
    assert wb["BOQ"].cell(row=2, column=5).value == 99.0


def test_populate_raises_when_no_rate_column(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Description", "Qty"])  # no rate-like header
    ws.append([1, "x", 3])
    p = tmp_path / "norate.xlsx"
    wb.save(p)
    with pytest.raises(ValueError):
        populate_template(str(p), str(tmp_path / "o.xlsx"), {2: 10.0})


def test_detect_prefers_unit_rate_over_price():
    # A generic "Price" column and a specific "Unit Rate" column coexist; the
    # specific alias must win so we never overwrite a Price total formula.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Item", "Description", "Qty", "Price", "Unit Rate"])
    # Unit Rate is column 5, Price is column 4
    assert detect_rate_column(ws) == 5


def test_populate_skips_formula_cells(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOQ"
    ws.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])
    # row 2's Rate cell already holds a formula — it must NOT be overwritten
    ws.append([1, "Split AC unit", "no", 5, "=B2*C2", "=D2*E2"])
    p = tmp_path / "formula.xlsx"
    wb.save(p)
    out = str(tmp_path / "o.xlsx")
    result = populate_template(str(p), out, {2: 1200.0})
    assert result["skipped_formula"] == 1
    assert result["written"] == 0
    written = openpyxl.load_workbook(out)["BOQ"]
    assert written.cell(row=2, column=5).value == "=B2*C2"  # formula preserved


def _make_cover_and_boq(path):
    wb = openpyxl.Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.append(["Tender", "Acme Project"])
    cover.append(["Date", "2026-06-09"])
    boq = wb.create_sheet("BOQ Pricing")
    boq.append(["Item", "Description", "Unit", "Qty", "Rate", "Amount"])  # row 1
    boq.append([1, "Split AC unit", "no", 5, None, "=D2*E2"])  # row 2
    boq.append([2, "VRF unit", "no", 2, None, "=D3*E3"])  # row 3
    wb.save(path)
    return str(path)


def test_populate_targets_hint_sheet_not_first(tmp_path):
    src = _make_cover_and_boq(tmp_path / "multi.xlsx")
    out = str(tmp_path / "out.xlsx")
    result = populate_template(src, out, {2: 1200.0})
    assert result["written"] == 1
    wb = openpyxl.load_workbook(out)
    boq = wb["BOQ Pricing"]
    # rate landed in the hint sheet at row 2 / detected rate col (5)
    assert boq.cell(row=2, column=result["rate_column"]).value == 1200.0
    assert boq.cell(row=2, column=6).value == "=D2*E2"  # Amount formula survived
    # the Cover sheet is untouched
    cover = wb["Cover"]
    assert cover.cell(row=1, column=1).value == "Tender"
    assert cover.cell(row=1, column=2).value == "Acme Project"


def test_parser_and_writer_pick_same_sheet(tmp_path):
    from app.services.boq.boq_parser import pick_sheet as parser_pick
    from app.services.pricing.template_writer import pick_sheet as writer_pick

    src = _make_cover_and_boq(tmp_path / "multi2.xlsx")
    wb = openpyxl.load_workbook(src)
    assert parser_pick(wb).title == writer_pick(wb).title == "BOQ Pricing"


def test_template_skips_none_row_index(tmp_path):
    src = _make_template(tmp_path / "client.xlsx")
    out = str(tmp_path / "o.xlsx")
    # a None key (no client row mapping) is skipped; only the real row is written
    result = populate_template(src, out, {None: 999.0, 2: 1200.0})
    assert result["written"] == 1
    written = openpyxl.load_workbook(out)["BOQ"]
    assert written.cell(row=2, column=5).value == 1200.0
