import openpyxl
import pytest
from sqlalchemy import select

from app.models.supplier import Supplier
from app.services.supplier.supplier_service import SupplierService


def _make_xlsx(path, rows, headers=("Name", "Email", "Trade", "Phone", "Region")):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(headers))
    for r in rows:
        ws.append(list(r))
    wb.save(path)
    return str(path)


async def test_import_creates_suppliers_and_parses_multivalue(db_session, tmp_path):
    f = _make_xlsx(
        tmp_path / "sup.xlsx",
        [
            ("Acme Steel", "a@x.test; b@x.test", "Structural Steel, MEP", "123", "North"),
            ("Concrete Co", "c@x.test", "Concrete", "456", "South"),
        ],
    )
    svc = SupplierService()
    res = await svc.import_excel(db_session, f)
    assert res["imported"] == 2
    assert res["skipped"] == 0
    acme = (await db_session.execute(
        select(Supplier).where(Supplier.name == "Acme Steel")
    )).scalar_one()
    assert acme.emails == ["a@x.test", "b@x.test"]
    assert acme.trade_categories == ["structural_steel", "mep"]
    assert acme.code == "SUP-0001"


async def test_import_skips_blank_names(db_session, tmp_path):
    f = _make_xlsx(tmp_path / "s.xlsx", [("", "x@x.test", "concrete", "", ""), ("Real", "r@x.test", "mep", "", "")])
    svc = SupplierService()
    res = await svc.import_excel(db_session, f)
    assert res["imported"] == 1
    assert res["skipped"] == 1


async def test_import_update_existing(db_session, tmp_path):
    svc = SupplierService()
    await svc.create(db_session, name="Acme Steel", emails=["old@x.test"], trade_categories=["concrete"])
    f = _make_xlsx(tmp_path / "s.xlsx", [("Acme Steel", "new@x.test", "MEP", "999", "East")])
    res = await svc.import_excel(db_session, f, update_existing=True)
    assert res["updated"] == 1
    assert res["imported"] == 0
    acme = (await db_session.execute(select(Supplier).where(Supplier.name == "Acme Steel"))).scalar_one()
    assert acme.emails == ["new@x.test"]
    assert acme.trade_categories == ["mep"]


async def test_import_missing_name_column_raises(db_session, tmp_path):
    wb = openpyxl.Workbook(); ws = wb.active; ws.append(["Company", "Email"]); ws.append(["X", "x@x.test"])
    p = tmp_path / "bad.xlsx"; wb.save(p)
    svc = SupplierService()
    with pytest.raises(ValueError):
        await svc.import_excel(db_session, str(p))


async def test_export_roundtrips(db_session, tmp_path):
    svc = SupplierService()
    await svc.create(db_session, name="Acme", emails=["a@x.test"], trade_categories=["mep"], region="North", rating=4.0)
    out = await svc.export_excel(db_session, str(tmp_path / "out.xlsx"))
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[0] == "Code" and "Name" in header
    # Column order: Code, Name, Email(s), Trades, Contact, Phone, Region,
    # Country, Rating, Active.
    assert ws.cell(row=2, column=2).value == "Acme"
    assert ws.cell(row=2, column=3).value == "a@x.test"
    assert ws.cell(row=2, column=4).value == "mep"
    assert ws.cell(row=2, column=7).value == "North"
    assert ws.cell(row=2, column=9).value == 4.0
    assert ws.cell(row=2, column=10).value == "yes"
