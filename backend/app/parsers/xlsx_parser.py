"""Excel spreadsheet parser using openpyxl."""

import time
from pathlib import Path
from typing import Optional

from app.parsers.base import BaseParser, ParsedContent, ParserRegistry


@ParserRegistry.register
class XlsxParser(BaseParser):
    """Parser for Microsoft Excel documents (.xlsx, .xls)."""

    supported_extensions = [".xlsx", ".xls"]
    supported_mimetypes = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ]

    async def parse(self, file_path: str) -> ParsedContent:
        """Parse an Excel document.

        Args:
            file_path: Path to Excel file

        Returns:
            ParsedContent with extracted text and metadata
        """
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException

        self.validate_file(file_path)
        start_time = time.time()

        try:
            # Load workbook (data_only=True to get values instead of formulas)
            wb = load_workbook(file_path, data_only=True, read_only=True)

            all_text = []
            tables = []

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                all_text.append(f"\n=== Sheet: {sheet_name} ===\n")

                sheet_data = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        row_values = [
                            str(cell) if cell is not None else ""
                            for cell in row
                        ]
                        sheet_data.append(row_values)
                        all_text.append(" | ".join(row_values))

                if sheet_data:
                    tables.append({
                        "sheet": sheet_name,
                        "data": sheet_data,
                        "rows": len(sheet_data),
                        "cols": len(sheet_data[0]) if sheet_data else 0,
                    })

            wb.close()

            full_text = "\n".join(all_text)

            # Extract metadata
            metadata = await self.extract_metadata(file_path)
            metadata["sheet_count"] = len(wb.sheetnames)
            metadata["sheet_names"] = wb.sheetnames

            processing_time = int((time.time() - start_time) * 1000)

            return ParsedContent(
                text=full_text,
                metadata=metadata,
                tables=tables if tables else None,
                processing_time_ms=processing_time,
            )

        except InvalidFileException:
            raise ValueError(f"Invalid or corrupted Excel file: {file_path}")
        except Exception as e:
            raise Exception(f"Failed to parse Excel: {str(e)}") from e

    async def extract_metadata(self, file_path: str) -> dict:
        """Extract metadata from Excel file.

        Args:
            file_path: Path to Excel file

        Returns:
            Dictionary of metadata
        """
        from openpyxl import load_workbook

        try:
            wb = load_workbook(file_path, read_only=True)
            props = wb.properties

            metadata = {
                "title": props.title or None,
                "creator": props.creator or None,
                "subject": props.subject or None,
                "description": props.description or None,
                "created": props.created.isoformat() if props.created else None,
                "modified": props.modified.isoformat() if props.modified else None,
                "last_modified_by": props.lastModifiedBy or None,
                "file_size": Path(file_path).stat().st_size,
            }

            wb.close()
            return metadata

        except Exception:
            return {
                "file_size": Path(file_path).stat().st_size,
            }

    async def parse_as_dataframe(self, file_path: str, sheet_name: Optional[str] = None):
        """Parse Excel file as pandas DataFrame.

        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet to parse (None for all)

        Returns:
            DataFrame or dict of DataFrames
        """
        import pandas as pd

        self.validate_file(file_path)

        if sheet_name:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            return pd.read_excel(file_path, sheet_name=None)
