"""Package creation and management service."""

import os
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy import select, func, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import get_settings, get_rules
from app.database import get_db_context
from app.models import BOQItem, Document, Package, PackageDocument, Project
from app.models.base import DocumentCategory, DocumentStatus, PackageStatus
from app.services.vector_store import VectorStoreService

settings = get_settings()
rules = get_rules()


class PackagingService:
    """Service for creating and managing procurement packages.

    Handles:
    - Automatic package generation from BOQ
    - Document-to-package linking
    - Package folder creation
    - Package brief generation
    """

    # Trade abbreviations for package codes
    TRADE_ABBREVIATIONS = {
        "CIVIL": "CIV",
        "CONCRETE": "CON",
        "STRUCTURAL_STEEL": "STL",
        "MASONRY": "MAS",
        "WATERPROOFING": "WPF",
        "ROOFING": "ROF",
        "DOORS_WINDOWS": "DW",
        "FINISHES": "FIN",
        "MEP_MECHANICAL": "MEC",
        "MEP_ELECTRICAL": "ELE",
        "MEP_PLUMBING": "PLB",
        "FIRE_PROTECTION": "FP",
        "ELEVATORS": "ELV",
        "LANDSCAPING": "LND",
        "FURNITURE": "FFE",
        "GENERAL": "GEN",
    }

    def __init__(self):
        """Initialize packaging service."""
        self.vector_store = VectorStoreService()

    async def generate_packages_from_boq(
        self,
        project_id: int,
        grouping: str = "trade",
        min_items: int = 5,
        max_items: int = 100,
    ) -> dict:
        """Automatically generate packages from BOQ items.

        Args:
            project_id: Project ID
            grouping: Grouping strategy ("trade" or "section")
            min_items: Minimum items per package
            max_items: Maximum items per package

        Returns:
            Package generation results
        """
        async with get_db_context() as db:
            # Get project
            result = await db.execute(
                select(Project).where(Project.id == project_id)
            )
            project = result.scalar_one_or_none()

            if not project:
                raise ValueError(f"Project not found: {project_id}")

            # Get all BOQ items not yet in packages
            result = await db.execute(
                select(BOQItem).where(
                    BOQItem.project_id == project_id,
                    BOQItem.package_id == None,
                    BOQItem.is_excluded == False,
                )
            )
            items = list(result.scalars().all())

            if not items:
                return {"packages_created": 0, "message": "No unassigned items found"}

            # Group items
            if grouping == "trade":
                groups = self._group_by_trade(items)
            else:
                groups = self._group_by_section(items)

            # Create packages
            packages_created = []
            package_seq = await self._get_next_package_seq(db, project_id)

            for group_name, group_items in groups.items():
                if len(group_items) < min_items:
                    # Merge small groups into "General" or skip
                    continue

                # Split if too large
                while group_items:
                    batch = group_items[:max_items]
                    group_items = group_items[max_items:]

                    # Generate package code
                    trade_abbr = self.TRADE_ABBREVIATIONS.get(group_name, "GEN")
                    project_code = project.code or f"P{project_id}"
                    package_code = f"PKG-{project_code}-{trade_abbr}-{package_seq:03d}"

                    # Create package
                    package = Package(
                        project_id=project_id,
                        name=f"{group_name.replace('_', ' ').title()} Package",
                        code=package_code,
                        trade_category=group_name,
                        description=f"Package for {group_name.lower().replace('_', ' ')} works",
                        status=PackageStatus.DRAFT,
                        total_items=len(batch),
                    )
                    db.add(package)
                    await db.flush()

                    # Assign items to package
                    for item in batch:
                        item.package_id = package.id

                    packages_created.append({
                        "id": package.id,
                        "code": package_code,
                        "name": package.name,
                        "items": len(batch),
                    })

                    package_seq += 1

            await db.commit()

            return {
                "packages_created": len(packages_created),
                "packages": packages_created,
            }

    def _group_by_trade(self, items: list[BOQItem]) -> dict[str, list[BOQItem]]:
        """Group items by trade category."""
        groups = {}
        for item in items:
            trade = item.trade_category or "GENERAL"
            if trade not in groups:
                groups[trade] = []
            groups[trade].append(item)
        return groups

    def _group_by_section(self, items: list[BOQItem]) -> dict[str, list[BOQItem]]:
        """Group items by BOQ section."""
        groups = {}
        for item in items:
            section = item.section or "GENERAL"
            if section not in groups:
                groups[section] = []
            groups[section].append(item)
        return groups

    async def _get_next_package_seq(self, db: AsyncSession, project_id: int) -> int:
        """Get next package sequence number."""
        result = await db.execute(
            select(func.count(Package.id)).where(Package.project_id == project_id)
        )
        count = result.scalar() or 0
        return count + 1

    async def link_documents_to_package(
        self,
        package_id: int,
        auto_link: bool = True,
        document_ids: Optional[list[int]] = None,
    ) -> dict:
        """Link relevant documents to a package.

        Args:
            package_id: Package ID
            auto_link: Automatically find relevant documents
            document_ids: Specific document IDs to link

        Returns:
            Linking results
        """
        async with get_db_context() as db:
            # Get package with items
            result = await db.execute(
                select(Package)
                .options(selectinload(Package.items))
                .where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()

            if not package:
                raise ValueError(f"Package not found: {package_id}")

            linked = []

            if document_ids:
                # Link specific documents
                for doc_id in document_ids:
                    link = PackageDocument(
                        package_id=package_id,
                        document_id=doc_id,
                        include_in_package=True,
                    )
                    db.add(link)
                    linked.append(doc_id)

            elif auto_link:
                # Auto-find relevant documents
                # Build search query from package items
                item_descriptions = [item.description for item in package.items[:10]]
                search_query = f"{package.trade_category} {' '.join(item_descriptions[:5])}"

                # Search for relevant document chunks
                results = await self.vector_store.search(
                    query=search_query,
                    limit=20,
                    filter_conditions={"project_id": package.project_id},
                    min_score=0.5,
                )

                # Group by document
                doc_scores = {}
                for result in results:
                    doc_id = result["metadata"].get("document_id")
                    if doc_id:
                        if doc_id not in doc_scores:
                            doc_scores[doc_id] = []
                        doc_scores[doc_id].append(result["score"])

                # Link documents with good relevance
                for doc_id, scores in doc_scores.items():
                    avg_score = sum(scores) / len(scores)
                    if avg_score >= 0.6:
                        # Check if already linked
                        existing = await db.execute(
                            select(PackageDocument).where(
                                PackageDocument.package_id == package_id,
                                PackageDocument.document_id == doc_id,
                            )
                        )
                        if not existing.scalar_one_or_none():
                            link = PackageDocument(
                                package_id=package_id,
                                document_id=doc_id,
                                relevance_score=avg_score,
                                include_in_package=True,
                            )
                            db.add(link)
                            linked.append(doc_id)

            await db.commit()

            return {
                "package_id": package_id,
                "documents_linked": len(linked),
                "document_ids": linked,
            }

    async def create_package_folder(
        self,
        package_id: int,
        base_path: Optional[str] = None,
    ) -> dict:
        """Create folder structure for a package.

        Args:
            package_id: Package ID
            base_path: Base output path

        Returns:
            Folder creation results
        """
        async with get_db_context() as db:
            # Get package with relationships
            result = await db.execute(
                select(Package)
                .options(
                    selectinload(Package.items),
                    selectinload(Package.linked_documents).selectinload(PackageDocument.document),
                )
                .where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()

            if not package:
                raise ValueError(f"Package not found: {package_id}")

            # Get project
            result = await db.execute(
                select(Project).where(Project.id == package.project_id)
            )
            project = result.scalar_one_or_none()

            # Determine base path
            if base_path:
                folder_base = Path(base_path)
            elif project.folder_path:
                folder_base = Path(project.folder_path) / "Packages"
            else:
                folder_base = settings.PROJECTS_PATH / str(project.id) / "Packages"

            # Create package folder
            package_folder = folder_base / package.code
            package_folder.mkdir(parents=True, exist_ok=True)

            # Create subfolders
            (package_folder / "BOQ").mkdir(exist_ok=True)
            (package_folder / "Specifications").mkdir(exist_ok=True)
            (package_folder / "Drawings").mkdir(exist_ok=True)
            (package_folder / "Offers").mkdir(exist_ok=True)

            # Export BOQ subset
            boq_file = package_folder / "BOQ" / f"{package.code}_BOQ.xlsx"
            await self._export_package_boq(package, str(boq_file))

            # Copy linked documents
            docs_copied = 0
            for link in package.linked_documents:
                if link.include_in_package and link.document:
                    doc = link.document
                    src = Path(doc.file_path)
                    if src.exists():
                        # Determine target folder based on category
                        if doc.category == DocumentCategory.SPECS:
                            target_dir = package_folder / "Specifications"
                        elif doc.category == DocumentCategory.DRAWINGS:
                            target_dir = package_folder / "Drawings"
                        else:
                            target_dir = package_folder

                        target = target_dir / doc.filename
                        shutil.copy2(src, target)
                        docs_copied += 1

            # Update package
            package.folder_path = str(package_folder)
            await db.commit()

            return {
                "package_id": package_id,
                "folder_path": str(package_folder),
                "boq_file": str(boq_file),
                "documents_copied": docs_copied,
            }

    async def _export_package_boq(self, package: Package, output_path: str) -> None:
        """Export BOQ items for a package to Excel."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Bill of Quantities - {package.name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Package Code: {package.code}"
        ws['A2'].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["No.", "Description", "Unit", "Quantity", "Unit Rate", "Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data
        for row_idx, item in enumerate(package.items, 5):
            ws.cell(row=row_idx, column=1, value=item.line_number).border = border
            ws.cell(row=row_idx, column=2, value=item.description).border = border
            ws.cell(row=row_idx, column=3, value=item.unit).border = border
            ws.cell(row=row_idx, column=4, value=item.quantity).border = border
            ws.cell(row=row_idx, column=5, value="").border = border  # For supplier to fill
            ws.cell(row=row_idx, column=6, value="").border = border

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        wb.save(output_path)

    async def generate_package_brief(
        self,
        package_id: int,
        output_path: Optional[str] = None,
    ) -> str:
        """Generate PDF brief for a package.

        Args:
            package_id: Package ID
            output_path: Output file path

        Returns:
            Path to generated PDF
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        async with get_db_context() as db:
            # Get package with relationships
            result = await db.execute(
                select(Package)
                .options(
                    selectinload(Package.items),
                    selectinload(Package.linked_documents).selectinload(PackageDocument.document),
                )
                .where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()

            if not package:
                raise ValueError(f"Package not found: {package_id}")

            # Get project
            result = await db.execute(
                select(Project).where(Project.id == package.project_id)
            )
            project = result.scalar_one_or_none()

        # Determine output path
        if not output_path:
            if package.folder_path:
                output_path = str(Path(package.folder_path) / f"{package.code}_Brief.pdf")
            else:
                output_path = str(settings.TEMP_PATH / f"{package.code}_Brief.pdf")

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # Create PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor("#2F5496"),
        )

        story = []

        # Title
        story.append(Paragraph(f"Package Brief", title_style))
        story.append(Paragraph(f"{package.name}", styles['Heading1']))
        story.append(Spacer(1, 20))

        # Package info table
        info_data = [
            ["Package Code:", package.code],
            ["Trade Category:", package.trade_category.replace("_", " ").title()],
            ["Project:", project.name if project else "N/A"],
            ["Total Items:", str(len(package.items))],
            ["Status:", package.status.value.title()],
        ]

        if package.submission_deadline:
            info_data.append(["Submission Deadline:", package.submission_deadline.strftime("%Y-%m-%d %H:%M")])

        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 30))

        # Description
        if package.description:
            story.append(Paragraph("Scope of Work", heading_style))
            story.append(Paragraph(package.description, styles['Normal']))
            story.append(Spacer(1, 20))

        # BOQ Summary
        story.append(Paragraph("Bill of Quantities Summary", heading_style))

        boq_headers = ["No.", "Description", "Unit", "Qty"]
        boq_data = [boq_headers]

        for item in package.items[:20]:  # First 20 items
            boq_data.append([
                item.line_number,
                item.description[:80] + "..." if len(item.description) > 80 else item.description,
                item.unit,
                f"{item.quantity:,.2f}",
            ])

        if len(package.items) > 20:
            boq_data.append(["...", f"... and {len(package.items) - 20} more items", "", ""])

        boq_table = Table(boq_data, colWidths=[0.6*inch, 4*inch, 0.6*inch, 0.8*inch])
        boq_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(boq_table)
        story.append(Spacer(1, 20))

        # Linked documents
        if package.linked_documents:
            story.append(Paragraph("Reference Documents", heading_style))
            for link in package.linked_documents:
                if link.document:
                    story.append(Paragraph(
                        f"• {link.document.filename}",
                        styles['Normal']
                    ))
            story.append(Spacer(1, 20))

        # Footer
        story.append(Spacer(1, 40))
        story.append(Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        ))

        # Build PDF
        doc.build(story)

        # Update package
        async with get_db_context() as db:
            result = await db.execute(
                select(Package).where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()
            if package:
                package.brief_path = output_path
                await db.commit()

        return output_path

    async def get_package_statistics(self, project_id: int) -> dict:
        """Get packaging statistics for a project.

        Args:
            project_id: Project ID

        Returns:
            Statistics dictionary
        """
        async with get_db_context() as db:
            # Total packages
            total_packages = await db.execute(
                select(func.count(Package.id)).where(Package.project_id == project_id)
            )
            total = total_packages.scalar() or 0

            # By status
            by_status = await db.execute(
                select(Package.status, func.count(Package.id))
                .where(Package.project_id == project_id)
                .group_by(Package.status)
            )
            status_counts = {row[0].value: row[1] for row in by_status}

            # Total BOQ items
            total_items = await db.execute(
                select(func.count(BOQItem.id)).where(BOQItem.project_id == project_id)
            )
            items_total = total_items.scalar() or 0

            # Assigned items
            assigned_items = await db.execute(
                select(func.count(BOQItem.id)).where(
                    BOQItem.project_id == project_id,
                    BOQItem.package_id != None,
                )
            )
            items_assigned = assigned_items.scalar() or 0

            return {
                "total_packages": total,
                "by_status": status_counts,
                "total_boq_items": items_total,
                "assigned_items": items_assigned,
                "unassigned_items": items_total - items_assigned,
                "assignment_rate": (items_assigned / items_total * 100) if items_total > 0 else 0,
            }
