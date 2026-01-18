"""Supplier management service."""

from pathlib import Path
from typing import Optional

from sqlalchemy import func, select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings
from app.database import get_db_context
from app.models import Supplier
from app.models.supplier import SupplierOffer

settings = get_settings()


class SupplierService:
    """Service for supplier management.

    Handles:
    - Supplier CRUD operations
    - Import from Excel/CSV
    - Supplier search and filtering
    - Performance tracking
    """

    async def create_supplier(
        self,
        organization_id: int,
        name: str,
        emails: list[str],
        trade_categories: list[str],
        **kwargs,
    ) -> Supplier:
        """Create a new supplier.

        Args:
            organization_id: Organization ID
            name: Supplier name
            emails: List of email addresses
            trade_categories: List of trade categories
            **kwargs: Additional supplier fields

        Returns:
            Created supplier
        """
        async with get_db_context() as db:
            # Generate code if not provided
            code = kwargs.get("code")
            if not code:
                count = await db.execute(
                    select(func.count(Supplier.id)).where(
                        Supplier.organization_id == organization_id
                    )
                )
                seq = (count.scalar() or 0) + 1
                code = f"SUP-{seq:04d}"

            supplier = Supplier(
                organization_id=organization_id,
                name=name,
                code=code,
                emails=emails,
                trade_categories=trade_categories,
                name_ar=kwargs.get("name_ar"),
                phone=kwargs.get("phone"),
                fax=kwargs.get("fax"),
                address=kwargs.get("address"),
                website=kwargs.get("website"),
                contact_name=kwargs.get("contact_name"),
                contact_email=kwargs.get("contact_email"),
                contact_phone=kwargs.get("contact_phone"),
                region=kwargs.get("region"),
                country=kwargs.get("country"),
                rating=kwargs.get("rating"),
                preferred_language=kwargs.get("preferred_language", "en"),
                notes=kwargs.get("notes"),
                custom_fields=kwargs.get("custom_fields"),
            )
            db.add(supplier)
            await db.commit()
            await db.refresh(supplier)

            return supplier

    async def get_supplier(
        self,
        supplier_id: int,
        organization_id: int,
    ) -> Optional[Supplier]:
        """Get supplier by ID.

        Args:
            supplier_id: Supplier ID
            organization_id: Organization ID for access check

        Returns:
            Supplier or None
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(Supplier).where(
                    Supplier.id == supplier_id,
                    Supplier.organization_id == organization_id,
                )
            )
            return result.scalar_one_or_none()

    async def search_suppliers(
        self,
        organization_id: int,
        query: Optional[str] = None,
        trade_categories: Optional[list[str]] = None,
        region: Optional[str] = None,
        is_active: Optional[bool] = True,
        min_rating: Optional[float] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> tuple[list[Supplier], int]:
        """Search suppliers with filters.

        Args:
            organization_id: Organization ID
            query: Search query (name, code, email)
            trade_categories: Filter by trades
            region: Filter by region
            is_active: Filter by active status
            min_rating: Minimum rating filter
            page: Page number
            page_size: Items per page

        Returns:
            Tuple of (suppliers list, total count)
        """
        async with get_db_context() as db:
            base_query = select(Supplier).where(
                Supplier.organization_id == organization_id
            )

            if query:
                search_term = f"%{query}%"
                base_query = base_query.where(
                    or_(
                        Supplier.name.ilike(search_term),
                        Supplier.code.ilike(search_term),
                        Supplier.contact_name.ilike(search_term),
                    )
                )

            if trade_categories:
                # JSON array contains any of the specified trades
                for trade in trade_categories:
                    base_query = base_query.where(
                        Supplier.trade_categories.contains([trade])
                    )

            if region:
                base_query = base_query.where(Supplier.region == region)

            if is_active is not None:
                base_query = base_query.where(Supplier.is_active == is_active)

            if min_rating:
                base_query = base_query.where(Supplier.rating >= min_rating)

            # Count total
            count_query = select(func.count()).select_from(base_query.subquery())
            total = (await db.execute(count_query)).scalar() or 0

            # Fetch paginated
            base_query = base_query.order_by(Supplier.name)
            base_query = base_query.offset((page - 1) * page_size).limit(page_size)
            result = await db.execute(base_query)
            suppliers = list(result.scalars().all())

            return suppliers, total

    async def get_suppliers_for_trade(
        self,
        organization_id: int,
        trade_category: str,
        limit: int = 50,
    ) -> list[Supplier]:
        """Get active suppliers for a specific trade.

        Args:
            organization_id: Organization ID
            trade_category: Trade category
            limit: Maximum suppliers to return

        Returns:
            List of suppliers
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(Supplier)
                .where(
                    Supplier.organization_id == organization_id,
                    Supplier.is_active == True,
                    Supplier.is_blacklisted == False,
                    Supplier.trade_categories.contains([trade_category]),
                )
                .order_by(Supplier.rating.desc().nullslast())
                .limit(limit)
            )
            return list(result.scalars().all())

    async def import_from_excel(
        self,
        file_path: str,
        organization_id: int,
        update_existing: bool = False,
    ) -> dict:
        """Import suppliers from Excel file.

        Args:
            file_path: Path to Excel file
            organization_id: Organization ID
            update_existing: Update if supplier exists

        Returns:
            Import results
        """
        import pandas as pd

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            raise ValueError(f"Failed to read Excel file: {str(e)}")

        # Normalize column names
        df.columns = [str(c).lower().strip().replace(" ", "_") for c in df.columns]

        # Column mapping
        column_map = {
            "name": ["name", "supplier_name", "company", "vendor"],
            "email": ["email", "emails", "email_address", "e-mail"],
            "trade": ["trade", "trades", "trade_category", "category", "specialization"],
            "phone": ["phone", "telephone", "tel", "mobile"],
            "contact": ["contact", "contact_name", "contact_person", "person"],
            "region": ["region", "area", "location"],
            "country": ["country"],
            "address": ["address", "full_address"],
            "website": ["website", "web", "url"],
        }

        def find_column(df, options):
            for opt in options:
                if opt in df.columns:
                    return opt
            return None

        col_mapping = {key: find_column(df, opts) for key, opts in column_map.items()}

        if not col_mapping["name"]:
            raise ValueError("Required column 'name' not found")

        imported = 0
        updated = 0
        skipped = 0
        errors = []

        async with get_db_context() as db:
            for idx, row in df.iterrows():
                try:
                    name = str(row[col_mapping["name"]]).strip()
                    if not name or name.lower() == "nan":
                        skipped += 1
                        continue

                    # Parse emails
                    emails = []
                    if col_mapping["email"]:
                        email_val = str(row[col_mapping["email"]])
                        if email_val and email_val.lower() != "nan":
                            emails = [e.strip() for e in email_val.replace(";", ",").split(",") if "@" in e]

                    # Parse trades
                    trades = []
                    if col_mapping["trade"]:
                        trade_val = str(row[col_mapping["trade"]])
                        if trade_val and trade_val.lower() != "nan":
                            trades = [t.strip().upper().replace(" ", "_") for t in trade_val.replace(";", ",").split(",")]

                    # Check if exists
                    existing = await db.execute(
                        select(Supplier).where(
                            Supplier.organization_id == organization_id,
                            Supplier.name == name,
                        )
                    )
                    supplier = existing.scalar_one_or_none()

                    if supplier:
                        if update_existing:
                            supplier.emails = emails if emails else supplier.emails
                            supplier.trade_categories = trades if trades else supplier.trade_categories
                            if col_mapping["phone"]:
                                val = row[col_mapping["phone"]]
                                if pd.notna(val):
                                    supplier.phone = str(val)
                            if col_mapping["contact"]:
                                val = row[col_mapping["contact"]]
                                if pd.notna(val):
                                    supplier.contact_name = str(val)
                            if col_mapping["region"]:
                                val = row[col_mapping["region"]]
                                if pd.notna(val):
                                    supplier.region = str(val)
                            if col_mapping["country"]:
                                val = row[col_mapping["country"]]
                                if pd.notna(val):
                                    supplier.country = str(val)
                            updated += 1
                        else:
                            skipped += 1
                        continue

                    # Generate code
                    count = await db.execute(
                        select(func.count(Supplier.id)).where(
                            Supplier.organization_id == organization_id
                        )
                    )
                    seq = (count.scalar() or 0) + imported + 1
                    code = f"SUP-{seq:04d}"

                    # Create supplier
                    supplier = Supplier(
                        organization_id=organization_id,
                        name=name,
                        code=code,
                        emails=emails,
                        trade_categories=trades,
                        phone=str(row[col_mapping["phone"]]) if col_mapping["phone"] and pd.notna(row[col_mapping["phone"]]) else None,
                        contact_name=str(row[col_mapping["contact"]]) if col_mapping["contact"] and pd.notna(row[col_mapping["contact"]]) else None,
                        region=str(row[col_mapping["region"]]) if col_mapping["region"] and pd.notna(row[col_mapping["region"]]) else None,
                        country=str(row[col_mapping["country"]]) if col_mapping["country"] and pd.notna(row[col_mapping["country"]]) else None,
                        address=str(row[col_mapping["address"]]) if col_mapping["address"] and pd.notna(row[col_mapping["address"]]) else None,
                        website=str(row[col_mapping["website"]]) if col_mapping["website"] and pd.notna(row[col_mapping["website"]]) else None,
                    )
                    db.add(supplier)
                    imported += 1

                except Exception as e:
                    errors.append(f"Row {idx + 2}: {str(e)}")

            await db.commit()

        return {
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:10],  # First 10 errors
            "total_errors": len(errors),
        }

    async def update_performance_stats(
        self,
        supplier_id: int,
    ) -> dict:
        """Update supplier performance statistics.

        Args:
            supplier_id: Supplier ID

        Returns:
            Updated statistics
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(Supplier).where(Supplier.id == supplier_id)
            )
            supplier = result.scalar_one_or_none()

            if not supplier:
                raise ValueError(f"Supplier not found: {supplier_id}")

            # Count RFQs (emails sent)
            from app.models import EmailLog
            from app.models.base import EmailType

            rfq_count = await db.execute(
                select(func.count(EmailLog.id)).where(
                    EmailLog.supplier_id == supplier_id,
                    EmailLog.email_type == EmailType.RFQ,
                )
            )
            supplier.total_rfqs_sent = rfq_count.scalar() or 0

            # Count offers
            offer_count = await db.execute(
                select(func.count(SupplierOffer.id)).where(
                    SupplierOffer.supplier_id == supplier_id
                )
            )
            supplier.total_offers_received = offer_count.scalar() or 0

            # Count awards
            from app.models.base import OfferStatus
            award_count = await db.execute(
                select(func.count(SupplierOffer.id)).where(
                    SupplierOffer.supplier_id == supplier_id,
                    SupplierOffer.status == OfferStatus.SELECTED,
                )
            )
            supplier.total_awards = award_count.scalar() or 0

            await db.commit()

            return {
                "total_rfqs_sent": supplier.total_rfqs_sent,
                "total_offers_received": supplier.total_offers_received,
                "total_awards": supplier.total_awards,
                "response_rate": supplier.response_rate,
            }

    async def blacklist_supplier(
        self,
        supplier_id: int,
        organization_id: int,
        reason: str,
    ) -> Supplier:
        """Blacklist a supplier.

        Args:
            supplier_id: Supplier ID
            organization_id: Organization ID
            reason: Blacklist reason

        Returns:
            Updated supplier
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(Supplier).where(
                    Supplier.id == supplier_id,
                    Supplier.organization_id == organization_id,
                )
            )
            supplier = result.scalar_one_or_none()

            if not supplier:
                raise ValueError(f"Supplier not found: {supplier_id}")

            supplier.is_blacklisted = True
            supplier.blacklist_reason = reason
            supplier.is_active = False

            await db.commit()
            await db.refresh(supplier)

            return supplier

    async def export_to_excel(
        self,
        organization_id: int,
        output_path: str,
        trade_categories: Optional[list[str]] = None,
    ) -> str:
        """Export suppliers to Excel.

        Args:
            organization_id: Organization ID
            output_path: Output file path
            trade_categories: Filter by trades

        Returns:
            Path to created file
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        async with get_db_context() as db:
            query = select(Supplier).where(
                Supplier.organization_id == organization_id,
                Supplier.is_active == True,
            )

            if trade_categories:
                for trade in trade_categories:
                    query = query.where(Supplier.trade_categories.contains([trade]))

            result = await db.execute(query.order_by(Supplier.name))
            suppliers = list(result.scalars().all())

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Suppliers"

        # Headers
        headers = ["Code", "Name", "Email(s)", "Trades", "Contact", "Phone", "Region", "Country", "Rating"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        for row_idx, supplier in enumerate(suppliers, 2):
            ws.cell(row=row_idx, column=1, value=supplier.code)
            ws.cell(row=row_idx, column=2, value=supplier.name)
            ws.cell(row=row_idx, column=3, value=", ".join(supplier.emails))
            ws.cell(row=row_idx, column=4, value=", ".join(supplier.trade_categories))
            ws.cell(row=row_idx, column=5, value=supplier.contact_name)
            ws.cell(row=row_idx, column=6, value=supplier.phone)
            ws.cell(row=row_idx, column=7, value=supplier.region)
            ws.cell(row=row_idx, column=8, value=supplier.country)
            ws.cell(row=row_idx, column=9, value=supplier.rating)

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 10

        wb.save(output_path)
        return output_path
