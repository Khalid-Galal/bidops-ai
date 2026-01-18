"""Export and report generation service."""

from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from app.config import get_settings
from app.database import get_db_context
from app.models import BOQItem, Package, Project, Document
from app.models.supplier import Supplier, SupplierOffer
from app.models.base import PackageStatus, OfferStatus

settings = get_settings()


class ExportService:
    """Service for exporting data and generating reports.

    Handles:
    - BOQ export in client format
    - Pricing summary reports
    - Project status reports
    - PDF report generation
    - Excel exports
    """

    async def export_priced_boq(
        self,
        project_id: int,
        output_path: str,
        include_breakdown: bool = True,
        format_style: str = "standard",
    ) -> str:
        """Export priced BOQ to Excel in client format.

        Args:
            project_id: Project ID
            output_path: Output file path
            include_breakdown: Include trade breakdown sheet
            format_style: Format style (standard, detailed, summary)

        Returns:
            Path to created file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        async with get_db_context() as db:
            # Get project
            result = await db.execute(
                select(Project).where(Project.id == project_id)
            )
            project = result.scalar_one_or_none()

            if not project:
                raise ValueError(f"Project not found: {project_id}")

            # Get all BOQ items ordered by section and line
            result = await db.execute(
                select(BOQItem)
                .where(BOQItem.project_id == project_id)
                .order_by(BOQItem.section, BOQItem.line_number)
            )
            items = list(result.scalars().all())

        wb = Workbook()

        # Styles
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        section_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '#,##0.00'

        # Main BOQ Sheet
        ws = wb.active
        ws.title = "Priced BOQ"

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Bill of Quantities - {project.name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Project Code: {project.code or 'N/A'}"
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.merge_cells('A3:G3')
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'].alignment = Alignment(horizontal="center")

        # Headers
        if format_style == "detailed":
            headers = ["No.", "Section", "Description", "Unit", "Quantity", "Unit Rate", "Total", "Trade"]
        else:
            headers = ["No.", "Description", "Unit", "Quantity", "Unit Rate", "Total"]

        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        current_section = None
        row = header_row + 1
        section_totals = {}
        grand_total = 0

        for item in items:
            # Section header
            if format_style == "detailed" and item.section and item.section != current_section:
                current_section = item.section
                ws.merge_cells(f'A{row}:G{row}' if format_style == "detailed" else f'A{row}:F{row}')
                ws.cell(row=row, column=1, value=current_section)
                ws.cell(row=row, column=1).fill = section_fill
                ws.cell(row=row, column=1).font = section_font
                row += 1

            # Data row
            if format_style == "detailed":
                cells = [
                    item.line_number,
                    item.section or "",
                    item.description,
                    item.unit,
                    item.quantity,
                    item.unit_rate or 0,
                    item.total_price or 0,
                    item.trade_category or "",
                ]
            else:
                cells = [
                    item.line_number,
                    item.description,
                    item.unit,
                    item.quantity,
                    item.unit_rate or 0,
                    item.total_price or 0,
                ]

            for col, value in enumerate(cells, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                if col in [5, 6, 7] if format_style == "detailed" else [4, 5, 6]:
                    cell.number_format = currency_format

            # Track totals
            if item.total_price:
                trade = item.trade_category or "GENERAL"
                section_totals[trade] = section_totals.get(trade, 0) + item.total_price
                grand_total += item.total_price

            row += 1

        # Grand total row
        row += 1
        total_col = 7 if format_style == "detailed" else 6
        ws.cell(row=row, column=total_col - 1, value="GRAND TOTAL:").font = total_font
        ws.cell(row=row, column=total_col, value=grand_total).font = total_font
        ws.cell(row=row, column=total_col).number_format = currency_format
        ws.cell(row=row, column=total_col).fill = total_fill

        # Column widths
        if format_style == "detailed":
            widths = [10, 20, 60, 10, 12, 15, 15, 18]
        else:
            widths = [10, 70, 10, 12, 15, 15]

        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Breakdown sheet
        if include_breakdown and section_totals:
            ws_breakdown = wb.create_sheet("Trade Breakdown")

            ws_breakdown['A1'] = "Trade Breakdown Summary"
            ws_breakdown['A1'].font = Font(bold=True, size=14)

            breakdown_headers = ["Trade Category", "Total Value", "Percentage"]
            for col, header in enumerate(breakdown_headers, 1):
                cell = ws_breakdown.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            sorted_trades = sorted(section_totals.items(), key=lambda x: -x[1])
            for row_idx, (trade, total) in enumerate(sorted_trades, 4):
                ws_breakdown.cell(row=row_idx, column=1, value=trade.replace("_", " ").title()).border = border
                ws_breakdown.cell(row=row_idx, column=2, value=total).border = border
                ws_breakdown.cell(row=row_idx, column=2).number_format = currency_format
                pct = (total / grand_total * 100) if grand_total else 0
                ws_breakdown.cell(row=row_idx, column=3, value=f"{pct:.1f}%").border = border

            # Total row
            row_idx += 1
            ws_breakdown.cell(row=row_idx, column=1, value="TOTAL").font = total_font
            ws_breakdown.cell(row=row_idx, column=2, value=grand_total).font = total_font
            ws_breakdown.cell(row=row_idx, column=2).number_format = currency_format
            ws_breakdown.cell(row=row_idx, column=3, value="100%").font = total_font

            ws_breakdown.column_dimensions['A'].width = 25
            ws_breakdown.column_dimensions['B'].width = 20
            ws_breakdown.column_dimensions['C'].width = 15

        wb.save(output_path)
        return output_path

    async def generate_pricing_report_pdf(
        self,
        project_id: int,
        output_path: str,
    ) -> str:
        """Generate pricing summary report as PDF.

        Args:
            project_id: Project ID
            output_path: Output file path

        Returns:
            Path to created file
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            PageBreak, Image
        )
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.piecharts import Pie

        async with get_db_context() as db:
            # Get project
            result = await db.execute(
                select(Project).where(Project.id == project_id)
            )
            project = result.scalar_one_or_none()

            if not project:
                raise ValueError(f"Project not found: {project_id}")

            # Get packages
            result = await db.execute(
                select(Package)
                .options(selectinload(Package.items))
                .where(Package.project_id == project_id)
            )
            packages = list(result.scalars().all())

            # Get BOQ items
            result = await db.execute(
                select(BOQItem).where(BOQItem.project_id == project_id)
            )
            items = list(result.scalars().all())

        # Calculate statistics
        total_value = sum(i.total_price or 0 for i in items)
        priced_items = len([i for i in items if i.total_price])
        by_trade = {}
        for item in items:
            trade = item.trade_category or "GENERAL"
            if trade not in by_trade:
                by_trade[trade] = 0
            by_trade[trade] += item.total_price or 0

        # Create PDF
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=20,
            spaceAfter=30,
            textColor=colors.HexColor("#2F5496"),
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor("#2F5496"),
        )
        normal_style = styles['Normal']

        story = []

        # Title
        story.append(Paragraph("Pricing Summary Report", title_style))
        story.append(Paragraph(f"{project.name}", styles['Heading1']))
        story.append(Spacer(1, 20))

        # Project info
        info_data = [
            ["Project Code:", project.code or "N/A"],
            ["Report Date:", datetime.now().strftime("%Y-%m-%d")],
            ["Total Packages:", str(len(packages))],
            ["Total BOQ Items:", str(len(items))],
            ["Priced Items:", f"{priced_items} ({priced_items/len(items)*100:.0f}%)" if items else "0"],
        ]

        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 30))

        # Total value highlight
        story.append(Paragraph("Total Project Value", heading_style))
        total_table = Table(
            [[f"{total_value:,.2f}"]],
            colWidths=[4*inch]
        )
        total_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 24),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#FFC000")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))
        story.append(total_table)
        story.append(Spacer(1, 30))

        # Trade breakdown
        story.append(Paragraph("Cost Breakdown by Trade", heading_style))

        sorted_trades = sorted(by_trade.items(), key=lambda x: -x[1])
        breakdown_data = [["Trade Category", "Value", "%"]]
        for trade, value in sorted_trades:
            pct = (value / total_value * 100) if total_value else 0
            breakdown_data.append([
                trade.replace("_", " ").title(),
                f"{value:,.2f}",
                f"{pct:.1f}%",
            ])

        breakdown_table = Table(breakdown_data, colWidths=[3*inch, 1.5*inch, 1*inch])
        breakdown_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(breakdown_table)
        story.append(Spacer(1, 30))

        # Package summary
        story.append(PageBreak())
        story.append(Paragraph("Package Summary", heading_style))

        package_data = [["Package", "Trade", "Items", "Value", "Status"]]
        for pkg in packages:
            pkg_value = sum(i.total_price or 0 for i in pkg.items)
            package_data.append([
                pkg.name[:30],
                pkg.trade_category.replace("_", " ").title()[:15],
                str(len(pkg.items)),
                f"{pkg_value:,.2f}",
                pkg.status.value.title(),
            ])

        package_table = Table(package_data, colWidths=[2*inch, 1.2*inch, 0.7*inch, 1.2*inch, 1*inch])
        package_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(package_table)

        # Footer
        story.append(Spacer(1, 50))
        story.append(Paragraph(
            f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M')} by BidOps AI",
            ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.grey)
        ))

        doc.build(story)
        return output_path

    async def export_offer_evaluation_report(
        self,
        package_id: int,
        output_path: str,
    ) -> str:
        """Export offer evaluation report to Excel.

        Args:
            package_id: Package ID
            output_path: Output file path

        Returns:
            Path to created file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        async with get_db_context() as db:
            # Get package
            result = await db.execute(
                select(Package)
                .options(selectinload(Package.items))
                .where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()

            if not package:
                raise ValueError(f"Package not found: {package_id}")

            # Get offers with suppliers
            result = await db.execute(
                select(SupplierOffer)
                .options(selectinload(SupplierOffer.supplier))
                .where(SupplierOffer.package_id == package_id)
                .order_by(SupplierOffer.overall_score.desc().nullslast())
            )
            offers = list(result.scalars().all())

        wb = Workbook()

        # Styles
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        selected_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary Sheet
        ws = wb.active
        ws.title = "Evaluation Summary"

        ws['A1'] = f"Offer Evaluation Report - {package.name}"
        ws['A1'].font = Font(bold=True, size=14)

        ws['A3'] = f"Package Code: {package.code}"
        ws['A4'] = f"Trade: {package.trade_category.replace('_', ' ').title()}"
        ws['A5'] = f"Total Items: {len(package.items)}"
        ws['A6'] = f"Offers Received: {len(offers)}"

        # Comparison table
        headers = ["Rank", "Supplier", "Total Price", "Validity", "Delivery",
                   "Commercial Score", "Technical Score", "Overall Score", "Status"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border

        for row_idx, offer in enumerate(offers, 9):
            cells = [
                offer.rank or "-",
                offer.supplier.name,
                offer.total_price or 0,
                f"{offer.validity_days} days" if offer.validity_days else "-",
                f"{offer.delivery_weeks} weeks" if offer.delivery_weeks else "-",
                round(offer.commercial_score or 0, 1),
                round(offer.technical_score or 0, 1),
                round(offer.overall_score or 0, 1),
                offer.status.value.title(),
            ]

            for col, value in enumerate(cells, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if offer.rank == 1:
                    cell.fill = best_fill
                if offer.status == OfferStatus.SELECTED:
                    cell.fill = selected_fill
                    cell.font = Font(bold=True, color="FFFFFF")

        # Column widths
        widths = [8, 30, 15, 12, 12, 15, 15, 15, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Detail sheets for each offer
        for offer in offers[:5]:  # Top 5 offers
            ws_offer = wb.create_sheet(f"{offer.supplier.name[:20]}")

            ws_offer['A1'] = f"Offer from: {offer.supplier.name}"
            ws_offer['A1'].font = Font(bold=True, size=12)

            details = [
                ["Total Price:", f"{offer.total_price:,.2f} {offer.currency}" if offer.total_price else "N/A"],
                ["Validity:", f"{offer.validity_days} days" if offer.validity_days else "N/A"],
                ["Delivery:", f"{offer.delivery_weeks} weeks" if offer.delivery_weeks else "N/A"],
                ["Payment Terms:", offer.payment_terms or "N/A"],
                ["Commercial Score:", round(offer.commercial_score or 0, 1)],
                ["Technical Score:", round(offer.technical_score or 0, 1)],
                ["Overall Score:", round(offer.overall_score or 0, 1)],
                ["Status:", offer.status.value.title()],
            ]

            for row_idx, (label, value) in enumerate(details, 3):
                ws_offer.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                ws_offer.cell(row=row_idx, column=2, value=value)

            # Exclusions
            if offer.exclusions:
                row = 12
                ws_offer.cell(row=row, column=1, value="Exclusions:").font = Font(bold=True)
                for exc in offer.exclusions:
                    row += 1
                    ws_offer.cell(row=row, column=1, value=f"• {exc}")

            # Deviations
            if offer.deviations:
                row += 2
                ws_offer.cell(row=row, column=1, value="Deviations:").font = Font(bold=True)
                for dev in offer.deviations:
                    row += 1
                    if isinstance(dev, dict):
                        ws_offer.cell(row=row, column=1, value=f"• {dev.get('deviation', dev)}")
                    else:
                        ws_offer.cell(row=row, column=1, value=f"• {dev}")

            ws_offer.column_dimensions['A'].width = 20
            ws_offer.column_dimensions['B'].width = 50

        wb.save(output_path)
        return output_path

    async def generate_project_status_report(
        self,
        project_id: int,
        output_path: str,
    ) -> str:
        """Generate comprehensive project status report.

        Args:
            project_id: Project ID
            output_path: Output file path

        Returns:
            Path to created file
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        async with get_db_context() as db:
            # Get project
            result = await db.execute(
                select(Project).where(Project.id == project_id)
            )
            project = result.scalar_one_or_none()

            if not project:
                raise ValueError(f"Project not found: {project_id}")

            # Get packages
            result = await db.execute(
                select(Package).where(Package.project_id == project_id)
            )
            packages = list(result.scalars().all())

            # Get offers
            result = await db.execute(
                select(SupplierOffer)
                .join(Package)
                .where(Package.project_id == project_id)
            )
            offers = list(result.scalars().all())

            # Get BOQ items
            result = await db.execute(
                select(BOQItem).where(BOQItem.project_id == project_id)
            )
            items = list(result.scalars().all())

        # Calculate statistics
        total_value = sum(i.total_price or 0 for i in items)
        packages_by_status = {}
        for pkg in packages:
            status = pkg.status.value
            packages_by_status[status] = packages_by_status.get(status, 0) + 1

        offers_by_status = {}
        for offer in offers:
            status = offer.status.value
            offers_by_status[status] = offers_by_status.get(status, 0) + 1

        # Create PDF
        doc = SimpleDocTemplate(output_path, pagesize=A4, rightMargin=50, leftMargin=50)
        styles = getSampleStyleSheet()

        story = []

        # Title
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=18, textColor=colors.HexColor("#2F5496"))
        story.append(Paragraph("Project Status Report", title_style))
        story.append(Paragraph(f"{project.name}", styles['Heading1']))
        story.append(Spacer(1, 20))

        # Overview
        story.append(Paragraph("Project Overview", styles['Heading2']))
        overview_data = [
            ["Project Code:", project.code or "N/A"],
            ["Status:", project.status.value.title()],
            ["Total Packages:", str(len(packages))],
            ["Total BOQ Items:", str(len(items))],
            ["Total Offers:", str(len(offers))],
            ["Total Value:", f"{total_value:,.2f}"],
        ]
        overview_table = Table(overview_data, colWidths=[2*inch, 4*inch])
        overview_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(overview_table)
        story.append(Spacer(1, 20))

        # Package Status
        story.append(Paragraph("Package Status", styles['Heading2']))
        pkg_status_data = [["Status", "Count"]]
        for status, count in packages_by_status.items():
            pkg_status_data.append([status.title(), str(count)])
        pkg_table = Table(pkg_status_data, colWidths=[3*inch, 1.5*inch])
        pkg_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        story.append(pkg_table)
        story.append(Spacer(1, 20))

        # Offer Status
        story.append(Paragraph("Offer Status", styles['Heading2']))
        offer_status_data = [["Status", "Count"]]
        for status, count in offers_by_status.items():
            offer_status_data.append([status.replace("_", " ").title(), str(count)])
        offer_table = Table(offer_status_data, colWidths=[3*inch, 1.5*inch])
        offer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        story.append(offer_table)

        # Footer
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", footer_style))

        doc.build(story)
        return output_path

    async def get_dashboard_statistics(
        self,
        project_id: int,
    ) -> dict:
        """Get dashboard statistics for a project.

        Args:
            project_id: Project ID

        Returns:
            Dashboard statistics
        """
        async with get_db_context() as db:
            # Get project
            result = await db.execute(
                select(Project).where(Project.id == project_id)
            )
            project = result.scalar_one_or_none()

            if not project:
                raise ValueError(f"Project not found: {project_id}")

            # Package stats
            pkg_result = await db.execute(
                select(
                    Package.status,
                    func.count(Package.id),
                )
                .where(Package.project_id == project_id)
                .group_by(Package.status)
            )
            packages_by_status = {row[0].value: row[1] for row in pkg_result}

            # BOQ stats
            boq_result = await db.execute(
                select(
                    func.count(BOQItem.id),
                    func.sum(BOQItem.total_price),
                )
                .where(BOQItem.project_id == project_id)
            )
            boq_row = boq_result.first()
            total_items = boq_row[0] or 0
            total_value = float(boq_row[1] or 0)

            priced_result = await db.execute(
                select(func.count(BOQItem.id))
                .where(
                    BOQItem.project_id == project_id,
                    BOQItem.total_price != None,
                )
            )
            priced_items = priced_result.scalar() or 0

            # Offer stats
            offer_result = await db.execute(
                select(
                    SupplierOffer.status,
                    func.count(SupplierOffer.id),
                )
                .join(Package)
                .where(Package.project_id == project_id)
                .group_by(SupplierOffer.status)
            )
            offers_by_status = {row[0].value: row[1] for row in offer_result}

            # Trade breakdown
            trade_result = await db.execute(
                select(
                    BOQItem.trade_category,
                    func.sum(BOQItem.total_price),
                )
                .where(
                    BOQItem.project_id == project_id,
                    BOQItem.total_price != None,
                )
                .group_by(BOQItem.trade_category)
            )
            by_trade = {
                (row[0] or "GENERAL"): float(row[1] or 0)
                for row in trade_result
            }

            return {
                "project_id": project_id,
                "project_name": project.name,
                "project_status": project.status.value,
                "summary": {
                    "total_packages": sum(packages_by_status.values()),
                    "total_items": total_items,
                    "priced_items": priced_items,
                    "pricing_completion": round(priced_items / total_items * 100, 1) if total_items else 0,
                    "total_value": round(total_value, 2),
                    "total_offers": sum(offers_by_status.values()),
                },
                "packages_by_status": packages_by_status,
                "offers_by_status": offers_by_status,
                "value_by_trade": by_trade,
            }
