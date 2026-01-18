"""BOQ (Bill of Quantities) parsing and management service."""

import re
from pathlib import Path
from typing import Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import get_settings, get_rules
from app.database import get_db_context
from app.models import BOQItem, Document, Project
from app.models.base import DocumentCategory, DocumentStatus
from app.services.llm_service import LLMService

settings = get_settings()
rules = get_rules()


class BOQService:
    """Service for BOQ parsing and management.

    Handles:
    - Excel BOQ parsing with multiple format support
    - Item standardization (units, descriptions)
    - Trade category classification
    - BOQ statistics and analysis
    """

    # Common unit mappings for standardization
    UNIT_MAPPINGS = {
        # Area
        "sqm": "m2", "sq.m": "m2", "sq.m.": "m2", "square meter": "m2",
        "square meters": "m2", "sq m": "m2", "m²": "m2",
        # Linear
        "lm": "m", "lin.m": "m", "lin.m.": "m", "linear meter": "m",
        "linear meters": "m", "l.m": "m", "rm": "m", "running meter": "m",
        # Volume
        "cum": "m3", "cu.m": "m3", "cu.m.": "m3", "cubic meter": "m3",
        "cubic meters": "m3", "m³": "m3",
        # Count
        "nr": "no", "nos": "no", "nos.": "no", "number": "no",
        "ea": "no", "each": "no", "pcs": "no", "pieces": "no", "pc": "no",
        # Weight
        "kg": "kg", "kgs": "kg", "kilogram": "kg", "kilograms": "kg",
        "ton": "t", "tons": "t", "tonne": "t", "tonnes": "t", "mt": "t",
        # Lump sum
        "ls": "ls", "lump sum": "ls", "lumpsum": "ls", "l.s": "ls",
        "item": "item", "lot": "lot", "set": "set",
    }

    # Trade category keywords
    TRADE_KEYWORDS = {
        "CIVIL": [
            "excavation", "earthwork", "backfill", "grading", "road",
            "pavement", "curb", "drainage", "manhole", "utility", "site work",
        ],
        "CONCRETE": [
            "concrete", "formwork", "rebar", "reinforcement", "slab",
            "beam", "column", "foundation", "footing", "pile", "retaining",
        ],
        "STRUCTURAL_STEEL": [
            "steel structure", "fabrication", "erection", "steel beam",
            "steel column", "truss", "purlins", "grating", "handrail",
        ],
        "MASONRY": [
            "block", "brick", "masonry", "cmu", "aac", "stone", "wall",
        ],
        "WATERPROOFING": [
            "waterproof", "damp proof", "membrane", "insulation",
            "vapor barrier", "sealant", "joint", "expansion",
        ],
        "ROOFING": [
            "roof", "cladding", "skylight", "gutter", "flashing",
            "metal deck", "sandwich panel",
        ],
        "DOORS_WINDOWS": [
            "door", "window", "glazing", "frame", "hardware", "curtain wall",
            "aluminum", "automatic door", "rolling shutter",
        ],
        "FINISHES": [
            "flooring", "tile", "carpet", "paint", "coating", "ceiling",
            "partition", "gypsum", "plaster", "render", "screed",
        ],
        "MEP_MECHANICAL": [
            "hvac", "air conditioning", "chiller", "ahu", "fcu", "duct",
            "diffuser", "grille", "damper", "vav", "cooling", "heating",
            "ventilation", "exhaust fan", "bms",
        ],
        "MEP_ELECTRICAL": [
            "electrical", "cable", "wire", "panel", "switchgear", "mdb",
            "smdb", "lighting", "fixture", "socket", "switch", "conduit",
            "tray", "busbar", "transformer", "generator", "ups",
        ],
        "MEP_PLUMBING": [
            "plumbing", "pipe", "sanitary", "water supply", "drainage",
            "sewer", "pump", "tank", "valve", "fitting", "fixture",
            "wpipe", "cpvc", "ppr", "hdpe",
        ],
        "FIRE_PROTECTION": [
            "fire", "sprinkler", "alarm", "smoke detector", "extinguisher",
            "hydrant", "hose", "suppression", "fm200", "fire fighting",
        ],
        "ELEVATORS": [
            "elevator", "lift", "escalator", "dumbwaiter", "conveyor",
            "hoist", "traction", "hydraulic",
        ],
        "LANDSCAPING": [
            "landscape", "planting", "irrigation", "lawn", "tree",
            "shrub", "hardscape", "paving", "pergola", "fountain",
        ],
        "FURNITURE": [
            "furniture", "ff&e", "fixture", "equipment", "signage",
            "workstation", "cabinet", "countertop", "millwork",
        ],
    }

    def __init__(self):
        """Initialize BOQ service."""
        self.llm = LLMService()

    async def parse_boq_excel(
        self,
        file_path: str,
        project_id: int,
        sheet_name: Optional[str] = None,
        header_row: int = 0,
        column_mapping: Optional[dict] = None,
    ) -> dict:
        """Parse BOQ from Excel file.

        Args:
            file_path: Path to Excel file
            project_id: Project ID to associate items with
            sheet_name: Specific sheet to parse (None for first/all)
            header_row: Row index containing headers (0-based)
            column_mapping: Custom column name mapping

        Returns:
            Parsing result with statistics
        """
        import pandas as pd

        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Default column mapping
        default_mapping = {
            "line_number": ["item", "no", "no.", "line", "ref", "s.no", "sno"],
            "description": ["description", "desc", "item description", "work item", "particulars"],
            "unit": ["unit", "uom", "u/m"],
            "quantity": ["quantity", "qty", "qnty", "q'ty"],
            "section": ["section", "category", "trade", "division"],
        }

        mapping = {**default_mapping, **(column_mapping or {})}

        try:
            # Read Excel
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
            else:
                # Try to find BOQ sheet
                xl = pd.ExcelFile(file_path)
                boq_sheet = None
                for sheet in xl.sheet_names:
                    if any(kw in sheet.lower() for kw in ["boq", "bill", "quantity", "pricing"]):
                        boq_sheet = sheet
                        break
                boq_sheet = boq_sheet or xl.sheet_names[0]
                df = pd.read_excel(file_path, sheet_name=boq_sheet, header=header_row)

            # Normalize column names
            df.columns = [str(c).lower().strip() for c in df.columns]

            # Map columns
            column_map = {}
            for target, sources in mapping.items():
                for source in sources:
                    if source in df.columns:
                        column_map[source] = target
                        break

            df = df.rename(columns=column_map)

            # Filter rows with valid data
            required_cols = ["description"]
            for col in required_cols:
                if col not in df.columns:
                    raise ValueError(f"Required column '{col}' not found in BOQ")

            # Remove empty rows
            df = df.dropna(subset=["description"])
            df = df[df["description"].str.strip().astype(bool)]

            # Parse items
            items = []
            current_section = None

            async with get_db_context() as db:
                # Clear existing BOQ items for project
                await db.execute(
                    BOQItem.__table__.delete().where(BOQItem.project_id == project_id)
                )

                for idx, row in df.iterrows():
                    # Detect section headers
                    desc = str(row.get("description", "")).strip()

                    # Check if this is a section header (no quantity usually)
                    qty = row.get("quantity")
                    if pd.isna(qty) or qty == "" or qty == 0:
                        # Likely a section header
                        if len(desc) < 100 and not any(c.isdigit() for c in desc[:5]):
                            current_section = desc
                            continue

                    # Parse quantity
                    try:
                        quantity = float(qty) if not pd.isna(qty) else 0.0
                    except (ValueError, TypeError):
                        quantity = 0.0

                    # Standardize unit
                    unit = str(row.get("unit", "")).strip().lower()
                    unit = self.UNIT_MAPPINGS.get(unit, unit) or "no"

                    # Create BOQ item
                    item = BOQItem(
                        project_id=project_id,
                        line_number=str(row.get("line_number", idx + 1)),
                        section=current_section,
                        description=desc,
                        unit=unit,
                        quantity=quantity,
                        client_ref=str(row.get("line_number", "")),
                        client_row_index=idx,
                    )

                    # Classify trade category
                    item.trade_category = self._classify_trade_simple(desc)

                    db.add(item)
                    items.append(item)

                await db.commit()

                # Get statistics
                stats = await self._get_boq_stats(db, project_id)

            return {
                "success": True,
                "file": path.name,
                "items_parsed": len(items),
                "sections_found": len(set(i.section for i in items if i.section)),
                "statistics": stats,
            }

        except Exception as e:
            raise ValueError(f"Failed to parse BOQ: {str(e)}")

    def _classify_trade_simple(self, description: str) -> str:
        """Simple keyword-based trade classification.

        Args:
            description: Item description

        Returns:
            Trade category
        """
        desc_lower = description.lower()

        for trade, keywords in self.TRADE_KEYWORDS.items():
            if any(kw in desc_lower for kw in keywords):
                return trade

        return "GENERAL"

    async def classify_items_with_ai(
        self,
        project_id: int,
        batch_size: int = 20,
    ) -> dict:
        """Classify BOQ items using AI for better accuracy.

        Args:
            project_id: Project ID
            batch_size: Items to process per batch

        Returns:
            Classification results
        """
        from app.prompts.classification import BOQ_CLASSIFICATION_PROMPT

        async with get_db_context() as db:
            # Get items without high-confidence classification
            result = await db.execute(
                select(BOQItem).where(
                    BOQItem.project_id == project_id,
                    (BOQItem.classification_confidence == None) |
                    (BOQItem.classification_confidence < 0.8),
                )
            )
            items = list(result.scalars().all())

            classified = 0
            failed = 0

            # Process in batches
            for i in range(0, len(items), batch_size):
                batch = items[i:i + batch_size]

                for item in batch:
                    try:
                        prompt = BOQ_CLASSIFICATION_PROMPT.format(
                            line_number=item.line_number,
                            description=item.description,
                            unit=item.unit,
                            section=item.section or "N/A",
                        )

                        response = await self.llm.generate(
                            prompt=prompt,
                            task_type="classification",
                            json_mode=True,
                        )

                        import json
                        result = json.loads(response)

                        item.trade_category = result.get("trade_category", "GENERAL")
                        item.trade_subcategory = result.get("trade_subcategory")
                        item.classification_confidence = float(result.get("confidence", 0.5))

                        classified += 1

                    except Exception:
                        failed += 1

                await db.commit()

            return {
                "total_items": len(items),
                "classified": classified,
                "failed": failed,
            }

    async def _get_boq_stats(self, db: AsyncSession, project_id: int) -> dict:
        """Get BOQ statistics.

        Args:
            db: Database session
            project_id: Project ID

        Returns:
            Statistics dictionary
        """
        # Total items
        total = await db.execute(
            select(func.count(BOQItem.id)).where(BOQItem.project_id == project_id)
        )
        total_count = total.scalar() or 0

        # By trade category
        by_trade = await db.execute(
            select(BOQItem.trade_category, func.count(BOQItem.id))
            .where(BOQItem.project_id == project_id)
            .group_by(BOQItem.trade_category)
        )
        trade_counts = {row[0] or "UNCLASSIFIED": row[1] for row in by_trade}

        # By section
        by_section = await db.execute(
            select(BOQItem.section, func.count(BOQItem.id))
            .where(BOQItem.project_id == project_id)
            .group_by(BOQItem.section)
        )
        section_counts = {row[0] or "NO SECTION": row[1] for row in by_section}

        return {
            "total_items": total_count,
            "by_trade": trade_counts,
            "by_section": section_counts,
        }

    async def get_items_by_trade(
        self,
        project_id: int,
        trade_category: str,
    ) -> list[BOQItem]:
        """Get BOQ items for a specific trade.

        Args:
            project_id: Project ID
            trade_category: Trade category to filter

        Returns:
            List of BOQ items
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(BOQItem).where(
                    BOQItem.project_id == project_id,
                    BOQItem.trade_category == trade_category,
                ).order_by(BOQItem.line_number)
            )
            return list(result.scalars().all())

    async def get_unassigned_items(self, project_id: int) -> list[BOQItem]:
        """Get BOQ items not assigned to any package.

        Args:
            project_id: Project ID

        Returns:
            List of unassigned items
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(BOQItem).where(
                    BOQItem.project_id == project_id,
                    BOQItem.package_id == None,
                ).order_by(BOQItem.trade_category, BOQItem.line_number)
            )
            return list(result.scalars().all())

    def standardize_unit(self, unit: str) -> str:
        """Standardize a unit of measurement.

        Args:
            unit: Raw unit string

        Returns:
            Standardized unit
        """
        unit_lower = unit.lower().strip()
        return self.UNIT_MAPPINGS.get(unit_lower, unit_lower)

    async def export_boq_excel(
        self,
        project_id: int,
        output_path: str,
        include_pricing: bool = False,
    ) -> str:
        """Export BOQ to Excel file.

        Args:
            project_id: Project ID
            output_path: Output file path
            include_pricing: Include pricing columns

        Returns:
            Path to created file
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        async with get_db_context() as db:
            result = await db.execute(
                select(BOQItem).where(
                    BOQItem.project_id == project_id
                ).order_by(BOQItem.section, BOQItem.line_number)
            )
            items = list(result.scalars().all())

        if not items:
            raise ValueError("No BOQ items found for project")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "BOQ"

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Headers
        headers = ["No.", "Section", "Description", "Unit", "Quantity", "Trade"]
        if include_pricing:
            headers.extend(["Unit Rate", "Total"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data rows
        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.line_number).border = border
            ws.cell(row=row_idx, column=2, value=item.section or "").border = border
            ws.cell(row=row_idx, column=3, value=item.description).border = border
            ws.cell(row=row_idx, column=4, value=item.unit).border = border
            ws.cell(row=row_idx, column=5, value=item.quantity).border = border
            ws.cell(row=row_idx, column=6, value=item.trade_category).border = border

            if include_pricing:
                ws.cell(row=row_idx, column=7, value=item.unit_rate or 0).border = border
                ws.cell(row=row_idx, column=8, value=item.total_price or 0).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 18

        # Save
        wb.save(output_path)

        return output_path
