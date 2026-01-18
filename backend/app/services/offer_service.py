"""Offer evaluation and comparison service."""

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import get_settings, get_rules
from app.database import get_db_context
from app.models import Package, Supplier, Document, BOQItem
from app.models.supplier import SupplierOffer
from app.models.base import OfferStatus, DocumentStatus
from app.services.document_service import DocumentService
from app.services.llm_service import LLMService
from app.prompts.classification import OFFER_EXTRACTION_PROMPT, COMPLIANCE_CHECK_PROMPT

settings = get_settings()
rules = get_rules()


class OfferService:
    """Service for offer evaluation and comparison.

    Handles:
    - Offer file upload and parsing
    - AI-powered data extraction
    - Compliance checking
    - Offer comparison and ranking
    - Commercial analysis
    """

    def __init__(self):
        """Initialize offer service."""
        self.llm = LLMService()
        self.document_service = DocumentService()

    async def create_offer(
        self,
        package_id: int,
        supplier_id: int,
        file_paths: list[str],
    ) -> SupplierOffer:
        """Create a new supplier offer.

        Args:
            package_id: Package ID
            supplier_id: Supplier ID
            file_paths: List of offer file paths

        Returns:
            Created offer
        """
        async with get_db_context() as db:
            # Verify package exists
            result = await db.execute(
                select(Package).where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()
            if not package:
                raise ValueError(f"Package not found: {package_id}")

            # Verify supplier exists
            result = await db.execute(
                select(Supplier).where(Supplier.id == supplier_id)
            )
            supplier = result.scalar_one_or_none()
            if not supplier:
                raise ValueError(f"Supplier not found: {supplier_id}")

            # Validate files exist
            valid_paths = []
            for path in file_paths:
                if Path(path).exists():
                    valid_paths.append(path)
                else:
                    raise ValueError(f"File not found: {path}")

            # Create offer
            offer = SupplierOffer(
                package_id=package_id,
                supplier_id=supplier_id,
                file_paths=valid_paths,
                status=OfferStatus.RECEIVED,
                received_at=datetime.now(timezone.utc),
            )
            db.add(offer)

            # Update package stats
            package.offers_received += 1

            # Update supplier stats
            supplier.total_offers_received += 1

            await db.commit()
            await db.refresh(offer)

            return offer

    async def extract_offer_data(
        self,
        offer_id: int,
    ) -> dict:
        """Extract commercial data from offer documents using AI.

        Args:
            offer_id: Offer ID

        Returns:
            Extracted data
        """
        async with get_db_context() as db:
            # Get offer with relationships
            result = await db.execute(
                select(SupplierOffer)
                .options(
                    selectinload(SupplierOffer.package).selectinload(Package.items),
                    selectinload(SupplierOffer.supplier),
                )
                .where(SupplierOffer.id == offer_id)
            )
            offer = result.scalar_one_or_none()

            if not offer:
                raise ValueError(f"Offer not found: {offer_id}")

            # Parse offer files to extract text
            all_content = []
            for file_path in offer.file_paths:
                try:
                    text = await self.document_service.parse_document(file_path)
                    all_content.append(text[:10000])  # Limit per file
                except Exception:
                    pass

            combined_content = "\n\n---\n\n".join(all_content)[:30000]

            # Prepare BOQ items summary
            required_items = []
            for item in offer.package.items[:20]:
                required_items.append(f"- {item.line_number}: {item.description[:100]} ({item.unit})")

            # Call AI for extraction
            prompt = OFFER_EXTRACTION_PROMPT.format(
                content=combined_content,
                package_name=offer.package.name,
                required_items="\n".join(required_items),
            )

            response = await self.llm.generate(
                prompt=prompt,
                task_type="offer_extraction",
                json_mode=True,
            )

            try:
                extracted = json.loads(response)
            except json.JSONDecodeError:
                extracted = {}

            # Update offer with extracted data
            if "total_price" in extracted:
                offer.total_price = extracted["total_price"].get("amount")
                offer.currency = extracted["total_price"].get("currency")
                offer.vat_included = extracted["total_price"].get("includes_vat")

            offer.validity_days = extracted.get("validity_days")
            offer.payment_terms = extracted.get("payment_terms")
            offer.delivery_weeks = extracted.get("delivery_weeks")
            offer.exclusions = extracted.get("exclusions", [])
            offer.deviations = extracted.get("deviations", [])
            offer.line_items = extracted.get("line_items", [])

            offer.status = OfferStatus.UNDER_REVIEW

            await db.commit()

            return extracted

    async def check_compliance(
        self,
        offer_id: int,
    ) -> dict:
        """Check offer compliance against requirements.

        Args:
            offer_id: Offer ID

        Returns:
            Compliance analysis
        """
        async with get_db_context() as db:
            # Get offer
            result = await db.execute(
                select(SupplierOffer)
                .options(
                    selectinload(SupplierOffer.package).selectinload(Package.items),
                )
                .where(SupplierOffer.id == offer_id)
            )
            offer = result.scalar_one_or_none()

            if not offer:
                raise ValueError(f"Offer not found: {offer_id}")

            # Get project requirements (checklist)
            from app.models import Project
            result = await db.execute(
                select(Project).where(Project.id == offer.package.project_id)
            )
            project = result.scalar_one_or_none()

            requirements = []
            if project and project.checklist:
                for category in project.checklist:
                    for req in category.get("requirements", []):
                        requirements.append(f"- {req.get('requirement', '')}")

            # Add package-specific requirements
            for item in offer.package.items[:30]:
                requirements.append(f"- BOQ Item: {item.description[:100]}")

            requirements_text = "\n".join(requirements[:50])

            # Build offer content summary
            offer_content = f"""
Total Price: {offer.total_price} {offer.currency or ''}
Payment Terms: {offer.payment_terms or 'Not specified'}
Delivery: {offer.delivery_weeks or 'Not specified'} weeks
Validity: {offer.validity_days or 'Not specified'} days

Exclusions:
{chr(10).join(['- ' + str(e) for e in (offer.exclusions or [])])}

Deviations:
{chr(10).join(['- ' + str(d.get('deviation', d)) for d in (offer.deviations or [])])}

Line Items Provided: {len(offer.line_items or [])}
"""

            # Call AI for compliance check
            prompt = COMPLIANCE_CHECK_PROMPT.format(
                requirements=requirements_text,
                offer_content=offer_content,
            )

            response = await self.llm.generate(
                prompt=prompt,
                task_type="compliance_check",
                json_mode=True,
            )

            try:
                compliance = json.loads(response)
            except json.JSONDecodeError:
                compliance = {
                    "overall_compliance": "UNKNOWN",
                    "compliance_score": 0,
                }

            # Update offer
            offer.compliance_analysis = compliance
            offer.clarifications_needed = compliance.get("clarifications_needed", [])

            # Set status based on compliance
            if compliance.get("overall_compliance") == "COMPLIANT":
                offer.status = OfferStatus.COMPLIANT
            elif compliance.get("overall_compliance") == "NON_COMPLIANT":
                offer.status = OfferStatus.NON_COMPLIANT
            else:
                offer.status = OfferStatus.UNDER_REVIEW

            await db.commit()

            return compliance

    async def evaluate_offer(
        self,
        offer_id: int,
        technical_score: Optional[float] = None,
        commercial_weight: float = 0.6,
        technical_weight: float = 0.4,
    ) -> dict:
        """Evaluate and score an offer.

        Args:
            offer_id: Offer ID
            technical_score: Manual technical score (0-100)
            commercial_weight: Weight for commercial score
            technical_weight: Weight for technical score

        Returns:
            Evaluation results
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(SupplierOffer)
                .options(selectinload(SupplierOffer.package))
                .where(SupplierOffer.id == offer_id)
            )
            offer = result.scalar_one_or_none()

            if not offer:
                raise ValueError(f"Offer not found: {offer_id}")

            # Calculate commercial score based on price ranking
            # Get all offers for the package
            all_offers = await db.execute(
                select(SupplierOffer).where(
                    SupplierOffer.package_id == offer.package_id,
                    SupplierOffer.total_price != None,
                    SupplierOffer.total_price > 0,
                )
            )
            offers_list = list(all_offers.scalars().all())

            # Find min price for normalization
            prices = [o.total_price for o in offers_list if o.total_price]
            if prices:
                min_price = min(prices)
                if offer.total_price:
                    # Score = 100 * (min_price / offer_price)
                    commercial_score = 100 * (min_price / offer.total_price)
                    commercial_score = min(100, max(0, commercial_score))
                else:
                    commercial_score = 0
            else:
                commercial_score = 50  # Default if no prices

            offer.commercial_score = commercial_score

            # Use provided technical score or compliance-based score
            if technical_score is not None:
                offer.technical_score = technical_score
            elif offer.compliance_analysis:
                # Use compliance score as technical score
                offer.technical_score = offer.compliance_analysis.get("compliance_score", 50)
            else:
                offer.technical_score = 50

            # Calculate overall weighted score
            offer.overall_score = (
                (offer.commercial_score * commercial_weight) +
                (offer.technical_score * technical_weight)
            )

            offer.evaluated_at = datetime.now(timezone.utc)
            offer.status = OfferStatus.EVALUATED

            # Update package stats
            offer.package.offers_evaluated += 1

            await db.commit()

            return {
                "offer_id": offer_id,
                "commercial_score": offer.commercial_score,
                "technical_score": offer.technical_score,
                "overall_score": offer.overall_score,
            }

    async def rank_offers(
        self,
        package_id: int,
    ) -> list[dict]:
        """Rank all offers for a package.

        Args:
            package_id: Package ID

        Returns:
            Ranked list of offers
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(SupplierOffer)
                .options(selectinload(SupplierOffer.supplier))
                .where(
                    SupplierOffer.package_id == package_id,
                    SupplierOffer.overall_score != None,
                )
                .order_by(SupplierOffer.overall_score.desc())
            )
            offers = list(result.scalars().all())

            ranked = []
            for rank, offer in enumerate(offers, 1):
                offer.rank = rank
                ranked.append({
                    "rank": rank,
                    "offer_id": offer.id,
                    "supplier_name": offer.supplier.name,
                    "total_price": offer.total_price,
                    "currency": offer.currency,
                    "commercial_score": offer.commercial_score,
                    "technical_score": offer.technical_score,
                    "overall_score": offer.overall_score,
                    "status": offer.status.value,
                })

            await db.commit()

            return ranked

    async def compare_offers(
        self,
        package_id: int,
    ) -> dict:
        """Generate offer comparison summary.

        Args:
            package_id: Package ID

        Returns:
            Comparison data
        """
        async with get_db_context() as db:
            # Get package
            result = await db.execute(
                select(Package)
                .options(selectinload(Package.items))
                .where(Package.id == package_id)
            )
            package = result.scalar_one_or_none()

            if not package:
                raise ValueError(f"Package not found: {package_id}")

            # Get all offers
            result = await db.execute(
                select(SupplierOffer)
                .options(selectinload(SupplierOffer.supplier))
                .where(SupplierOffer.package_id == package_id)
                .order_by(SupplierOffer.overall_score.desc().nullslast())
            )
            offers = list(result.scalars().all())

            if not offers:
                return {"message": "No offers found for this package"}

            # Calculate statistics
            prices = [o.total_price for o in offers if o.total_price]

            comparison = {
                "package_id": package_id,
                "package_name": package.name,
                "total_boq_items": len(package.items),
                "total_offers": len(offers),
                "evaluated_offers": len([o for o in offers if o.overall_score]),
                "price_statistics": {
                    "min": min(prices) if prices else None,
                    "max": max(prices) if prices else None,
                    "average": sum(prices) / len(prices) if prices else None,
                    "currency": offers[0].currency if offers and offers[0].currency else "USD",
                },
                "offers": [],
            }

            for offer in offers:
                comparison["offers"].append({
                    "offer_id": offer.id,
                    "supplier_id": offer.supplier.id,
                    "supplier_name": offer.supplier.name,
                    "total_price": offer.total_price,
                    "currency": offer.currency,
                    "validity_days": offer.validity_days,
                    "delivery_weeks": offer.delivery_weeks,
                    "payment_terms": offer.payment_terms,
                    "commercial_score": offer.commercial_score,
                    "technical_score": offer.technical_score,
                    "overall_score": offer.overall_score,
                    "rank": offer.rank,
                    "status": offer.status.value,
                    "exclusions_count": len(offer.exclusions or []),
                    "deviations_count": len(offer.deviations or []),
                    "line_items_count": len(offer.line_items or []),
                })

            return comparison

    async def select_offer(
        self,
        offer_id: int,
        notes: Optional[str] = None,
    ) -> SupplierOffer:
        """Select an offer as the winner.

        Args:
            offer_id: Offer ID
            notes: Selection notes

        Returns:
            Selected offer
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(SupplierOffer)
                .options(
                    selectinload(SupplierOffer.package),
                    selectinload(SupplierOffer.supplier),
                )
                .where(SupplierOffer.id == offer_id)
            )
            offer = result.scalar_one_or_none()

            if not offer:
                raise ValueError(f"Offer not found: {offer_id}")

            # Unselect any previously selected offer for this package
            await db.execute(
                SupplierOffer.__table__.update()
                .where(
                    SupplierOffer.package_id == offer.package_id,
                    SupplierOffer.status == OfferStatus.SELECTED,
                )
                .values(status=OfferStatus.EVALUATED)
            )

            # Select this offer
            offer.status = OfferStatus.SELECTED
            offer.recommendation = notes

            # Update supplier stats
            offer.supplier.total_awards += 1

            await db.commit()
            await db.refresh(offer)

            return offer

    async def export_comparison_excel(
        self,
        package_id: int,
        output_path: str,
    ) -> str:
        """Export offer comparison to Excel.

        Args:
            package_id: Package ID
            output_path: Output file path

        Returns:
            Path to created file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        comparison = await self.compare_offers(package_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Offer Comparison"

        # Styles
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = f"Offer Comparison - {comparison['package_name']}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Summary
        ws['A3'] = "Total Offers:"
        ws['B3'] = comparison['total_offers']
        ws['C3'] = "Lowest Price:"
        ws['D3'] = comparison['price_statistics']['min']
        ws['E3'] = comparison['price_statistics']['currency']

        # Headers
        headers = [
            "Rank", "Supplier", "Total Price", "Currency", "Validity (days)",
            "Delivery (weeks)", "Commercial Score", "Technical Score",
            "Overall Score", "Status"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, offer in enumerate(comparison['offers'], 6):
            cells = [
                offer['rank'] or '-',
                offer['supplier_name'],
                offer['total_price'] or 0,
                offer['currency'] or '',
                offer['validity_days'] or '-',
                offer['delivery_weeks'] or '-',
                round(offer['commercial_score'] or 0, 1),
                round(offer['technical_score'] or 0, 1),
                round(offer['overall_score'] or 0, 1),
                offer['status'],
            ]

            for col, value in enumerate(cells, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if offer['rank'] == 1:
                    cell.fill = best_fill

        # Column widths
        widths = [8, 30, 15, 10, 15, 15, 15, 15, 15, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(output_path)
        return output_path

    async def get_offer_details(
        self,
        offer_id: int,
    ) -> dict:
        """Get detailed offer information.

        Args:
            offer_id: Offer ID

        Returns:
            Offer details
        """
        async with get_db_context() as db:
            result = await db.execute(
                select(SupplierOffer)
                .options(
                    selectinload(SupplierOffer.package),
                    selectinload(SupplierOffer.supplier),
                )
                .where(SupplierOffer.id == offer_id)
            )
            offer = result.scalar_one_or_none()

            if not offer:
                raise ValueError(f"Offer not found: {offer_id}")

            return {
                "id": offer.id,
                "package": {
                    "id": offer.package.id,
                    "name": offer.package.name,
                    "code": offer.package.code,
                },
                "supplier": {
                    "id": offer.supplier.id,
                    "name": offer.supplier.name,
                    "code": offer.supplier.code,
                },
                "status": offer.status.value,
                "received_at": offer.received_at.isoformat() if offer.received_at else None,
                "evaluated_at": offer.evaluated_at.isoformat() if offer.evaluated_at else None,
                "commercial": {
                    "total_price": offer.total_price,
                    "currency": offer.currency,
                    "vat_included": offer.vat_included,
                    "vat_amount": offer.vat_amount,
                    "validity_days": offer.validity_days,
                    "payment_terms": offer.payment_terms,
                    "delivery_weeks": offer.delivery_weeks,
                    "delivery_terms": offer.delivery_terms,
                },
                "compliance": {
                    "exclusions": offer.exclusions,
                    "deviations": offer.deviations,
                    "missing_items": offer.missing_items,
                    "clarifications_needed": offer.clarifications_needed,
                    "analysis": offer.compliance_analysis,
                },
                "scoring": {
                    "commercial_score": offer.commercial_score,
                    "technical_score": offer.technical_score,
                    "overall_score": offer.overall_score,
                    "rank": offer.rank,
                },
                "line_items": offer.line_items,
                "files": offer.file_paths,
                "notes": offer.evaluator_notes,
                "recommendation": offer.recommendation,
            }
